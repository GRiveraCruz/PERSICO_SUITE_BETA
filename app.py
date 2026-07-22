"""
Persico Mex — Suite Unificada
===============================
Combina:  Job Register         (orig. puerto 5001)
          Hourly Rate Register  (orig. puerto 5002)
          GERC Quote Register   (orig. puerto 5000)
          Purchase Orders       (nuevo módulo)

Ejecutar:  python app.py
Acceso:    http://<IP-del-servidor>:5000
"""

import io, json, re, datetime, shutil, hashlib
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory, Response, session, redirect, url_for, send_file
from threading import Lock
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════
#  CONFIG  — rutas locales / volumen persistente (Railway)
# ══════════════════════════════════════════════════════════════════
import os as _os
_BASE = _os.path.dirname(_os.path.abspath(__file__))

# En Railway: setear variable DATA_DIR apuntando al mount path del volumen.
# Localmente: usa la carpeta data/ junto al .py (Windows/Mac/Linux).
_DATA = _os.environ.get("DATA_DIR", _os.path.join(_BASE, "data"))

# Seed automático: si el volumen está vacío, copiamos los datos iniciales
_SEED = _os.path.join(_BASE, "data_seed")
if _os.path.isdir(_SEED) and not _os.path.exists(_os.path.join(_DATA, "JOBs")):
    import shutil as _shutil
    print(f"[INIT] Volumen vacío — copiando datos iniciales: data_seed/ → {_DATA}")
    _shutil.copytree(_SEED, _DATA, dirs_exist_ok=True)
    print("[INIT] Seed OK")

JOBS_FOLDER  = _os.path.join(_DATA, "JOBs")
RATES_FOLDER = _os.path.join(_DATA, "HOUR_RATE")
XLSM_PATH    = _os.path.join(_DATA, "QUOTE_REG", "quotes.json")   # migrado a JSON
QUOTE_BASE   = _os.path.join(_DATA, "QUOTE_REG")
PO_FOLDER    = _os.path.join(_DATA, "IPOs")
FX_FOLDER    = _os.path.join(_DATA, "FX")

HOST         = "0.0.0.0"
PORT         = int(_os.environ.get("PORT", 5000))
CURRENT_YEAR = datetime.date.today().year

# Quote Register constants
QUOTE_DATA_ROW = 4
QUOTE_MAX_ROWS = 200
# ══════════════════════════════════════════════════════════════════

app  = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = _os.environ.get("SECRET_KEY", "persico-suite-secret-2026")
lock = Lock()
JOB_RE = re.compile(r"^\d+-\d+$")

# ══════════════════════════════════════════════════════════════════
#  AUTH — usuarios con contraseña
# ══════════════════════════════════════════════════════════════════
def _hash(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def _auth_file():
    p = Path(_DATA) / "users_auth.json"
    return p

def _ensure_auth_file():
    """Migra usuarios de env vars a users_auth.json si no existe."""
    p = _auth_file()
    if p.exists():
        return
    users = {}
    for i in range(1, 10):
        val = _os.environ.get(f"USER{i}", "")
        if ":" in val:
            u, pw = val.split(":", 1)
            u = u.strip()
            users[u] = {
                "password_hash": _hash(pw.strip()),
                "active": True,
                "created_at": datetime.datetime.now().isoformat()
            }
    # Defaults si no hay vars de entorno
    if not users:
        for u in ["guillermo","luz","pablo","omar"]:
            users[u] = {
                "password_hash": _hash("Persico2026!"),
                "active": True,
                "created_at": datetime.datetime.now().isoformat()
            }
    Path(_DATA).mkdir(parents=True, exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def _load_auth():
    _ensure_auth_file()
    with open(_auth_file(), "r", encoding="utf-8") as f:
        return json.load(f)

def _save_auth(auth):
    with open(_auth_file(), "w", encoding="utf-8") as f:
        json.dump(auth, f, ensure_ascii=False, indent=2)

def _check_login(username, password):
    try:
        auth = _load_auth()
        u = auth.get(username)
        if not u or not u.get("active", True): return False
        return u.get("password_hash") == _hash(password)
    except: return False

def _login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "no autenticado"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════════
#  IN-MEMORY CACHE — evita leer disco en cada request
# ══════════════════════════════════════════════════════════════════
import threading as _thr_cache

_cache      = {}
_cache_lock = _thr_cache.Lock()

def _cache_get(key):
    with _cache_lock:
        return _cache.get(key)

def _cache_set(key, data):
    with _cache_lock:
        _cache[key] = data

def _cache_del(key):
    with _cache_lock:
        _cache.pop(key, None)

@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "").strip()
        if _check_login(u, p):
            session["user"] = u
            return redirect("/")
        error = "Usuario o contraseña incorrectos"
    return f'''<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Persico Suite — Login</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: Arial, sans-serif; background: #1a1a2e; display: flex;
            justify-content: center; align-items: center; min-height: 100vh; }}
    .card {{ background: #16213e; border-radius: 12px; padding: 36px 40px 40px;
             width: 380px; box-shadow: 0 8px 32px rgba(0,0,0,0.4); }}
    .logo-wrap {{ text-align: center; margin-bottom: 6px; }}
    .logo-wrap img {{ height: 52px; object-fit: contain; }}
    p.sub {{ color: #718096; text-align: center; margin-bottom: 28px; font-size: 12px;
             letter-spacing: .5px; text-transform: uppercase; margin-top: 6px; }}
    .divider {{ border: none; border-top: 1px solid #2d3748; margin-bottom: 24px; }}
    label {{ color: #a0aec0; font-size: 13px; display: block; margin-bottom: 6px; }}
    input {{ width: 100%; padding: 10px 14px; border-radius: 8px;
             border: 1px solid #2d3748; background: #0f3460; color: #e2e8f0;
             font-size: 14px; margin-bottom: 18px; outline: none; }}
    input:focus {{ border-color: #c8102e; }}
    button {{ width: 100%; padding: 12px; background: #c8102e; color: white;
              border: none; border-radius: 8px; font-size: 15px;
              cursor: pointer; font-weight: bold; letter-spacing: .5px; }}
    button:hover {{ background: #a00d24; }}
    .error {{ background: #742a2a; color: #feb2b2; padding: 10px 14px;
              border-radius: 8px; font-size: 13px; margin-bottom: 18px;
              text-align: center; }}
  </style>
</head>
<body>
  <div class="card">
    <div class="logo-wrap">
      <img src="/static/persico_logo.webp" alt="Persico" onerror="this.style.display='none';document.getElementById('brand-text').style.display='block'">
      <div id="brand-text" style="display:none;font-size:26px;font-weight:900;color:#c8102e;letter-spacing:2px">PERSICO</div>
    </div>
    <p class="sub">Suite · Gestión Interna</p>
    <hr class="divider">
    {'<div class="error">' + error + '</div>' if error else ''}
    <form method="POST">
      <label>Usuario</label>
      <input type="text" name="username" autocomplete="username" required>
      <label>Contraseña</label>
      <input type="password" name="password" autocomplete="current-password" required>
      <button type="submit">Entrar</button>
    </form>
  </div>
</body>
</html>'''

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/emergency-reset-admin")
def emergency_reset_admin():
    """Endpoint de emergencia — regenera el superusuario desde variables de entorno."""
    secret = request.args.get("key","")
    expected = _os.environ.get("SECRET_KEY", "persico-suite-secret-2026")
    if secret != expected:
        return "Clave incorrecta", 403
    try:
        users = {}
        try:
            p = Path(USERS_FILE)
            if p.exists():
                with open(p,"r",encoding="utf-8") as f:
                    users = json.load(f)
        except: pass
        # Restore admin user from env
        admin = _os.environ.get("ADMIN_USER","guillermo")
        users[admin] = {
            "role": "admin",
            "permissions": {m: "full" for m in MODULES}
        }
        Path(USERS_FILE).parent.mkdir(parents=True, exist_ok=True)
        with open(USERS_FILE,"w",encoding="utf-8") as f:
            json.dump(users, f, ensure_ascii=False, indent=2)
        return f"<h2>✅ Usuario '{admin}' restaurado como administrador.</h2><a href='/login'>Ir al login</a>"
    except Exception as e:
        return f"Error: {e}", 500


# ══════════════════════════════════════════════════════════════════
#  SHARED HELPERS
# ══════════════════════════════════════════════════════════════════
def to_str(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool): return v
    return str(v).strip() or None

def esc_csv(s):
    return str(s or "").replace(",", "")

# ══════════════════════════════════════════════════════════════════
#  JOB REGISTER HELPERS
# ══════════════════════════════════════════════════════════════════
def validate_subindex(sub):
    try:
        n = int(sub)
    except ValueError:
        return False
    return (n == 0 or n == 1 or (2 <= n <= 50) or
            (51 <= n <= 60) or (61 <= n <= 97) or n == 99)

def subindex_label(sub):
    try:
        n = int(sub)
    except ValueError:
        return "Desconocido"
    if n == 0:           return "Máquina / equipo principal"
    if n == 1:           return "Instalación y puesta en marcha"
    if 2  <= n <= 50:    return f"Cambio de ingeniería ({n:02d})"
    if 51 <= n <= 60:    return f"Refacción pagada por cliente ({n})"
    if 61 <= n <= 97:    return f"Servicio pagado por cliente ({n})"
    if n == 99:          return "Servicio de garantía"
    return "Índice no válido"

def jobs_root(): return Path(JOBS_FOLDER)
def job_folder(job_number): return jobs_root() / job_number
def meta_path(job_number): return job_folder(job_number) / "job_info.json"

def read_meta(job_number):
    mp = meta_path(job_number)
    if mp.exists():
        try:
            with open(mp, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    parts = job_number.split("-")
    sub = parts[1] if len(parts) > 1 else "00"
    return {
        "job_number": job_number,
        "main_index": int(parts[0]) if parts[0].isdigit() else 0,
        "subindex": sub.zfill(2),
        "subindex_label": subindex_label(sub),
        "customer": "", "pm": "", "description": "",
        "product_group": "", "product_subgroup": "",
        "revenue": 0, "estimated_cost": 0,
        "po_number": "", "ship_date": "",
        "approval_fc": "ToApprove", "status": "Open",
        "notes": "", "created_at": "",
    }

def write_meta(job_number, data):
    with open(meta_path(job_number), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

def scan_jobs():
    root = jobs_root()
    if not root.exists(): return []
    result = []
    for item in sorted(root.iterdir()):
        if item.is_dir() and JOB_RE.match(item.name):
            result.append(read_meta(item.name))
    result.sort(key=lambda j: (j.get("main_index", 0), int(j.get("subindex", "0"))))
    return result

def next_main_index():
    root = jobs_root()
    if not root.exists(): return 100
    indices = []
    for item in root.iterdir():
        if item.is_dir() and JOB_RE.match(item.name):
            try: indices.append(int(item.name.split("-")[0]))
            except ValueError: pass
    return max(indices) + 1 if indices else 100

def all_job_numbers():
    root = jobs_root()
    if not root.exists(): return set()
    return {item.name for item in root.iterdir()
            if item.is_dir() and JOB_RE.match(item.name)}

def extract_customer(full_addr):
    if not full_addr: return ""
    s = str(full_addr).strip()
    m = re.match(r'^([A-Z][A-Z &]+)-', s)
    if m: return m.group(1).strip()
    return re.split(r'[,\n]', s)[0].strip()[:60]

# ══════════════════════════════════════════════════════════════════
#  HOURLY RATE HELPERS
# ══════════════════════════════════════════════════════════════════
def rates_root(): return Path(RATES_FOLDER)
def rates_file(year): return rates_root() / f"rates_{year}.json"

def load_rates(year):
    p = rates_file(year)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_rates(year, records):
    root = rates_root()
    root.mkdir(parents=True, exist_ok=True)
    with open(rates_file(year), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)

def available_years():
    root = rates_root()
    if not root.exists(): return []
    years = []
    for p in root.iterdir():
        m = re.match(r"^rates_(\d{4})\.json$", p.name)
        if m: years.append(int(m.group(1)))
    return sorted(years, reverse=True)

def normalize_name(name):
    return re.sub(r"\s+", " ", str(name).strip().upper())

# ══════════════════════════════════════════════════════════════════
#  QUOTE REGISTER HELPERS
# ══════════════════════════════════════════════════════════════════
# (Quote Register migrado a JSON — sin dependencia de .xlsm)

def _int_or_none(v):
    try: return int(v) if v not in (None, "", "0", 0) else None
    except: return None

def _gen_qnum(records):
    seq = len(records) + 1
    return f"Q-{datetime.date.today().year}-{seq:03d}"

def _quotes_path():
    p = Path(QUOTE_BASE)
    p.mkdir(parents=True, exist_ok=True)
    return p / "quotes.json"

def _load_quotes():
    p = _quotes_path()
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _save_quotes(records):
    with open(_quotes_path(), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)

def read_quote_records():
    with lock:
        records = _load_quotes()
        for i, r in enumerate(records):
            r["row"] = i   # row = índice lógico, compatible con la API existente
        return records

def write_quote_record(data, target_row=None):
    with lock:
        records = _load_quotes()
        if target_row is None:
            qnum = data.get("qnum") or _gen_qnum(records)
            rec = {
                "qnum":       qnum,
                "customer":   data.get("customer", ""),
                "desc":       data.get("desc", ""),
                "machine":    _int_or_none(data.get("machine")),
                "tool":       _int_or_none(data.get("tool")),
                "machTool":   _int_or_none(data.get("machTool")),
                "robotic":    _int_or_none(data.get("robotic")),
                "service":    _int_or_none(data.get("service")),
                "rfq":        data.get("rfq") or None,
                "received":   data.get("received") or None,
                "done":       bool(data.get("done")),
                "sentMgmt":   data.get("sentMgmt") or None,
                "sentClient": data.get("sentClient") or None,
                "deadline":   data.get("deadline") or None,
                "notes":      data.get("notes") or None,
                "awarded":    bool(data.get("awarded")),
                "technicalSales":     data.get("technicalSales") or None,
                "keyAccountManager":  data.get("keyAccountManager") or None,
                "created_at": datetime.datetime.now().isoformat(),
            }
            records.append(rec)
            idx = len(records) - 1
        else:
            idx  = target_row
            if idx < 0 or idx >= len(records):
                raise ValueError(f"Fila {target_row} fuera de rango")
            rec  = records[idx]
            qnum = data.get("qnum", rec.get("qnum"))
            rec.update({
                "qnum":       qnum,
                "customer":   data.get("customer", rec.get("customer", "")),
                "desc":       data.get("desc", rec.get("desc", "")),
                "machine":    _int_or_none(data.get("machine")),
                "tool":       _int_or_none(data.get("tool")),
                "machTool":   _int_or_none(data.get("machTool")),
                "robotic":    _int_or_none(data.get("robotic")),
                "service":    _int_or_none(data.get("service")),
                "rfq":        data.get("rfq") or None,
                "received":   data.get("received") or None,
                "done":       bool(data.get("done")),
                "sentMgmt":   data.get("sentMgmt") or None,
                "sentClient": data.get("sentClient") or None,
                "deadline":   data.get("deadline") or None,
                "notes":      data.get("notes") or None,
                "awarded":    bool(data.get("awarded")),
                "technicalSales":     data.get("technicalSales", rec.get("technicalSales")) or None,
                "keyAccountManager":  data.get("keyAccountManager", rec.get("keyAccountManager")) or None,
                "updated_at": datetime.datetime.now().isoformat(),
            })
        _save_quotes(records)
        rec["row"] = idx
        return rec

def delete_quote_record(target_row):
    with lock:
        records = _load_quotes()
        if 0 <= target_row < len(records):
            records.pop(target_row)
            _save_quotes(records)

# ══════════════════════════════════════════════════════════════════
#  ROUTES — GENERAL
# ══════════════════════════════════════════════════════════════════
@app.before_request
def require_login():
    public = ("/login", "/logout")
    if request.path in public:
        return None
    if not session.get("user"):
        if request.path.startswith("/api/"):
            return jsonify({"error": "no autenticado"}), 401
        return redirect("/login")

@app.route("/")
@_login_required
def index():
    return send_from_directory("static", "index.html")

@app.route("/api/ping")
def ping():
    jobs_ok   = jobs_root().exists()
    rates_ok  = rates_root().exists()
    quotes_file = _quotes_path()
    xlsm_ok   = quotes_file.exists()
    quote_ok  = Path(QUOTE_BASE).exists()
    po_ok     = Path(PO_FOLDER).exists()
    job_count = 0
    if jobs_ok:
        job_count = sum(1 for f in jobs_root().iterdir()
                        if f.is_dir() and JOB_RE.match(f.name))
    return jsonify({
        "jobs_folder":  JOBS_FOLDER,
        "jobs_ok":      jobs_ok,
        "job_count":    job_count,
        "rates_folder": RATES_FOLDER,
        "rates_ok":     rates_ok,
        "years":        available_years(),
        "current_year": CURRENT_YEAR,
        "xlsm_path":    str(quotes_file),
        "xlsm_ok":      xlsm_ok,
        "quote_base":   QUOTE_BASE,
        "quote_ok":     quote_ok,
        "po_folder":    PO_FOLDER,
        "cpo_folder":   CPO_FOLDER,
        "cpo_ok":       Path(CPO_FOLDER).exists(),
        "po_ok":        po_ok,
        "wh_folder":    WH_FOLDER,
        "wh_ok":        Path(WH_FOLDER).exists(),
        "ivp_folder":   IVP_FOLDER,
        "ivp_ok":       Path(IVP_FOLDER).exists(),
        "fx_folder":    FX_FOLDER,
        "fx_ok":        Path(FX_FOLDER).exists(),
    })

# ══════════════════════════════════════════════════════════════════
#  ROUTES — JOB REGISTER  (/api/jobs/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/jobs", methods=["GET"])
def api_get_jobs():
    try: return jsonify(scan_jobs())
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/next-index", methods=["GET"])
def api_next_index():
    try: return jsonify({"next": next_main_index()})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/jobs", methods=["POST"])
def api_create_job():
    try:
        data = request.json
        sub  = str(data.get("subindex", "00")).zfill(2)
        if not validate_subindex(sub):
            return jsonify({"error": f"Subíndice '{sub}' no válido."}), 400
        with lock:
            # Soporte para asociar a job existente (main_index_override)
            main_override = data.get("main_index_override")
            if main_override is not None:
                main = int(main_override)
            else:
                main = next_main_index()
            job_number = f"{main}-{sub}"
            if job_number in all_job_numbers():
                return jsonify({"error": f"El Job {job_number} ya existe."}), 409
            folder = job_folder(job_number)
            try: folder.mkdir(parents=True, exist_ok=True)
            except Exception as fe:
                return jsonify({"error": f"No se pudo crear carpeta en NAS: {fe}"}), 500
            record = {
                "job_number": job_number, "main_index": main,
                "subindex": sub, "subindex_label": subindex_label(sub),
                "customer": data.get("customer", ""),
                "pm": data.get("pm", ""),
                "description": data.get("description", ""),
                "product_group": data.get("product_group", ""),
                "product_subgroup": data.get("product_subgroup", ""),
                "revenue": data.get("revenue", 0),
                "estimated_cost": data.get("estimated_cost", 0),
                "po_number": data.get("po_number", ""),
                "ship_date": data.get("ship_date", ""),
                "approval_fc": data.get("approval_fc", "ToApprove"),
                "status": data.get("status", "Open"),
                "notes": data.get("notes", ""),
                "created_at": datetime.datetime.now().isoformat(),
            }
            write_meta(job_number, record)
            return jsonify(record), 201
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/jobs/<job_number>", methods=["PUT"])
def api_update_job(job_number):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    try:
        data = request.json
        with lock:
            if not job_folder(job_number).exists():
                return jsonify({"error": "Job no encontrado"}), 404
            meta = read_meta(job_number)
            for k in ["customer","pm","description","product_group","product_subgroup",
                      "revenue","estimated_cost","po_number","ship_date",
                      "approval_fc","status","notes"]:
                if k in data: meta[k] = data[k]
            meta["updated_at"] = datetime.datetime.now().isoformat()
            write_meta(job_number, meta)
            return jsonify(meta)
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/jobs/<job_number>", methods=["DELETE"])
def api_delete_job(job_number):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    try:
        folder = job_folder(job_number)
        if folder.exists():
            import shutil as _shutil
            _shutil.rmtree(str(folder))
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/files/<job_number>", methods=["GET"])
def api_list_job_files(job_number):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    folder = job_folder(job_number)
    if not folder.exists(): return jsonify([])
    files = []
    for f in sorted(folder.iterdir()):
        if f.is_file() and f.name != "job_info.json":
            st = f.stat()
            files.append({
                "name": f.name, "size": st.st_size,
                "modified": datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return jsonify(files)

@app.route("/api/files/<job_number>", methods=["POST"])
def api_upload_job_file(job_number):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    folder = job_folder(job_number)
    try: folder.mkdir(parents=True, exist_ok=True)
    except Exception as e: return jsonify({"error": f"No se pudo acceder a la carpeta: {e}"}), 500
    saved = []
    for f in request.files.getlist("files"):
        dest = folder / f.filename
        f.save(str(dest))
        saved.append({"name": f.filename, "size": dest.stat().st_size})
    return jsonify({"saved": saved})

@app.route("/api/files/<job_number>/<filename>", methods=["GET"])
def api_download_job_file(job_number, filename):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    folder = job_folder(job_number)
    if not (folder / filename).exists():
        return jsonify({"error": "Archivo no encontrado"}), 404
    return send_from_directory(str(folder), filename, as_attachment=True)

@app.route("/api/files/<job_number>/<filename>", methods=["DELETE"])
def api_delete_job_file(job_number, filename):
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    target = job_folder(job_number) / filename
    if target.exists() and target.is_file() and target.name != "job_info.json":
        target.unlink()
        return jsonify({"ok": True})
    return jsonify({"error": "Archivo no encontrado"}), 404

@app.route("/api/import-jobs-excel", methods=["POST"])
def api_import_jobs_excel():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error": "No se recibió archivo"}), 400
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value: headers[str(cell.value).strip()] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                if a in headers: return headers[a]
            return None

        ci_job  = col("Job Number", "Job Sequnce#", "Job #")
        ci_pm   = col("PM Assig.", "PM", "PM Assigned")
        ci_desc = col("Job Description", "Description")
        ci_cust = col("Customer and Ship To:", "Customer", "Customer/Ship To")
        ci_rev  = col("Revenue Amount:", "Revenue Amount", "Revenue")
        ci_cost = col("Estimated Cost:", "Estimated Cost", "Cost")
        ci_fc   = col("Approval By FC", "Approval FC", "FC")
        ci_pg   = col("Product Group")
        ci_psg  = col("Product SubGroup", "Product Subgroup")
        ci_po   = col("PO Number", "PO #")
        ci_ship = col("Ship Date")
        ci_date = col("Date Created", "Created")
        ci_note = col("Notes")

        if ci_job is None:
            return jsonify({"error": "No se encontró la columna 'Job Number' en el Excel"}), 400

        year_filter = request.form.get("year", "")
        try: year_filter = int(year_filter) if year_filter else None
        except ValueError: year_filter = None

        def cv(row_vals, idx):
            if idx is None or idx >= len(row_vals): return None
            return row_vals[idx]
        def ts(v): return str(v).strip() if v is not None else ""
        def tf(v):
            try: return float(v) if v is not None else 0
            except: return 0
        def td(v):
            if v is None: return ""
            if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
            return str(v)[:10]

        existing = all_job_numbers()
        results  = {"created": [], "skipped": [], "errors": []}

        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            job_number = ts(cv(row, ci_job)).strip()
            if not job_number or not JOB_RE.match(job_number): continue
            if year_filter and ci_date is not None:
                raw_date = cv(row, ci_date)
                if raw_date and hasattr(raw_date, "year"):
                    if raw_date.year < year_filter:
                        results["skipped"].append({"job": job_number, "reason": f"Año {raw_date.year} < {year_filter}"})
                        continue
            if job_number in existing:
                results["skipped"].append({"job": job_number, "reason": "Ya existe"})
                continue
            parts = job_number.split("-")
            sub = parts[1].zfill(2) if len(parts) > 1 else "00"
            meta = {
                "job_number": job_number,
                "main_index": int(parts[0]) if parts[0].isdigit() else 0,
                "subindex": sub, "subindex_label": subindex_label(sub),
                "customer": extract_customer(ts(cv(row, ci_cust))),
                "customer_full": ts(cv(row, ci_cust)),
                "pm": ts(cv(row, ci_pm)),
                "description": ts(cv(row, ci_desc)),
                "product_group": ts(cv(row, ci_pg)),
                "product_subgroup": ts(cv(row, ci_psg)),
                "revenue": tf(cv(row, ci_rev)),
                "estimated_cost": tf(cv(row, ci_cost)),
                "po_number": ts(cv(row, ci_po)),
                "ship_date": td(cv(row, ci_ship)),
                "approval_fc": ts(cv(row, ci_fc)) or "ToApprove",
                "status": "Open", "notes": ts(cv(row, ci_note)),
                "created_at": td(cv(row, ci_date)), "imported": True,
            }
            with lock:
                folder = job_folder(job_number)
                try:
                    folder.mkdir(parents=True, exist_ok=True)
                    write_meta(job_number, meta)
                    existing.add(job_number)
                    results["created"].append(job_number)
                except Exception as fe:
                    results["errors"].append({"job": job_number, "error": str(fe)})

        results["summary"] = {
            "created": len(results["created"]),
            "skipped": len(results["skipped"]),
            "errors":  len(results["errors"]),
        }
        return jsonify(results)
    except Exception as e: return jsonify({"error": str(e)}), 500

# ══════════════════════════════════════════════════════════════════
#  ROUTES — HOURLY RATE  (/api/rates/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/rates", methods=["GET"])
def api_get_rates():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = load_rates(year)
        return jsonify({"year": year, "records": data, "available_years": available_years()})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/rates", methods=["POST"])
def api_save_rates():
    try:
        payload = request.json
        year    = int(payload.get("year", CURRENT_YEAR))
        records = payload.get("records", [])
        for r in records:
            if not r.get("employee"):
                return jsonify({"error": "Todos los registros deben tener un nombre de empleado"}), 400
            try: float(r["rate"])
            except (ValueError, TypeError):
                return jsonify({"error": f"Tarifa inválida para {r.get('employee')}"}), 400
        with lock: save_rates(year, records)
        return jsonify({"ok": True, "year": year, "count": len(records)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/rates/employee", methods=["PUT"])
def api_update_employee():
    try:
        data     = request.json
        year     = int(data.get("year", CURRENT_YEAR))
        employee = str(data.get("employee", "")).strip()
        rate     = float(data.get("rate", 0))
        dept     = str(data.get("department", "")).strip()
        notes    = str(data.get("notes", "")).strip()
        if not employee: return jsonify({"error": "Nombre de empleado requerido"}), 400
        with lock:
            records = load_rates(year)
            norm = normalize_name(employee)
            found = False
            for r in records:
                if normalize_name(r["employee"]) == norm:
                    r["rate"] = rate; r["department"] = dept; r["notes"] = notes
                    r["updated_at"] = datetime.datetime.now().isoformat()
                    found = True; break
            if not found:
                records.append({
                    "employee": employee, "rate": rate,
                    "department": dept, "notes": notes,
                    "created_at": datetime.datetime.now().isoformat(),
                })
            save_rates(year, records)
            return jsonify({"ok": True, "found": found, "records": records})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/rates/employee", methods=["DELETE"])
def api_delete_employee():
    try:
        data = request.json
        year = int(data.get("year", CURRENT_YEAR))
        norm = normalize_name(str(data.get("employee", "")).strip())
        with lock:
            records = load_rates(year)
            before  = len(records)
            records = [r for r in records if normalize_name(r["employee"]) != norm]
            save_rates(year, records)
        return jsonify({"ok": True, "removed": before - len(records)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/rates/copy-year", methods=["POST"])
def api_copy_year():
    try:
        data        = request.json
        source_year = int(data.get("source_year"))
        target_year = int(data.get("target_year"))
        if source_year == target_year:
            return jsonify({"error": "El año origen y destino deben ser distintos"}), 400
        with lock:
            src = load_rates(source_year)
            if not src: return jsonify({"error": f"No hay tarifas para {source_year}"}), 404
            if rates_file(target_year).exists():
                return jsonify({"error": f"Ya existe una tabla para {target_year}. Elimínala primero."}), 409
            new_records = [{
                "employee": r["employee"], "rate": r["rate"],
                "department": r.get("department", ""), "notes": r.get("notes", ""),
                "copied_from": source_year, "created_at": datetime.datetime.now().isoformat(),
            } for r in src]
            save_rates(target_year, new_records)
        return jsonify({"ok": True, "count": len(new_records), "target_year": target_year})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/import-rates-excel", methods=["POST"])
def api_import_rates_excel():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error": "No se recibió archivo"}), 400
        year = int(request.form.get("year", CURRENT_YEAR))
        mode = request.form.get("mode", "replace")
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value: headers[str(cell.value).strip().upper()] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                if a.upper() in headers: return headers[a.upper()]
            return None

        ci_emp  = col("EMPLOYEE", "NOMBRE", "NAME", "EMPLEADO")
        ci_rate = col("HOURLY RATE", "RATE", "TARIFA", "HOURLY_RATE", "HR RATE")
        ci_dept = col("DEPARTMENT", "DEPT", "DEPARTAMENTO", "AREA")
        ci_note = col("NOTES", "NOTE", "NOTAS", "NOTA")

        if ci_emp is None or ci_rate is None:
            return jsonify({"error": "No se encontraron columnas EMPLOYEE / HOURLY RATE en el Excel"}), 400

        imported = []; errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            emp = str(row[ci_emp]).strip() if row[ci_emp] is not None else ""
            if not emp or emp.upper() == "NONE": continue
            try: rate = float(row[ci_rate]) if row[ci_rate] is not None else 0
            except (ValueError, TypeError):
                errors.append({"employee": emp, "error": "Tarifa no numérica"}); continue
            dept  = str(row[ci_dept]).strip() if ci_dept is not None and row[ci_dept] else ""
            notes = str(row[ci_note]).strip() if ci_note is not None and row[ci_note] else ""
            imported.append({
                "employee": emp, "rate": rate, "department": dept, "notes": notes,
                "imported": True, "created_at": datetime.datetime.now().isoformat(),
            })

        if not imported: return jsonify({"error": "No se encontraron registros válidos en el archivo"}), 400

        with lock:
            if mode == "replace":
                final = imported
            else:
                existing = load_rates(year)
                existing_map = {normalize_name(r["employee"]): r for r in existing}
                for rec in imported:
                    key = normalize_name(rec["employee"])
                    if key in existing_map:
                        existing_map[key]["rate"] = rec["rate"]
                        existing_map[key]["department"] = rec["department"] or existing_map[key].get("department","")
                        existing_map[key]["updated_at"] = rec["created_at"]
                    else:
                        existing_map[key] = rec
                final = list(existing_map.values())
            save_rates(year, final)

        return jsonify({
            "ok": True, "year": year, "mode": mode,
            "imported": len(imported), "total": len(final), "errors": errors,
        })
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/export-rates/<int:year>")
def api_export_rates(year):
    records = load_rates(year)
    lines   = ["EMPLOYEE,HOURLY RATE,DEPARTMENT,NOTES"]
    for r in records:
        lines.append(f"{esc_csv(r.get('employee',''))},{r.get('rate',0)},{esc_csv(r.get('department',''))},{esc_csv(r.get('notes',''))}")
    return Response("\n".join(lines), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=hourly_rates_{year}.csv"})

# ══════════════════════════════════════════════════════════════════
#  ROUTES — QUOTE REGISTER  (/api/quotes/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/quotes", methods=["GET"])
def api_get_quotes():
    try: return jsonify(read_quote_records())
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/quotes", methods=["POST"])
def api_create_quote():
    try:
        data   = request.json
        result = write_quote_record(data, target_row=None)
        return jsonify(result), 201
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/quotes/<int:row>", methods=["PUT"])
def api_update_quote(row):
    try:
        data   = request.json
        result = write_quote_record(data, target_row=row)
        return jsonify(result)
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/quotes/<int:row>", methods=["DELETE"])
def api_delete_quote(row):
    try:
        delete_quote_record(row)
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/quotes/upload/<qnum>", methods=["POST"])
def api_upload_quote(qnum):
    if not re.match(r"^Q-\d{4}-\d{3}$", qnum):
        return jsonify({"error": "Q-Number inválido"}), 400
    folder = Path(QUOTE_BASE) / qnum
    try: folder.mkdir(parents=True, exist_ok=True)
    except Exception as e: return jsonify({"error": f"No se pudo acceder a la carpeta: {e}"}), 500
    saved = []
    for f in request.files.getlist("files"):
        dest = folder / f.filename
        f.save(str(dest))
        saved.append({"name": f.filename, "size": dest.stat().st_size})
    return jsonify({"saved": saved})

@app.route("/api/quotes/files/<qnum>", methods=["GET"])
def api_list_quote_files(qnum):
    if not re.match(r"^Q-\d{4}-\d{3}$", qnum):
        return jsonify({"error": "Q-Number inválido"}), 400
    folder = Path(QUOTE_BASE) / qnum
    if not folder.exists(): return jsonify([])
    files = []
    for f in sorted(folder.iterdir()):
        if f.is_file():
            st = f.stat()
            files.append({
                "name": f.name, "size": st.st_size,
                "modified": datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return jsonify(files)

@app.route("/api/quotes/files/<qnum>/<filename>", methods=["GET"])
def api_download_quote_file(qnum, filename):
    if not re.match(r"^Q-\d{4}-\d{3}$", qnum):
        return jsonify({"error": "Q-Number inválido"}), 400
    folder = Path(QUOTE_BASE) / qnum
    if not (folder / filename).exists():
        return jsonify({"error": "Archivo no encontrado"}), 404
    return send_from_directory(str(folder), filename, as_attachment=True)

@app.route("/api/quotes/files/<qnum>/<filename>", methods=["DELETE"])
def api_delete_quote_file(qnum, filename):
    if not re.match(r"^Q-\d{4}-\d{3}$", qnum):
        return jsonify({"error": "Q-Number inválido"}), 400
    target = Path(QUOTE_BASE) / qnum / filename
    if target.exists() and target.is_file():
        target.unlink()
        return jsonify({"ok": True})
    return jsonify({"error": "Archivo no encontrado"}), 404

@app.route("/api/quotes/import", methods=["POST"])
def api_import_quotes_excel():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error": "No se recibió archivo"}), 400
        mode = request.form.get("mode", "append")  # append | replace

        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        # Intentar la hoja "QUOTE REGISTER", si no existe usar la activa
        ws = wb["QUOTE REGISTER"] if "QUOTE REGISTER" in wb.sheetnames else wb.active

        # Detectar fila de encabezado (buscar celda con Q-NUMBER o CUSTOMER)
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_vals = [str(v).strip().upper() for v in row if v]
            if any(k in row_vals for k in ["Q-NUMBER", "CUSTOMER", "QNUM"]):
                header_row = i
                break
        if header_row is None:
            return jsonify({"error": "No se encontró fila de encabezados (Q-NUMBER / CUSTOMER)"}), 400

        # Mapear columnas
        headers = {}
        for cell in list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]:
            if cell.value:
                headers[str(cell.value).strip().upper().replace("  "," ")] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                if a.upper() in headers: return headers[a.upper()]
            return None

        ci_qnum     = col("Q-NUMBER", "QNUM", "Q NUMBER", "QUOTE")
        ci_cust     = col("CUSTOMER")
        ci_desc     = col("JOB DESCRIPTION", "DESCRIPTION", "DESC")
        ci_machine  = col("MACHINE (BASE)", "MACHINE")
        ci_tool     = col("TOOL / TOLLING", "TOOL", "TOLLING")
        ci_machtool = col("MACHINE & TOOL", "MACHINE & TOOL")
        ci_robotic  = col("ROBOTIC CELL", "ROBOTIC")
        ci_service  = col("SERVICE")
        ci_rfq      = col("RFQ REF.", "RFQ", "RFQ REF")
        ci_received = col("DATE RECEIVED", "RECEIVED")
        ci_done     = col("DONE ✓", "DONE")
        ci_mgmt     = col("SENT TO MANAGEMENT", "SENT MGMT", "SENTMGMT")
        ci_client   = col("SENT TO CUSTOMER", "SENT CUSTOMER", "SENTCLIENT", "SENT TO CLIENT")
        ci_notes    = col("NOTES")
        ci_awarded  = col("AWARDED")

        if ci_cust is None:
            return jsonify({"error": "No se encontró columna CUSTOMER en el Excel"}), 400

        def parse_date(v):
            if v is None: return None
            if isinstance(v, (datetime.date, datetime.datetime)):
                return v.strftime("%Y-%m-%d")
            try:
                # Excel serial date
                base = datetime.date(1899, 12, 30)
                return (base + datetime.timedelta(days=int(float(str(v))))).strftime("%Y-%m-%d")
            except:
                return str(v)[:10] if v else None

        def parse_bool(v):
            if v is None: return False
            if isinstance(v, bool): return v
            return str(v).strip().upper() in ("TRUE", "1", "YES", "SI", "✓", "X")

        def parse_int(v):
            try: return int(v) if v not in (None, "", 0) else None
            except: return None

        imported = []
        errors = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cust = str(row[ci_cust]).strip() if ci_cust is not None and row[ci_cust] else ""
            if not cust or cust.upper() in ("NONE", ""):
                continue
            qnum = str(row[ci_qnum]).strip() if ci_qnum is not None and row[ci_qnum] else None
            rec = {
                "qnum":       qnum,
                "customer":   cust,
                "desc":       str(row[ci_desc]).strip() if ci_desc is not None and row[ci_desc] else "",
                "machine":    parse_int(row[ci_machine])  if ci_machine  is not None else None,
                "tool":       parse_int(row[ci_tool])     if ci_tool     is not None else None,
                "machTool":   parse_int(row[ci_machtool]) if ci_machtool is not None else None,
                "robotic":    parse_int(row[ci_robotic])  if ci_robotic  is not None else None,
                "service":    parse_int(row[ci_service])  if ci_service  is not None else None,
                "rfq":        str(row[ci_rfq]).strip()    if ci_rfq      is not None and row[ci_rfq] else None,
                "received":   parse_date(row[ci_received]) if ci_received is not None else None,
                "done":       parse_bool(row[ci_done])    if ci_done     is not None else False,
                "sentMgmt":   parse_date(row[ci_mgmt])   if ci_mgmt     is not None else None,
                "sentClient": parse_date(row[ci_client]) if ci_client   is not None else None,
                "notes":      str(row[ci_notes]).strip()  if ci_notes    is not None and row[ci_notes] else None,
                "awarded":    parse_bool(row[ci_awarded]) if ci_awarded  is not None else False,
                "created_at": datetime.datetime.now().isoformat(),
            }
            imported.append(rec)

        if not imported:
            return jsonify({"error": "No se encontraron registros válidos en el archivo"}), 400

        with lock:
            if mode == "replace":
                final = imported
            else:  # append
                existing = _load_quotes()
                existing_qnums = {r.get("qnum") for r in existing if r.get("qnum")}
                for rec in imported:
                    if rec["qnum"] and rec["qnum"] in existing_qnums:
                        errors.append({"qnum": rec["qnum"], "error": "Ya existe, omitido"})
                    else:
                        existing.append(rec)
                final = existing
            _save_quotes(final)

        return jsonify({
            "ok": True, "mode": mode,
            "imported": len(imported), "total": len(final), "errors": errors,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ══════════════════════════════════════════════════════════════════
#  PURCHASE ORDERS HELPERS
# ══════════════════════════════════════════════════════════════════
PO_COLS = [
    "clave", "fecha_doc", "entregar_a", "nombre",
    "subtotal", "tipo_cambio", "estatus",
    "descuento_financiero", "pct_descuento", "fecha_recepcion"
]

def po_root(): return Path(PO_FOLDER)
def po_json_file(year): return po_root() / f"po_{year}.json"

def po_available_years():
    root = po_root()
    if not root.exists(): return []
    years = []
    for p in root.iterdir():
        m = re.match(r"^po_(\d{4})\.json$", p.name)
        if m: years.append(int(m.group(1)))
    return sorted(years, reverse=True)

def po_load(year):
    _h = _cache_get(f"po_{year}")
    if _h is not None: return _h
    p = po_json_file(year)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def po_save(year, records):
    root = po_root()
    root.mkdir(parents=True, exist_ok=True)
    with open(po_json_file(year), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)
    _cache_set(f"po_{year}", records)

def po_to_str(v):
    if v is None: return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()

def po_to_float(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

# ══════════════════════════════════════════════════════════════════
#  ROUTES — PURCHASE ORDERS  (/api/po/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/po", methods=["GET"])
def api_get_po():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = po_load(year)
        return jsonify({"year": year, "records": data, "available_years": po_available_years()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/po/import", methods=["POST"])
def api_import_po_excel():
    """
    Importa Purchase Orders desde el Excel con columnas:
      Clave | Fecha de documento | Entregar a | Nombre |
      Subtotal | Tipo de cambio | Estatus |
      Descuento financiero | Porcentaje de descuento | Fecha de recepción
    mode=replace → reemplaza toda la tabla del año
    mode=merge   → agrega / actualiza sin borrar los que no aparecen
    """
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No se recibió archivo"}), 400

        year = int(request.form.get("year", CURRENT_YEAR))
        mode = request.form.get("mode", "append")

        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active

        # Build header map (0-based)
        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers[str(cell.value).strip().lower()] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                if a.lower() in headers: return headers[a.lower()]
            return None

        ci_clave  = col("clave")
        ci_fdoc   = col("fecha de documento")
        ci_dest   = col("entregar a")
        ci_nombre = col("nombre")
        ci_sub    = col("subtotal")
        ci_tc     = col("tipo de cambio")
        ci_est    = col("estatus")
        ci_desc   = col("descuento financiero")
        ci_pct    = col("porcentaje de descuento financ", "porcentaje de descuento financiero", "porcentaje de descuento")
        ci_frec   = col("fecha de recepción", "fecha de recepcion")

        if ci_clave is None:
            return jsonify({"error": "No se encontró la columna 'Clave' en el Excel"}), 400

        imported = []
        errors   = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            def cv(idx):
                if idx is None or idx >= len(row): return None
                return row[idx]

            clave = cv(ci_clave)
            if clave is None or str(clave).strip() == "": continue

            try:
                clave_int = int(clave)
            except (ValueError, TypeError):
                errors.append({"clave": str(clave), "error": "Clave no numérica"})
                continue

            subtotal = po_to_float(cv(ci_sub))
            tc       = po_to_float(cv(ci_tc)) or 1.0
            desc_fin = po_to_float(cv(ci_desc))
            pct_desc = po_to_float(cv(ci_pct))

            imported.append({
                "clave":               clave_int,
                "fecha_doc":           po_to_str(cv(ci_fdoc)),
                "entregar_a":          po_to_str(cv(ci_dest)),
                "nombre":              po_to_str(cv(ci_nombre)),
                "subtotal":            subtotal,
                "tipo_cambio":         tc,
                "subtotal_mxn":        round(subtotal * tc, 2),
                "estatus":             po_to_str(cv(ci_est)),
                "descuento_financiero":desc_fin,
                "pct_descuento":       pct_desc,
                "fecha_recepcion":     po_to_str(cv(ci_frec)),
            })

        if not imported:
            return jsonify({"error": "No se encontraron registros válidos en el archivo"}), 400

        with lock:
            if mode == "replace":
                final = imported
            else:
                existing = po_load(year)
                # Preserve split records (they share a clave but are distinct)
                # Merge: imported records update by clave, but don't touch _split_origin records
                split_records  = [r for r in existing if r.get("_split_origin")]
                normal_records = [r for r in existing if not r.get("_split_origin")]
                existing_map   = {str(r["clave"]): r for r in normal_records}
                for rec in imported:
                    existing_map[str(rec["clave"])] = rec
                final = list(existing_map.values()) + split_records
            po_save(year, final)

        return jsonify({
            "ok":       True,
            "year":     year,
            "mode":     mode,
            "imported": len(imported),
            "total":    len(final),
            "errors":   errors,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/po/<int:year>/pdf/<path:clave>")
def api_po_item_pdf(year, clave):
    """Generate a simple PDF for a single IPO record (no GPO required)."""
    try:
        idx = int(request.args.get("idx", 0))
        records = po_load(year)
        matches = [r for r in records if str(r.get("clave","")).upper()==clave.upper()]
        if not matches or idx >= len(matches):
            return jsonify({"error":"Registro no encontrado"}), 404
        r = matches[idx]
        logo_path = _os.path.join(_BASE,"static","persico_logo.webp"); logo_b64=""
        if _os.path.exists(logo_path):
            import base64 as _b64
            with open(logo_path,"rb") as lf: logo_b64=_b64.b64encode(lf.read()).decode()
        logo_tag=f'<img src="data:image/webp;base64,{logo_b64}" style="height:44px">' if logo_b64 else '<b style="color:#c8102e">PERSICO</b>'
        fmt=lambda v:f"${float(v or 0):,.2f}"
        html=f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>OC {clave}</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:Arial,sans-serif;font-size:11px;padding:28px}}
table{{width:100%;border-collapse:collapse;margin-bottom:12px}}
th{{background:#1f3864;color:#fff;padding:7px 8px;font-size:9px;text-transform:uppercase;text-align:left}}
td{{padding:6px 8px;border-bottom:1px solid #eee}}</style></head>
<body>
<div style="display:flex;justify-content:space-between;border-bottom:3px solid #c8102e;padding-bottom:12px;margin-bottom:16px">
  <div>{logo_tag}</div>
  <div style="text-align:right">
    <div style="font-size:10px;color:#888;text-transform:uppercase">Orden de Compra</div>
    <div style="font-size:20px;font-weight:900;color:#c8102e">{clave}</div>
    <div style="font-size:10px;color:#888">{r.get('fecha_doc','')}</div>
    <div style="font-size:11px;font-weight:700;background:rgba(200,16,46,.1);color:#c8102e;padding:2px 8px;border-radius:4px;margin-top:4px;display:inline-block">{r.get('estatus','—')}</div>
  </div>
</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px;padding:10px;background:#f7f7f7;border-radius:6px;font-size:10px">
  <div><b>Proveedor:</b> {r.get('nombre','—')}</div>
  <div><b>Job / Destino:</b> {r.get('entregar_a','—')}</div>
  <div><b>Moneda:</b> {r.get('moneda','USD')}</div>
  <div><b>Tipo de Cambio:</b> {float(r.get('tipo_cambio',1)):.4f}</div>
</div>
<table><thead><tr><th>No. Parte</th><th>Descripción</th><th style="text-align:right">Cantidad</th><th style="text-align:right">Costo Unit.</th><th style="text-align:right">Total</th></tr></thead>
<tbody><tr>
  <td style="font-family:monospace">{r.get('part_number','—')}</td>
  <td>{r.get('description','—')}</td>
  <td style="text-align:right">{r.get('quantity','—')}</td>
  <td style="text-align:right;font-family:monospace">{fmt(r.get('unit_price',0))}</td>
  <td style="text-align:right;font-family:monospace;font-weight:700">{fmt(r.get('subtotal',0))}</td>
</tr></tbody>
<tfoot><tr style="font-weight:700;background:#1f3864;color:#fff">
  <td colspan="4" style="padding:7px 8px">SUBTOTAL USD</td>
  <td style="padding:7px 8px;text-align:right;font-family:monospace">{fmt(r.get('subtotal',0))}</td>
</tr></tfoot></table>
</body></html>"""
        return Response(html, mimetype="text/html",
            headers={"Content-Disposition":f"inline;filename=OC_{clave}_{idx+1}.html"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500


def api_export_po(year):
    """Exporta Purchase Orders del año como CSV."""
    records = po_load(year)
    lines = ["Clave,Fecha Documento,Entregar A,Nombre,Subtotal,Tipo Cambio,Subtotal MXN,Estatus,Desc.Financiero,% Desc,Fecha Recepción"]
    for r in records:
        lines.append(",".join([
            str(r.get("clave", "")),
            r.get("fecha_doc", ""),
            '"' + r.get("entregar_a", "").replace('"', '') + '"',
            '"' + r.get("nombre", "").replace('"', '') + '"',
            str(r.get("subtotal", 0)),
            str(r.get("tipo_cambio", 1)),
            str(r.get("subtotal_mxn", 0)),
            r.get("estatus", ""),
            str(r.get("descuento_financiero", 0)),
            str(r.get("pct_descuento", 0)),
            r.get("fecha_recepcion", ""),
        ]))
    return Response(
        "\n".join(lines), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=purchase_orders_{year}.csv"}
    )

@app.route("/api/po/years")
def api_po_years():
    return jsonify({"available_years": po_available_years(), "current_year": CURRENT_YEAR})

@app.route("/api/po/<int:year>/edit", methods=["POST"])
def api_po_edit(year):
    """Manually edit a PO/IVP record."""
    try:
        data  = request.get_json()
        clave = str(data.get("clave","")).strip().upper()
        if not clave: return jsonify({"error":"Clave requerida"}), 400
        with lock:
            records = po_load(year)
            matches = [r for r in records
                if str(r.get("gpo_number","")).upper()==clave or
                   str(r.get("clave","")).upper()==clave]
            if not matches:
                return jsonify({"error":f"Registro {clave} no encontrado en {year}"}), 404
            target_pnum = str(data.get("part_number","")).strip().upper()
            if len(matches)>1 and target_pnum:
                rec = next((r for r in matches if str(r.get("part_number","")).upper()==target_pnum), matches[0])
            else:
                rec = matches[0]
            for field in ["nombre","entregar_a","fecha_doc","subtotal","moneda",
                          "tipo_cambio","estatus","fecha_recepcion",
                          "description","part_number","quantity","unit_price",
                          "subtotal_usd","subtotal_mxn"]:
                if field in data and data[field] is not None:
                    rec[field] = data[field]
            rec["_edited_by"] = session.get("user","")
            rec["_edited_at"] = datetime.datetime.now().isoformat()
            po_save(year, records)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/po/<int:year>/<path:clave>", methods=["DELETE"])
def api_delete_po(year, clave):
    if not is_admin(): return jsonify({"error": "Sin permiso — solo administradores"}), 403
    try:
        row_index = request.args.get("idx")   # optional: delete specific row by index
        with lock:
            records = po_load(year)
            if row_index is not None:
                # Delete only the specific row by position
                idx = int(row_index)
                matches = [(i,r) for i,r in enumerate(records)
                           if str(r.get("clave","")) == str(clave)]
                if not matches or idx >= len(matches):
                    return jsonify({"error": "Registro no encontrado"}), 404
                del_pos = matches[idx][0]
                new = [r for i,r in enumerate(records) if i != del_pos]
            else:
                # Delete ALL records with this clave (original behavior)
                new = [r for r in records if str(r.get("clave","")) != str(clave)]
            if len(new) == len(records):
                return jsonify({"error": "Registro no encontrado"}), 404
            po_save(year, new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/gpo/<po_number>", methods=["DELETE"])
def api_delete_gpo(po_number):
    if not is_admin(): return jsonify({"error": "Sin permiso — solo administradores"}), 403
    try:
        # Normalize — accept PO#, PO_, PO- all as the same
        po_clean = po_number.upper().replace("PO#","PO-").replace("PO_","PO-")
        with lock:
            records = gpo_load()
            new = [r for r in records if
                   r.get("po_number","").upper().replace("PO#","PO-").replace("PO_","PO-") != po_clean]
            if len(new) == len(records):
                return jsonify({"error": "PO no encontrada"}), 404
            gpo_save(new)
            # Remove related IPO items
            year = datetime.datetime.now().year
            ipo = po_load(year)
            ipo_new = [r for r in ipo if
                       str(r.get("gpo_number","")).upper().replace("PO#","PO-").replace("PO_","PO-") != po_clean
                       and str(r.get("clave","")).upper().replace("PO#","PO-").replace("PO_","PO-") != po_clean]
            po_save(year, ipo_new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/gpo/normalize", methods=["POST"])
def api_normalize_gpo():
    """Normaliza registros viejos con PO# o PO_ a PO-"""
    if not is_admin(): return jsonify({"error": "Sin permiso"}), 403
    try:
        fixed = 0
        with lock:
            records = gpo_load()
            for r in records:
                old = r.get("po_number","")
                new = old.replace("PO#","PO-").replace("PO_","PO-")
                if old != new:
                    r["po_number"] = new
                    fixed += 1
            gpo_save(records)
            year = datetime.datetime.now().year
            ipo = po_load(year)
            for r in ipo:
                for field in ["clave","gpo_number"]:
                    old = str(r.get(field,""))
                    new = old.replace("PO#","PO-").replace("PO_","PO-")
                    if old != new:
                        r[field] = new
                        fixed += 1
            po_save(year, ipo)
        return jsonify({"ok": True, "fixed": fixed})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  WORK HOURS HELPERS
# ══════════════════════════════════════════════════════════════════
WH_FOLDER  = _os.path.join(_DATA, "WHs")
IVP_FOLDER = _os.path.join(_DATA, "IVPs")


def _homologar_empleado(raw_name, canonical_list, _cache={}):
    """
    Convierte nombres del formato 'ID NOMBRE APELLIDO' al formato canónico
    'APELLIDOS NOMBRE(S)' usando la lista de HOURLY_RATE como referencia.
    """
    import re as _re
    # Quitar ID numérico al inicio
    clean = _re.sub(r"^\d+\s*", "", str(raw_name)).strip().upper()
    if not clean:
        return raw_name

    # Cache para no recalcular
    if clean in _cache:
        return _cache[clean]

    # Buscar mejor coincidencia por palabras compartidas
    words = set(clean.split())
    best_match = clean  # fallback: devolver limpio sin ID
    best_score = 0
    for c in canonical_list:
        c_words = set(c.upper().split())
        score = len(words & c_words)
        if score > best_score:
            best_score = score
            best_match = c

    # Solo usar el canónico si hay al menos 2 palabras en común
    result = best_match if best_score >= 2 else clean
    _cache[clean] = result
    return result

def _get_canonical_employees(year=None):
    """Carga la lista canónica de empleados desde HOURLY_RATE."""
    if year is None:
        year = CURRENT_YEAR
    rates_path = rates_root() / f"rates_{year}.json"
    if not rates_path.exists():
        # Intentar con cualquier año disponible
        root = rates_root()
        if root.exists():
            files = sorted(root.glob("rates_*.json"), reverse=True)
            if files:
                rates_path = files[0]
    if rates_path.exists():
        try:
            with open(rates_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return [r["employee"] for r in data if r.get("employee")]
        except:
            pass
    return []

def wh_root():  return Path(WH_FOLDER)
def ivp_root(): return Path(IVP_FOLDER)

def wh_json_file(year):  return wh_root()  / f"wh_{year}.json"
def ivp_json_file(year): return ivp_root() / f"ivp_{year}.json"

def _generic_available_years(root_fn, prefix):
    root = root_fn()
    if not root.exists(): return []
    years = []
    for p in root.iterdir():
        m = re.match(rf"^{prefix}_(\d{{4}})\.json$", p.name)
        if m: years.append(int(m.group(1)))
    return sorted(years, reverse=True)

def wh_available_years():  return _generic_available_years(wh_root,  "wh")
def ivp_available_years(): return _generic_available_years(ivp_root, "ivp")

def _generic_load(json_file_fn, year):
    p = json_file_fn(year)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def wh_load(year):  return _generic_load(wh_json_file,  year)
def ivp_load(year): return _generic_load(ivp_json_file, year)

def _generic_save(root_fn, json_file_fn, year, records):
    root = root_fn()
    root.mkdir(parents=True, exist_ok=True)
    with open(json_file_fn(year), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)

def wh_save(year, records):  _generic_save(wh_root,  wh_json_file,  year, records)
def ivp_save(year, records): _generic_save(ivp_root, ivp_json_file, year, records)

# ══════════════════════════════════════════════════════════════════
#  ROUTES — WORK HOURS  (/api/wh/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/wh", methods=["GET"])
def api_get_wh():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = wh_load(year)
        return jsonify({"year": year, "records": data, "available_years": wh_available_years()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/wh/import", methods=["POST"])
def api_import_wh():
    """
    Importa Work Hours desde Excel con columnas:
      cboReports | cboFilterFavorites | ID | Employee | Date Worked |
      Work Code  | Hours | Work Description
    Soporta filtro por fecha_inicio / fecha_fin y mode replace/merge.
    """
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No se recibió archivo"}), 400

        year       = int(request.form.get("year", CURRENT_YEAR))
        mode       = request.form.get("mode", "append")
        date_from  = request.form.get("date_from", "")   # YYYY-MM-DD
        date_to    = request.form.get("date_to",   "")   # YYYY-MM-DD

        dt_from = None
        dt_to   = None
        if date_from:
            try: dt_from = datetime.datetime.strptime(date_from[:10], "%Y-%m-%d")
            except: pass
        if date_to:
            try: dt_to = datetime.datetime.strptime(date_to[:10], "%Y-%m-%d")
            except: pass

        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        # Use the first sheet (may be named 'Work Hours List')
        ws = wb.active if len(wb.sheetnames) == 1 else wb[wb.sheetnames[0]]
        for sname in wb.sheetnames:
            if "work hours" in sname.lower():
                ws = wb[sname]; break

        # Detect header row — scan first 3 rows for 'Employee'
        header_row = 1
        headers = {}
        for ri in range(1, 4):
            row_vals = [c.value for c in next(ws.iter_rows(min_row=ri, max_row=ri))]
            if any(str(v).strip().lower() == "employee" for v in row_vals if v):
                header_row = ri
                for ci, v in enumerate(row_vals):
                    if v: headers[str(v).strip().lower()] = ci
                break

        def col(*aliases):
            for a in aliases:
                if a.lower() in headers: return headers[a.lower()]
            return None

        ci_id    = col("id")
        ci_emp   = col("employee")
        ci_date  = col("date worked", "date")
        ci_wcode = col("work code")
        ci_hours = col("hours")
        ci_desc  = col("work description", "description")

        if ci_emp is None or ci_date is None or ci_hours is None:
            return jsonify({"error": "No se encontraron columnas requeridas (Employee, Date Worked, Hours)"}), 400

        imported = []
        skipped  = 0
        errors   = []

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row = list(row)
            def cv(idx):
                if idx is None or idx >= len(row): return None
                return row[idx]

            emp = cv(ci_emp)
            if not emp or str(emp).strip() == "": continue

            date_val = cv(ci_date)
            if not isinstance(date_val, (datetime.datetime, datetime.date)):
                skipped += 1; continue

            # Date range filter
            if dt_from and date_val < dt_from: skipped += 1; continue
            if dt_to   and date_val > dt_to:   skipped += 1; continue

            try:
                hours = float(cv(ci_hours)) if cv(ci_hours) is not None else 0
            except (ValueError, TypeError):
                errors.append({"row": str(cv(ci_id)), "error": "Horas no numéricas"}); continue

            # Homologar nombre al formato canónico
            canonical = _get_canonical_employees(year)
            emp_homolog = _homologar_empleado(str(emp).strip(), canonical) if canonical else                           __import__("re").sub(r"^\d+\s*", "", str(emp).strip()).upper()

            imported.append({
                "id":          int(cv(ci_id)) if cv(ci_id) is not None else None,
                "employee":    emp_homolog,
                "date_worked": date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val)[:10],
                "work_code":   str(cv(ci_wcode) or "").strip(),
                "hours":       hours,
                "description": str(cv(ci_desc) or "").strip(),
            })

        if not imported:
            return jsonify({"error": f"No se encontraron registros válidos (omitidos: {skipped})"}), 400

        with lock:
            if mode == "replace":
                final = imported
            else:
                existing = wh_load(year)
                existing_ids = {r["id"] for r in existing if r.get("id")}
                for rec in imported:
                    if rec.get("id") and rec["id"] in existing_ids:
                        for i, ex in enumerate(existing):
                            if ex.get("id") == rec["id"]:
                                existing[i] = rec; break
                    else:
                        existing.append(rec)
                final = existing
            wh_save(year, final)

        return jsonify({
            "ok":       True,
            "year":     year,
            "mode":     mode,
            "imported": len(imported),
            "skipped":  skipped,
            "total":    len(final),
            "errors":   errors,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/wh/export/<int:year>")
def api_export_wh(year):
    records = wh_load(year)
    lines = ["ID,Employee,Date Worked,Work Code,Hours,Description"]
    for r in records:
        lines.append(",".join([
            str(r.get("id", "")),
            '"' + r.get("employee", "").replace('"', '') + '"',
            r.get("date_worked", ""),
            '"' + r.get("work_code", "").replace('"', '') + '"',
            str(r.get("hours", 0)),
            '"' + r.get("description", "").replace('"', '') + '"',
        ]))
    return Response(
        "\n".join(lines), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=work_hours_{year}.csv"}
    )

# ══════════════════════════════════════════════════════════════════
#  ROUTES — INVOICED POs  (/api/ivp/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/ivp", methods=["GET"])
def api_get_ivp():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = ivp_load(year)
        return jsonify({"year": year, "records": data, "available_years": ivp_available_years()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/ivp/import", methods=["POST"])
def api_import_ivp():
    """
    Importa Invoiced POs desde Excel con columnas:
      Clave | Entregar a | Nombre | Subtotal | Estatus |
      Fecha de recepción | Fecha de pago | Documento anterior
    Detecta USD por '(dolares)' en nombre del proveedor.
    """
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No se recibió archivo"}), 400

        year = int(request.form.get("year", CURRENT_YEAR))
        mode = request.form.get("mode", "append")

        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active

        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers[str(cell.value).strip().lower().rstrip()] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                k = a.lower().rstrip()
                if k in headers: return headers[k]
                # partial match
                for hk in headers:
                    if k in hk or hk in k: return headers[hk]
            return None

        ci_clave = col("clave")
        ci_dest  = col("entregar a")
        ci_nomb  = col("nombre")
        ci_sub   = col("subtotal")
        ci_est   = col("estatus")
        ci_frec  = col("fecha de recepción", "fecha de recepcion")
        ci_fpag  = col("fecha de pago")
        ci_doc   = col("documento anterior")

        if ci_clave is None:
            return jsonify({"error": "No se encontró la columna 'Clave' en el Excel"}), 400

        imported = []
        errors   = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)
            def cv(idx):
                if idx is None or idx >= len(row): return None
                return row[idx]

            clave = cv(ci_clave)
            if clave is None or str(clave).strip() in ("", "None"): continue

            try:
                clave_int = int(clave)
            except (ValueError, TypeError):
                errors.append({"clave": str(clave), "error": "Clave no numérica"}); continue

            nombre = str(cv(ci_nomb) or "").strip()
            is_usd = "(dolares)" in nombre.lower()

            try:
                subtotal = float(cv(ci_sub)) if cv(ci_sub) is not None else 0.0
            except (ValueError, TypeError):
                errors.append({"clave": clave_int, "error": "Subtotal no numérico"}); continue

            estatus = str(cv(ci_est) or "").strip()
            # Skip obviously bad estatus rows (date leaked into column)
            if estatus and re.match(r"\d{4}-\d{2}-\d{2}", estatus):
                estatus = ""

            def to_date(v):
                if v is None: return ""
                if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
                s = str(v)[:10]
                return s if re.match(r"\d{4}-\d{2}-\d{2}", s) else ""

            doc_ant = cv(ci_doc)
            try:
                doc_ant = int(doc_ant) if doc_ant is not None else None
            except (ValueError, TypeError):
                doc_ant = None

            imported.append({
                "clave":            clave_int,
                "entregar_a":       str(cv(ci_dest) or "").strip(),
                "nombre":           nombre,
                "subtotal":         subtotal,
                "moneda":           "USD" if is_usd else "MXN",
                "estatus":          estatus,
                "fecha_recepcion":  to_date(cv(ci_frec)),
                "fecha_pago":       to_date(cv(ci_fpag)),
                "doc_anterior":     doc_ant,
            })

        if not imported:
            return jsonify({"error": "No se encontraron registros válidos en el archivo"}), 400

        with lock:
            if mode == "replace":
                final = imported
            else:
                existing = ivp_load(year)
                existing_map = {r["clave"]: r for r in existing}
                for rec in imported:
                    existing_map[rec["clave"]] = rec
                final = list(existing_map.values())
            ivp_save(year, final)

        return jsonify({
            "ok":       True,
            "year":     year,
            "mode":     mode,
            "imported": len(imported),
            "total":    len(final),
            "errors":   errors,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/ivp/export/<int:year>")
def api_export_ivp(year):
    records = ivp_load(year)
    lines = ["Clave,Entregar A,Nombre,Subtotal,Moneda,Estatus,Fecha Recepcion,Fecha Pago,Doc Anterior"]
    for r in records:
        lines.append(",".join([
            str(r.get("clave", "")),
            '"' + r.get("entregar_a", "").replace('"', '') + '"',
            '"' + r.get("nombre", "").replace('"', '') + '"',
            str(r.get("subtotal", 0)),
            r.get("moneda", "MXN"),
            r.get("estatus", ""),
            r.get("fecha_recepcion", ""),
            r.get("fecha_pago", ""),
            str(r.get("doc_anterior", "") or ""),
        ]))
    return Response(
        "\n".join(lines), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=invoiced_pos_{year}.csv"}
    )

# (run block moved to end)

# ══════════════════════════════════════════════════════════════════
#  JOB REPORT  — /api/report/*
# ══════════════════════════════════════════════════════════════════

def _build_report_data(job_number, rate_year, wh_year, po_year):
    """Core logic: compile all report data for a Job."""
    job_meta = read_meta(job_number) if job_folder(job_number).exists() else {}

    rates_raw = load_rates(rate_year)
    rate_map  = {normalize_name(r["employee"]): float(r["rate"])
                 for r in rates_raw if r.get("employee")}

    wh_raw   = wh_load(wh_year)
    job_main = "-".join(job_number.split("-")[:2]) if "-" in job_number else job_number
    wh_f     = [r for r in wh_raw
                if job_main.upper() in (r.get("work_code") or "").upper()]

    emp_agg = {}
    for r in wh_f:
        emp = str(r.get("employee", "")).strip()
        hrs = float(r.get("hours", 0))
        if emp not in emp_agg:
            emp_agg[emp] = {"employee": emp, "hours": 0.0, "rate": 0.0, "amount": 0.0}
        emp_agg[emp]["hours"] += hrs
        rate = rate_map.get(normalize_name(emp), 0.0)
        emp_agg[emp]["rate"]   = rate
        emp_agg[emp]["amount"] = round(emp_agg[emp]["hours"] * rate, 2)

    workers   = sorted(emp_agg.values(), key=lambda x: x["hours"], reverse=True)
    accum_hrs = round(sum(w["hours"]  for w in workers), 2)
    amount_wh = round(sum(w["amount"] for w in workers), 2)

    po_raw = po_load(po_year)
    po_f   = [r for r in po_raw
              if job_main.upper() in (r.get("entregar_a") or "").upper()
              and r.get("estatus","") != "Cancelada"]  # exclude cancelled

    fx_all = fx_load_all()
    po_items = [{"clave":       r.get("clave"),
                 "nombre":      r.get("nombre", ""),
                 "subtotal":    float(r.get("subtotal", 0)),
                 "moneda":      r.get("moneda", "MXN"),
                 "subtotal_usd": _po_usd(r, fx_all),
                 "fx_rate":     fx_rate_for_date(r.get("fecha_recepcion",""), fx_all) or float(r.get("tipo_cambio",0)) or None,
                 "estatus":     r.get("estatus", ""),
                 "fecha_recepcion": r.get("fecha_recepcion", "")}
                for r in po_f]

    purch_tot = round(sum(p["subtotal_usd"] for p in po_items), 2)
    # Revenue: preferir suma de CPOs si existen
    cpo_rev = cpo_revenue_for_job(job_number, po_year)
    revenue = cpo_rev if cpo_rev > 0 else float(job_meta.get("revenue", 0))
    # Reasignaciones
    try:
        ra_data = reassign_load()
        reassign_items = [
            item for o in ra_data
            for item in o.get("items",[])
            if item.get("job","").upper() == job_number.upper()
        ]
        reassign_total = round(sum(float(i.get("total_cost",0)) for i in reassign_items), 2)
    except:
        reassign_items = []
        reassign_total = 0.0
    # Recovery: negative values that improve margin
    try:
        rc_data = recovery_load()
        recovery_items = [r for r in rc_data if r.get("job","").upper()==job_number.upper()]
        recovery_total = round(sum(float(r.get("total_value",0)) for r in recovery_items), 2)
    except:
        recovery_items = []
        recovery_total = 0.0

    # Service costs (viáticos + gastos de viaje + envíos) — all in USD
    via_items = [r for r in _svc_load(VIATICOS_FILE) if r.get("job","").upper()==job_number.upper()]
    gv_items  = [r for r in _svc_load(GASTOS_FILE)   if r.get("job","").upper()==job_number.upper()]
    env_items = [r for r in _svc_load(ENVIOS_FILE)   if r.get("job","").upper()==job_number.upper()]
    svc_via   = round(sum(r.get("valor_usd",0) for r in via_items), 4)
    svc_gv    = round(sum(r.get("valor_usd",0) for r in gv_items),  4)
    svc_env   = round(sum(r.get("valor_usd",0) for r in env_items), 4)
    svc_total = round(svc_via + svc_gv + svc_env, 4)

    # GM = Revenue - Manpower - POs - Reassignments - Services + Recoveries
    cost = round(amount_wh + purch_tot + reassign_total + svc_total, 2)
    gm   = round(revenue - cost + abs(recovery_total), 2)
    gm_pct    = round((gm / revenue * 100), 1) if revenue else 0.0

    return {
        "job_number":       job_number,
        "customer":         job_meta.get("customer", ""),
        "description":      job_meta.get("description", ""),
        "pm":               job_meta.get("pm", ""),
        "revenue":          revenue,
        "po_number":        job_meta.get("po_number", ""),
        "ship_date":        job_meta.get("ship_date", ""),
        "status":           job_meta.get("status", ""),
        "product_group":    job_meta.get("product_group", ""),
        "accum_hours":      accum_hrs,
        "amount_wh":        amount_wh,
        "workers":          workers,
        "purchasing_total": purch_tot,
        "po_items":         po_items,
        "reassign_total":   reassign_total,
        "reassign_items":   reassign_items,
        "recovery_total":   recovery_total,
        "recovery_items":   recovery_items,
        "svc_viaticos":      svc_via,
        "svc_gastos":        svc_gv,
        "svc_envios":        svc_env,
        "svc_total":         svc_total,
        "svc_viaticos_items":via_items,
        "svc_gastos_items":  gv_items,
        "svc_envios_items":  env_items,
        "cost":             cost,
        "gross_margin":     gm,
        "gm_pct":           gm_pct,
        "rate_year":        rate_year,
        "wh_year":          wh_year,
        "po_year":          po_year,
        "wh_matches":       len(wh_f),
        "po_matches":       len(po_f),
    }


@app.route("/api/report/data")
def api_report_data():
    try:
        job_number = request.args.get("job", "").strip()
        rate_year  = int(request.args.get("rate_year", CURRENT_YEAR))
        wh_year    = int(request.args.get("wh_year",   CURRENT_YEAR))
        po_year    = int(request.args.get("po_year",   CURRENT_YEAR))
        if not job_number:
            return jsonify({"error": "job_number requerido"}), 400
        return jsonify(_build_report_data(job_number, rate_year, wh_year, po_year))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/report/export-excel")
def api_report_export_excel():
    """Exporta el reporte como .xlsx siguiendo la estructura del template."""
    from flask import make_response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    job_number = request.args.get("job", "").strip()
    rate_year  = int(request.args.get("rate_year", CURRENT_YEAR))
    wh_year    = int(request.args.get("wh_year",   CURRENT_YEAR))
    po_year    = int(request.args.get("po_year",   CURRENT_YEAR))

    if not job_number:
        return jsonify({"error": "job_number requerido"}), 400

    d = _build_report_data(job_number, rate_year, wh_year, po_year)

    # Colour palette
    RED_H   = "C8102E"
    DARK    = "1F1F1F"
    DGRAY   = "2D2D2D"
    MGRAY   = "3D3D3D"
    LGRAY   = "F0F0F0"
    XLGRAY  = "FAFAFA"
    GOLD    = "F0A500"
    WHITE   = "FFFFFF"
    GREEN_C = "1E8449"
    RED_NEG = "C0392B"
    BLUE_H  = "1F618D"

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def _font(sz=9, bold=False, color=DARK, italic=False):
        return Font(name="Arial", size=sz, bold=bold, color=color, italic=italic)
    def _side():
        return Side(style="thin", color="AAAAAA")
    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)
    def _lft(indent=1):
        return Alignment(horizontal="left",   vertical="center", indent=indent, wrap_text=False)
    def _rgt():
        return Alignment(horizontal="right",  vertical="center")
    def _ctr():
        return Alignment(horizontal="center", vertical="center")
    MONEY = '#,##0.00'
    HRS   = '#,##0.0'
    PCT   = '0.0"%"'

    wb = Workbook()
    ws = wb.active
    ws.title = f"Report {job_number}"

    # Column widths
    widths = {"A":26,"B":18,"C":28,"D":10,"E":14,"F":16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Row 1: Title ─────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"  JOB COST REPORT  ·  {job_number}"
    c.font      = _font(16, True, WHITE)
    c.fill      = _fill(RED_H)
    c.alignment = _lft(2)

    # ── Row 2: Sub-header ─────────────────────────────────────────
    ws.row_dimensions[2].height = 14
    ws.merge_cells("A2:C2")
    ws["A2"].value = f"Generated: {datetime.date.today()}  |  Rate year: {rate_year}  |  WH year: {wh_year}  |  PO year: {po_year}"
    ws["A2"].font  = _font(8, False, "888888", True)
    ws["A2"].alignment = _lft()

    # ── Summary block (rows 3-14) ──────────────────────────────────
    def label_row(row, label, val, fmt=None, bg_lbl=DGRAY, bg_val=MGRAY,
                  fc_lbl=WHITE, fc_val=WHITE, bold_val=False, height=18):
        ws.row_dimensions[row].height = height
        cl = ws.cell(row, 1)
        cl.value = label; cl.font = _font(9, True, fc_lbl)
        cl.fill = _fill(bg_lbl); cl.alignment = _lft(); cl.border = _border()
        cv = ws.cell(row, 2)
        cv.value = val; cv.font = _font(10, bold_val, fc_val)
        cv.fill = _fill(bg_val); cv.alignment = _lft()
        cv.border = _border()
        if fmt: cv.number_format = fmt

    label_row(3,  "JOB NUMBER",            job_number,              bg_val=MGRAY, fc_val=GOLD, bold_val=True)
    label_row(4,  "CUSTOMER",              d["customer"] or "—",    bg_val=MGRAY)
    label_row(5,  "PM",                    d["pm"] or "—",          bg_val=MGRAY)
    label_row(6,  "DESCRIPTION",           d["description"] or "—", bg_val=MGRAY)
    label_row(7,  "STATUS",                d["status"] or "—",      bg_val=MGRAY)

    # Spacer
    ws.row_dimensions[8].height = 6
    ws.merge_cells("A8:F8"); ws["A8"].fill = _fill(DARK)

    label_row(9,  "REVENUE",              d["revenue"],       MONEY, bg_val=BLUE_H, fc_val=WHITE, bold_val=True)
    label_row(10, "ACUMULATED WORK HOURS",d["accum_hours"],   HRS,   bg_val=MGRAY,  fc_val=WHITE)
    label_row(11, "AMOUNT WORK HOURS",    d["amount_wh"],     MONEY, bg_val=MGRAY,  fc_val=WHITE, bold_val=True)
    label_row(12, "PURCHASINGS TOTAL",    d["purchasing_total"], MONEY, bg_val=MGRAY, fc_val=WHITE, bold_val=True)

    ws.row_dimensions[13].height = 6
    ws.merge_cells("A13:F13"); ws["A13"].fill = _fill(DARK)

    # COST
    ws.row_dimensions[14].height = 20
    cl = ws["A14"]; cl.value = "COST"
    cl.font = _font(11, True, WHITE); cl.fill = _fill(RED_H)
    cl.alignment = _lft(); cl.border = _border()
    cv = ws["B14"]; cv.value = d["cost"]
    cv.font = _font(12, True, WHITE); cv.fill = _fill(RED_H)
    cv.alignment = _rgt(); cv.number_format = MONEY; cv.border = _border()

    # GROSS MARGIN
    ws.row_dimensions[15].height = 22
    gm_bg = GREEN_C if d["gross_margin"] >= 0 else RED_NEG
    cl = ws["A15"]; cl.value = "GROSS MARGIN"
    cl.font = _font(12, True, WHITE); cl.fill = _fill(gm_bg)
    cl.alignment = _lft(); cl.border = _border()
    cv = ws["B15"]; cv.value = d["gross_margin"]
    cv.font = _font(13, True, WHITE); cv.fill = _fill(gm_bg)
    cv.alignment = _rgt(); cv.number_format = MONEY; cv.border = _border()

    # GM%
    ws.row_dimensions[16].height = 16
    ws.merge_cells("A16:B16")
    c16 = ws["A16"]
    c16.value = f"Gross Margin %:  {d['gm_pct']:.1f}%"
    c16.font  = _font(10, True, WHITE)
    c16.fill  = _fill(gm_bg); c16.alignment = _ctr(); c16.border = _border()

    # ── Detail tables (right side of summary, rows 3-16) ──────────
    # PO mini-list header (cols D-F, row 3)
    ws.row_dimensions[3].height = max(ws.row_dimensions[3].height, 18)
    for col, txt in [(4,"CLAVE PO"),(5,"MXN"),(6,"PROVEEDOR")]:
        c = ws.cell(3, col)
        c.value = txt; c.font = _font(8, True, WHITE)
        c.fill = _fill(RED_H); c.alignment = _ctr(); c.border = _border()

    for i, po in enumerate(d["po_items"][:12]):
        r = 4 + i
        ws.row_dimensions[r].height = 15
        bg = LGRAY if i % 2 == 0 else XLGRAY
        ws.cell(r,4).value = str(po["clave"]); ws.cell(r,4).font = _font(8,False,"333333")
        ws.cell(r,4).fill = _fill(bg); ws.cell(r,4).alignment = _ctr(); ws.cell(r,4).border = _border()
        ws.cell(r,5).value = po["subtotal_usd"]; ws.cell(r,5).font = _font(8,False,"333333")
        ws.cell(r,5).fill = _fill(bg); ws.cell(r,5).alignment = _rgt()
        ws.cell(r,5).number_format = MONEY; ws.cell(r,5).border = _border()
        ws.cell(r,6).value = po["nombre"][:30]; ws.cell(r,6).font = _font(7,False,"666666")
        ws.cell(r,6).fill = _fill(bg); ws.cell(r,6).alignment = _lft(); ws.cell(r,6).border = _border()

    # ── Spacer ────────────────────────────────────────────────────
    sr = 17
    ws.row_dimensions[sr].height = 8
    ws.merge_cells(f"A{sr}:F{sr}")
    ws[f"A{sr}"].fill = _fill(DARK)

    # ── Detail tables header row ──────────────────────────────────
    dh = sr + 1
    ws.row_dimensions[dh].height = 22
    for col, txt in [(1,"PO NUMBER"),(2,"VALUE (USD)"),(3,"PROVEEDOR"),
                     (4,"WORKER"),(5,"HOURS"),(6,"VALUE (USD)")]:
        c = ws.cell(dh, col)
        c.value = txt; c.font = _font(9, True, WHITE)
        c.fill = _fill(DGRAY); c.alignment = _ctr(); c.border = _border()

    # ── PO detail rows ────────────────────────────────────────────
    po_start = dh + 1
    for i, po in enumerate(d["po_items"]):
        r = po_start + i
        ws.row_dimensions[r].height = 15
        bg = LGRAY if i % 2 == 0 else XLGRAY
        ws.cell(r,1).value = str(po["clave"])
        ws.cell(r,1).font = _font(9,False,"222222"); ws.cell(r,1).fill = _fill(bg)
        ws.cell(r,1).alignment = _ctr(); ws.cell(r,1).border = _border()
        ws.cell(r,2).value = po["subtotal_usd"]
        ws.cell(r,2).font = _font(9,False,"222222"); ws.cell(r,2).fill = _fill(bg)
        ws.cell(r,2).alignment = _rgt(); ws.cell(r,2).number_format = MONEY
        ws.cell(r,2).border = _border()
        ws.cell(r,3).value = po["nombre"][:35]
        ws.cell(r,3).font = _font(8,False,"555555"); ws.cell(r,3).fill = _fill(bg)
        ws.cell(r,3).alignment = _lft(); ws.cell(r,3).border = _border()

    # ── Worker detail rows ────────────────────────────────────────
    wk_start = dh + 1
    for i, w in enumerate(d["workers"]):
        r = wk_start + i
        if r < po_start + len(d["po_items"]):
            ws.row_dimensions[r].height = max(ws.row_dimensions[r].height, 15)
        else:
            ws.row_dimensions[r].height = 15
        bg = LGRAY if i % 2 == 0 else XLGRAY
        ws.cell(r,4).value = w["employee"]
        ws.cell(r,4).font = _font(9,False,"222222"); ws.cell(r,4).fill = _fill(bg)
        ws.cell(r,4).alignment = _lft(); ws.cell(r,4).border = _border()
        ws.cell(r,5).value = w["hours"]
        ws.cell(r,5).font = _font(9,False,"222222"); ws.cell(r,5).fill = _fill(bg)
        ws.cell(r,5).alignment = _rgt(); ws.cell(r,5).number_format = HRS
        ws.cell(r,5).border = _border()
        ws.cell(r,6).value = w["amount"]
        ws.cell(r,6).font = _font(9,False,"222222"); ws.cell(r,6).fill = _fill(bg)
        ws.cell(r,6).alignment = _rgt(); ws.cell(r,6).number_format = MONEY
        ws.cell(r,6).border = _border()

    # ── Totals footer ─────────────────────────────────────────────
    foot = max(po_start+len(d["po_items"]), wk_start+len(d["workers"])) + 1
    ws.row_dimensions[foot].height = 20
    for col, val, fmt in [
        (1,"TOTAL",None),(2,d["purchasing_total"],MONEY),(3,"",None),
        (4,"TOTAL",None),(5,d["accum_hours"],HRS),(6,d["amount_wh"],MONEY)]:
        c = ws.cell(foot, col)
        c.value = val; c.font = _font(10, True, WHITE)
        c.fill = _fill(DARK); c.alignment = _rgt() if fmt else _ctr()
        c.border = _border()
        if fmt: c.number_format = fmt

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Report_{job_number.replace('-','_')}_{datetime.date.today()}.xlsx"
    resp  = make_response(buf.read())
    resp.headers["Content-Type"]        = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp

# ══════════════════════════════════════════════════════════════════
#  FX (TIPO DE CAMBIO) — /api/fx/*
# ══════════════════════════════════════════════════════════════════

def fx_root():           return Path(FX_FOLDER)
def fx_json_file(year):  return fx_root() / f"fx_{year}.json"

def fx_available_years():
    root = fx_root()
    if not root.exists(): return []
    years = []
    for p in root.iterdir():
        m = re.match(r"^fx_(\d{4})\.json$", p.name)
        if m: years.append(int(m.group(1)))
    return sorted(years, reverse=True)

def fx_load(year) -> dict:
    """Returns {YYYY-MM-DD: rate_float}"""
    p = fx_json_file(year)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def fx_save(year, data: dict):
    root = fx_root()
    root.mkdir(parents=True, exist_ok=True)
    with open(fx_json_file(year), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def fx_load_all() -> dict:
    """Merge all years into one lookup dict {YYYY-MM-DD: rate}"""
    combined = {}
    for year in fx_available_years():
        combined.update(fx_load(year))
    return combined

def fx_rate_for_date(date_str: str, fx_all: dict) -> float:
    """
    Returns the MXN/USD rate for a given date string (YYYY-MM-DD).
    Falls back up to 7 days earlier for weekends/holidays.
    Returns None if not found.
    """
    if not date_str or not fx_all:
        return None
    try:
        d = datetime.datetime.strptime(date_str[:10], "%Y-%m-%d").date()
    except ValueError:
        return None
    for offset in range(8):
        key = (d - datetime.timedelta(days=offset)).strftime("%Y-%m-%d")
        if key in fx_all:
            return fx_all[key]
    return None


@app.route("/api/fx", methods=["GET"])
def api_get_fx():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        data = fx_load(year)
        # Return as sorted list for frontend table
        records = [{"date": k, "rate": v} for k, v in sorted(data.items())]
        return jsonify({
            "year":            year,
            "records":         records,
            "available_years": fx_available_years(),
            "count":           len(records),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fx/import", methods=["POST"])
def api_import_fx():
    """
    Importa el archivo tipoCambio.xls del Banco de México.
    Formato: filas de datos a partir de fila 9 (idx 8).
    Col 0 = Fecha (dd/mm/yyyy string)
    Col 3 = Tipo de cambio 'Para solventar obligaciones'
    """
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No se recibió archivo"}), 400

        mode = request.form.get("mode", "merge")   # merge default — accumulate years
        raw  = f.read()

        # Support both .xls (legacy binary) and .xlsx
        fname_lower = (f.filename or "").lower()

        def parse_rows_from_raw(raw_bytes, is_xls):
            """
            Returns an iterable of rows starting from data row 9 (idx 8).
            For .xls we try xlrd, then fall back to converting via LibreOffice,
            then fall back to a minimal built-in compound-doc reader.
            """
            if is_xls:
                # ── Try xlrd first (if installed) ──────────────────
                try:
                    import xlrd
                    wb2 = xlrd.open_workbook(file_contents=raw_bytes)
                    ws2 = wb2.sheet_by_index(0)
                    return [[ws2.cell_value(r, c) for c in range(ws2.ncols)]
                            for r in range(0, ws2.nrows)]
                except ImportError:
                    pass

                # ── Try pandas with openpyxl-xlrd engine ───────────
                try:
                    import pandas as pd
                    df = pd.read_excel(io.BytesIO(raw_bytes), header=None,
                                       engine='xlrd')
                    return df.values.tolist()
                except Exception:
                    pass

                # ── Convert .xls → .xlsx via LibreOffice ────────────
                import tempfile, subprocess, os
                with tempfile.TemporaryDirectory() as tmpdir:
                    src = os.path.join(tmpdir, "tc.xls")
                    with open(src, "wb") as fh:
                        fh.write(raw_bytes)
                    result = subprocess.run(
                        ["libreoffice", "--headless", "--convert-to", "xlsx",
                         "--outdir", tmpdir, src],
                        capture_output=True, timeout=30
                    )
                    out_path = src.replace(".xls", ".xlsx")
                    if result.returncode != 0 or not os.path.exists(out_path):
                        raise RuntimeError(
                            "No se pudo convertir el archivo. "
                            "Instala xlrd: pip install xlrd==1.2.0 --break-system-packages"
                        )
                    with open(out_path, "rb") as fh:
                        xlsx_bytes = fh.read()

                wb3 = openpyxl.load_workbook(io.BytesIO(xlsx_bytes),
                                              read_only=True, data_only=True)
                ws3 = wb3.active
                return list(ws3.iter_rows(min_row=1, values_only=True))

            else:
                wb4 = openpyxl.load_workbook(io.BytesIO(raw_bytes),
                                              read_only=True, data_only=True)
                ws4 = wb4.active
                return list(ws4.iter_rows(min_row=1, values_only=True))

        is_xls = fname_lower.endswith(".xls") and not fname_lower.endswith(".xlsx")
        try:
            all_rows = parse_rows_from_raw(raw, is_xls)
        except RuntimeError as re_err:
            return jsonify({"error": str(re_err)}), 400

        rows_iter = all_rows

        # Parse rows
        by_year   = {}    # year → {YYYY-MM-DD: rate}
        imported  = 0
        skipped   = 0

        # Auto-detect header row and column positions
        header_map = {}
        data_start  = 0
        for ri, row in enumerate(rows_iter[:5]):
            if not row: continue
            vals = [str(v).strip().upper() if v else "" for v in row]
            if any(k in vals for k in ["FECHA","DATE","TASA","RATE","TASA_MXN_USD"]):
                for ci, v in enumerate(vals):
                    if v in ("FECHA","DATE"): header_map["date"] = ci
                    if v in ("TASA","RATE","TASA_MXN_USD","TASA (MXN/USD)"): header_map["rate"] = ci
                data_start = ri + 1
                break

        date_col = header_map.get("date", 0)
        rate_col = header_map.get("rate", 3)  # fallback: col 3 (Banxico format)

        for row in rows_iter[data_start:]:
            if not row or not row[date_col]:
                continue
            date_raw = str(row[date_col]).strip()
            rate_raw = row[rate_col] if len(row) > rate_col else None

            # Parse date — soporta dd/mm/yyyy y yyyy-mm-dd
            d = None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
                try:
                    d = datetime.datetime.strptime(date_raw[:10], fmt)
                    break
                except ValueError:
                    continue
            if d is None:
                skipped += 1
                continue

            # Parse rate — skip N/E
            try:
                rate = float(rate_raw)
                if rate <= 0:
                    raise ValueError
            except (TypeError, ValueError):
                skipped += 1
                continue

            year    = d.year
            iso_key = d.strftime("%Y-%m-%d")
            by_year.setdefault(year, {})[iso_key] = round(rate, 6)
            imported += 1

        if not by_year:
            return jsonify({"error": "No se encontraron registros válidos"}), 400

        total_saved = 0
        with lock:
            for year, new_data in by_year.items():
                if mode == "replace":
                    final = new_data
                else:   # merge
                    existing = fx_load(year)
                    existing.update(new_data)
                    final = existing
                fx_save(year, final)
                total_saved += len(final)

        years_touched = sorted(by_year.keys())
        return jsonify({
            "ok":       True,
            "imported": imported,
            "skipped":  skipped,
            "years":    years_touched,
            "total_saved": total_saved,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/fx/banxico", methods=["POST"])
def api_fx_banxico():
    """Consulta el tipo de cambio del día desde la API de Banxico y lo registra."""
    import urllib.request, json as _json
    try:
        token = _os.environ.get("BANXICO_TOKEN", "")
        # Endpoint público de Banxico — serie SF43718 = Fix MXN/USD
        url = "https://www.banxico.org.mx/SieAPIRest/service/v1/series/SF43718/datos/oportuno"
        headers_req = {"Bmx-Token": token} if token else {}
        req = urllib.request.Request(url, headers=headers_req)
        with urllib.request.urlopen(req, timeout=8) as resp:
            raw = _json.loads(resp.read().decode("utf-8"))

        # Parse Banxico response
        series = raw.get("bmx", {}).get("series", [])
        if not series:
            return jsonify({"error": "Respuesta vacía de Banxico"}), 502
        datos = series[0].get("datos", [])
        if not datos:
            return jsonify({"error": "Sin datos en la respuesta de Banxico"}), 502

        # Take the most recent entry
        entry    = datos[-1]
        fecha_mx = entry.get("fecha", "")   # format: DD/MM/YYYY
        dato     = entry.get("dato", "N/E")

        if dato in ("N/E", "N/D", ""):
            return jsonify({"error": f"Banxico no tiene dato disponible para {fecha_mx}"}), 400

        rate = round(float(dato), 6)

        # Convert date to ISO
        d = datetime.datetime.strptime(fecha_mx, "%d/%m/%Y")
        iso_key = d.strftime("%Y-%m-%d")
        year    = d.year

        with lock:
            existing = fx_load(year)
            existing[iso_key] = rate
            fx_save(year, existing)

        return jsonify({"ok": True, "fecha": iso_key, "rate": rate,
                        "source": "Banxico SF43718 (Fix MXN/USD)"})
    except urllib.error.URLError as e:
        return jsonify({"error": f"No se pudo conectar a Banxico: {str(e)}"}), 503
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def api_fx_lookup():
    """Quick single-date lookup: ?date=YYYY-MM-DD"""
    try:
        date_str = request.args.get("date", "")
        fx_all   = fx_load_all()
        rate     = fx_rate_for_date(date_str, fx_all)
        return jsonify({"date": date_str, "rate": rate, "found": rate is not None})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Helper used by PO and IVP endpoints: convert subtotal to USD ─
def _po_usd(record: dict, fx_all: dict) -> float:
    """Return the subtotal in USD for a PO record."""
    subtotal = float(record.get("subtotal", 0))
    # Zero subtotal (cancelled) returns 0 immediately
    if subtotal == 0:
        return 0.0
    moneda = str(record.get("moneda", "MXN")).upper()
    if moneda == "USD":
        return subtotal   # already in USD — no conversion
    # MXN → USD
    date_str = record.get("fecha_recepcion") or record.get("fecha_doc") or ""
    rate     = fx_rate_for_date(date_str, fx_all)
    if rate and rate > 0:
        return round(subtotal / rate, 6)
    tc = float(record.get("tipo_cambio", 0))
    if tc > 1:
        return round(subtotal / tc, 6)
    return subtotal   # can't convert — return as-is


@app.route("/api/po/usd-view")
def api_po_usd_view():
    """Return PO records with all amounts converted to USD."""
    try:
        year   = int(request.args.get("year", CURRENT_YEAR))
        po_raw = po_load(year)
        fx_all = fx_load_all()
        result = []
        for r in po_raw:
            rec = dict(r)
            rec["subtotal_usd"] = _po_usd(r, fx_all)
            # Determine which rate was used
            if r.get("moneda") == "USD":
                rec["fx_rate_used"] = 1.0
            else:
                date_str = r.get("fecha_recepcion") or r.get("fecha_doc") or ""
                rec["fx_rate_used"] = fx_rate_for_date(date_str, fx_all) or float(r.get("tipo_cambio", 0)) or None
            result.append(rec)
        return jsonify({"year": year, "records": result, "available_years": po_available_years()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ══════════════════════════════════════════════════════════════════
#  CUSTOMER POs (CPO) HELPERS
# ══════════════════════════════════════════════════════════════════
CPO_FOLDER = _os.path.join(_DATA, "CPOs")

def cpo_root(): return Path(CPO_FOLDER)
def cpo_json_file(year): return cpo_root() / f"cpo_{year}.json"

def cpo_available_years():
    root = cpo_root()
    if not root.exists(): return []
    years = []
    for p in root.iterdir():
        m = re.match(r"^cpo_(\d{4})\.json$", p.name)
        if m: years.append(int(m.group(1)))
    return sorted(years, reverse=True)

def cpo_load(year):
    p = cpo_json_file(year)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def cpo_save(year, records):
    root = cpo_root()
    root.mkdir(parents=True, exist_ok=True)
    with open(cpo_json_file(year), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)

def cpo_to_float(v):
    try: return float(v) if v not in (None, "", "N/A", "#N/A") else 0.0
    except: return 0.0

def cpo_to_str(v):
    if v is None or str(v).strip() in ("#N/A", "None"): return ""
    return str(v).strip()

def cpo_parse_date(v):
    if v is None: return None
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime("%Y-%m-%d")
    try:
        base = datetime.date(1899, 12, 30)
        return (base + datetime.timedelta(days=int(float(str(v))))).strftime("%Y-%m-%d")
    except:
        s = str(v).strip()
        return s[:10] if s else None

def cpo_revenue_for_job(job_number, year):
    """Suma de VALUE de todas las CPOs asociadas a este job en el año dado."""
    records = cpo_load(year)
    job_main = job_number.upper()
    total = sum(cpo_to_float(r.get("value")) for r in records
                if (r.get("job") or "").upper() == job_main)
    return round(total, 2)

# ══════════════════════════════════════════════════════════════════
#  ROUTES — CUSTOMER POs  (/api/cpo/*)
# ══════════════════════════════════════════════════════════════════
@app.route("/api/cpo", methods=["GET"])
def api_get_cpo():
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        job  = request.args.get("job", "").strip().upper()
        records = cpo_load(year)
        if job:
            records = [r for r in records if (r.get("job") or "").upper() == job]
        return jsonify({"year": year, "records": records,
                        "available_years": cpo_available_years()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cpo", methods=["POST"])
def api_create_cpo():
    try:
        data = request.get_json()
        year = int(data.get("year", CURRENT_YEAR))
        records = cpo_load(year)
        rec = {
            "id":           f"CPO-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
            "type_id":      cpo_to_str(data.get("type_id")),
            "po_number":    cpo_to_str(data.get("po_number")),
            "date":         cpo_to_str(data.get("date")),
            "job":          cpo_to_str(data.get("job")).upper(),
            "customer_supplier": cpo_to_str(data.get("customer_supplier")),
            "value":        cpo_to_float(data.get("value")),
            "type_name":    cpo_to_str(data.get("type_name", "01_REVENUE")),
            "customer":     cpo_to_str(data.get("customer")),
            "year":         year,
            "pm":           cpo_to_str(data.get("pm")),
            "status":       cpo_to_str(data.get("status", "WIP")),
            "est_finalize": cpo_to_str(data.get("est_finalize")),
            "created_at":   datetime.datetime.now().isoformat(),
        }
        records.append(rec)
        cpo_save(year, records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cpo/<cpo_id>", methods=["PUT"])
def api_update_cpo(cpo_id):
    try:
        data = request.get_json()
        year = int(data.get("year", CURRENT_YEAR))
        records = cpo_load(year)
        idx = next((i for i, r in enumerate(records) if r.get("id") == cpo_id), None)
        if idx is None:
            return jsonify({"error": "CPO no encontrada"}), 404
        rec = records[idx]
        for k in ["type_id","po_number","date","job","customer_supplier",
                  "type_name","customer","pm","status","est_finalize"]:
            if k in data: rec[k] = cpo_to_str(data[k])
        if "value" in data: rec["value"] = cpo_to_float(data["value"])
        rec["updated_at"] = datetime.datetime.now().isoformat()
        rec["year"] = year
        cpo_save(year, records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cpo/<cpo_id>", methods=["DELETE"])
def api_delete_cpo(cpo_id):
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        records = cpo_load(year)
        new_records = [r for r in records if r.get("id") != cpo_id]
        if len(new_records) == len(records):
            return jsonify({"error": "CPO no encontrada"}), 404
        cpo_save(year, new_records)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cpo/import", methods=["POST"])
def api_import_cpo_excel():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error": "No se recibió archivo"}), 400
        year = int(request.form.get("year", CURRENT_YEAR))
        mode = request.form.get("mode", "append")

        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            if cell.value:
                headers[str(cell.value).strip().upper()] = cell.column - 1

        def col(*aliases):
            for a in aliases:
                if a.upper() in headers: return headers[a.upper()]
            return None

        ci_tid   = col("TYPE ID")
        ci_po    = col("PO NUMBER", "NAME/NUMBER / ID", "NAME/NUMBER/ID")
        ci_date  = col("DATE")
        ci_job   = col("JOB")
        ci_cs    = col("CUSTOMER/SUPPLIER/CC", "CUSTOMER/SUPPLIER")
        ci_val   = col("VALUE")
        ci_tn    = col("TYPE NAME")
        ci_cust  = col("CUSTOMER")
        ci_yr    = col("YEAR")
        ci_pm    = col("PM")
        ci_pnum  = col("PO NUMBER")
        ci_stat  = col("STATUS")
        ci_est   = col("ESTIMATED TIME TO FINALIZE", "EST TIME TO FINALIZE")

        if ci_job is None or ci_val is None:
            return jsonify({"error": "No se encontraron columnas JOB / VALUE"}), 400

        imported = []; errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            job = cpo_to_str(row[ci_job] if ci_job is not None else "")
            if not job or job.upper() in ("NONE", ""): continue
            val = cpo_to_float(row[ci_val] if ci_val is not None else 0)
            # Determinar año del registro
            rec_year = year
            if ci_yr is not None and row[ci_yr] not in (None, "", "#N/A"):
                try: rec_year = int(float(str(row[ci_yr])))
                except: pass
            rec = {
                "id":           f"CPO-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                "type_id":      cpo_to_str(row[ci_tid]  if ci_tid  is not None else ""),
                "po_number":    cpo_to_str(row[ci_pnum] if ci_pnum is not None else (row[ci_po] if ci_po is not None else "")),
                "date":         cpo_parse_date(row[ci_date] if ci_date is not None else None),
                "job":          job.upper(),
                "customer_supplier": cpo_to_str(row[ci_cs]   if ci_cs   is not None else ""),
                "value":        val,
                "type_name":    cpo_to_str(row[ci_tn]   if ci_tn   is not None else "01_REVENUE"),
                "customer":     cpo_to_str(row[ci_cust] if ci_cust is not None else ""),
                "year":         rec_year,
                "pm":           cpo_to_str(row[ci_pm]   if ci_pm   is not None else ""),
                "status":       cpo_to_str(row[ci_stat] if ci_stat is not None else "WIP"),
                "est_finalize": cpo_parse_date(row[ci_est] if ci_est is not None else None),
                "created_at":   datetime.datetime.now().isoformat(),
            }
            imported.append((rec_year, rec))

        if not imported:
            return jsonify({"error": "No se encontraron registros válidos"}), 400

        with lock:
            if mode == "replace":
                # Agrupar por año
                by_year = {}
                for yr, rec in imported:
                    by_year.setdefault(yr, []).append(rec)
                for yr, recs in by_year.items():
                    cpo_save(yr, recs)
            else:
                # Append agrupado por año
                by_year = {}
                for yr, rec in imported:
                    by_year.setdefault(yr, []).append(rec)
                for yr, recs in by_year.items():
                    existing = cpo_load(yr)
                    existing.extend(recs)
                    cpo_save(yr, existing)

        total = sum(len(cpo_load(yr)) for yr in set(yr for yr, _ in imported))
        return jsonify({"ok": True, "mode": mode, "imported": len(imported),
                        "total": total, "errors": errors})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cpo/revenue/<job_number>")
def api_cpo_revenue(job_number):
    try:
        year = int(request.args.get("year", CURRENT_YEAR))
        rev  = cpo_revenue_for_job(job_number, year)
        cpos = [r for r in cpo_load(year)
                if (r.get("job") or "").upper() == job_number.upper()]
        return jsonify({"job": job_number, "year": year,
                        "revenue": rev, "cpo_count": len(cpos), "cpos": cpos})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/report/multi", methods=["POST"])
def api_report_multi():
    """Reporte agrupado de múltiples jobs."""
    try:
        data      = request.get_json()
        jobs      = [j.strip() for j in data.get("jobs", []) if j.strip()]
        rate_year = int(data.get("rate_year", CURRENT_YEAR))
        wh_year   = int(data.get("wh_year",   CURRENT_YEAR))
        po_year   = int(data.get("po_year",   CURRENT_YEAR))
        cpo_year  = int(data.get("cpo_year",  CURRENT_YEAR))
        label     = data.get("label", "Multi-Job Report")
        if not jobs:
            return jsonify({"error": "Se requiere al menos un job"}), 400

        rows = []
        totals = {"revenue": 0, "amount_wh": 0, "purchasing_total": 0,
                  "cost": 0, "gross_margin": 0, "accum_hours": 0,
                  "reassign_total": 0, "recovery_total": 0,
                  "svc_total": 0, "svc_viaticos": 0, "svc_gastos": 0, "svc_envios": 0}
        for jn in jobs:
            d = _build_report_data(jn, rate_year, wh_year, po_year)
            # Usar CPO como Revenue si hay registros
            cpo_rev = cpo_revenue_for_job(jn, cpo_year)
            if cpo_rev > 0:
                d["revenue"]      = cpo_rev
                d["cost"]         = round(d["amount_wh"] + d["purchasing_total"] + d.get("svc_total",0), 2)
                d["gross_margin"] = round(cpo_rev - d["cost"] + d.get("recovery_total",0), 2)
                d["gm_pct"]       = round((d["gross_margin"] / cpo_rev * 100), 1) if cpo_rev else 0.0
                d["revenue_source"] = "CPO"
            else:
                d["revenue_source"] = "job_meta"
            rows.append({
                "job_number":       d["job_number"],
                "customer":         d["customer"],
                "description":      d["description"],
                "pm":               d["pm"],
                "revenue":          d["revenue"],
                "accum_hours":      d["accum_hours"],
                "amount_wh":        d["amount_wh"],
                "purchasing_total": d["purchasing_total"],
                "reassign_total":   d.get("reassign_total", 0),
                "recovery_total":   d.get("recovery_total", 0),
                "svc_total":        d.get("svc_total", 0),
                "svc_viaticos":     d.get("svc_viaticos", 0),
                "svc_gastos":       d.get("svc_gastos", 0),
                "svc_envios":       d.get("svc_envios", 0),
                "cost":             d["cost"],
                "gross_margin":     d["gross_margin"],
                "gm_pct":           d["gm_pct"],
                "revenue_source":   d.get("revenue_source", "job_meta"),
            })
            for k in totals:
                totals[k] = round(totals[k] + d.get(k, 0), 2)

        totals["gm_pct"] = round((totals["gross_margin"] / totals["revenue"] * 100), 1) if totals["revenue"] else 0.0
        return jsonify({"label": label, "jobs": rows, "totals": totals,
                        "rate_year": rate_year, "wh_year": wh_year,
                        "po_year": po_year, "cpo_year": cpo_year})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ══════════════════════════════════════════════════════════════════
#  PT NUMBERS
# ══════════════════════════════════════════════════════════════════
PT_FILE = _os.path.join(_DATA, "pt_numbers.json")

def pt_load():
    p = Path(PT_FILE)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except: return []
    return []

def pt_save(records):
    Path(PT_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(PT_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

@app.route("/api/pt", methods=["GET"])
def api_get_pt():
    try:
        records = pt_load()
        q = request.args.get("q","").lower()
        if q:
            records = [r for r in records if q in json.dumps(r, ensure_ascii=False).lower()]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/pt", methods=["POST"])
def api_create_pt():
    try:
        data = request.get_json()
        with lock:
            records = pt_load()
            pt_num = str(data.get("pt_number","")).strip().upper()
            if not pt_num:
                return jsonify({"error": "PT Number es requerido"}), 400
            if any(r["pt_number"] == pt_num for r in records):
                return jsonify({"error": f"{pt_num} ya existe"}), 409
        rec = {
            "pt_number":        pt_num,
            "customer":         str(data.get("customer","")).strip(),
            "customer_program": str(data.get("customer_program","")).strip(),
            "pm":               str(data.get("pm","")).strip(),
            "jobs":             [j.strip().upper() for j in data.get("jobs",[]) if j.strip()],
            "notes":            str(data.get("notes","")).strip(),
            "created_at":       datetime.datetime.now().isoformat(),
        }
        records.append(rec)
        pt_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/pt/<pt_number>", methods=["PUT"])
def api_update_pt(pt_number):
    try:
        data = request.get_json()
        records = pt_load()
        idx = next((i for i,r in enumerate(records) if r["pt_number"]==pt_number.upper()), None)
        if idx is None:
            return jsonify({"error": "PT no encontrado"}), 404
        rec = records[idx]
        for k in ["customer","customer_program","pm","notes"]:
            if k in data: rec[k] = str(data[k]).strip()
        if "jobs" in data:
            rec["jobs"] = [j.strip().upper() for j in data["jobs"] if j.strip()]
        rec["updated_at"] = datetime.datetime.now().isoformat()
        pt_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/pt/<pt_number>", methods=["DELETE"])
def api_delete_pt(pt_number):
    try:
        records = pt_load()
        new_records = [r for r in records if r["pt_number"] != pt_number.upper()]
        if len(new_records) == len(records):
            return jsonify({"error": "PT no encontrado"}), 404
        pt_save(new_records)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/pt/<pt_number>/jobs", methods=["GET"])
def api_pt_jobs(pt_number):
    """Devuelve los jobs asociados a un PT con info básica de cada job."""
    try:
        records = pt_load()
        pt = next((r for r in records if r["pt_number"]==pt_number.upper()), None)
        if pt is None:
            return jsonify({"error": "PT no encontrado"}), 404
        jobs_info = []
        for jn in pt.get("jobs", []):
            info_path = job_folder(jn) / "job_info.json"
            if info_path.exists():
                try:
                    with open(info_path, "r", encoding="utf-8") as f:
                        ji = json.load(f)
                    jobs_info.append({"job_number": jn,
                                      "customer": ji.get("customer",""),
                                      "description": ji.get("description",""),
                                      "pm": ji.get("pm","")})
                except:
                    jobs_info.append({"job_number": jn, "customer":"","description":"","pm":""})
            else:
                jobs_info.append({"job_number": jn, "customer":"","description":"","pm":""})
        return jsonify({"pt": pt, "jobs": jobs_info})
    except Exception as e:
        return jsonify({"error": str(e)}), 500




# ══════════════════════════════════════════════════════════════════
#  USERS & PERMISSIONS
# ══════════════════════════════════════════════════════════════════
USERS_FILE  = _os.path.join(_DATA, "users.json")
ADMIN_USER  = _os.environ.get("ADMIN_USER", "guillermo")

MODULES = [
    # Proyectos
    "jobs", "pt", "sv", "rates", "quotes",
    # Ventas
    "cpo",
    # Compras — Catálogos
    "cat-electrico", "cat-mecanico", "cat-servicios",
    # Compras — Proveedores
    "proveedores",
    # Compras — Documentos
    "gpo", "po", "ivp", "reassign", "recovery",
    # Almacenes
    "stock", "ingreso", "apartados", "salida",
    # Servicio
    "viaticos", "gastos-viaje", "envios",
    # Reportes y Configuración
    "wh", "report", "multirpt", "fx", "projconfig",
]

# Levels: "none" | "view" | "create" | "full"
LEVEL_NONE   = "none"
LEVEL_VIEW   = "view"
LEVEL_CREATE = "create"
LEVEL_FULL   = "full"

def _default_perms(role):
    if role == "admin":
        return {m: LEVEL_FULL for m in MODULES}
    else:
        return {m: LEVEL_VIEW for m in MODULES}

def _level_gte(level, minimum):
    """Check if level >= minimum."""
    order = [LEVEL_NONE, LEVEL_VIEW, LEVEL_CREATE, LEVEL_FULL]
    try:
        return order.index(level) >= order.index(minimum)
    except: return False

def users_load():
    p = Path(USERS_FILE)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
                # Migrate old format {action: bool} → level string
                for uname, info in data.items():
                    perms = info.get("permissions", {})
                    migrated = {}
                    for mod, val in perms.items():
                        if isinstance(val, dict):
                            # Old format — convert
                            if val.get("delete") or val.get("import"):
                                migrated[mod] = LEVEL_FULL
                            elif val.get("create") or val.get("edit"):
                                migrated[mod] = LEVEL_CREATE
                            elif val.get("view"):
                                migrated[mod] = LEVEL_VIEW
                            else:
                                migrated[mod] = LEVEL_NONE
                        else:
                            migrated[mod] = val  # already new format
                    info["permissions"] = migrated
                return data
        except: pass
    # Build from auth file or env vars
    users = {}
    try:
        auth = _load_auth()
        for uname in auth:
            role = "admin" if uname == ADMIN_USER else "viewer"
            users[uname] = {"role": role, "permissions": _default_perms(role)}
    except:
        for i in range(1, 10):
            val = _os.environ.get(f"USER{i}", "")
            if ":" in val:
                uname = val.split(":", 1)[0].strip()
                role  = "admin" if uname == ADMIN_USER else "viewer"
                users[uname] = {"role": role, "permissions": _default_perms(role)}
    return users

def users_save(users):
    Path(USERS_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def get_user_perms(username):
    users = users_load()
    if username not in users:
        role = "admin" if username == ADMIN_USER else "viewer"
        return {"role": role, "permissions": _default_perms(role)}
    return users[username]

def _get_level(module):
    """Get permission level for current user on a module."""
    user = session.get("user")
    if not user: return LEVEL_NONE
    info = get_user_perms(user)
    if info.get("role") == "admin": return LEVEL_FULL
    return info.get("permissions", {}).get(module, LEVEL_NONE)

def can(action, module):
    """Backward-compatible permission check."""
    user = session.get("user")
    if not user: return False
    info = get_user_perms(user)
    if info.get("role") == "admin": return True
    level = info.get("permissions", {}).get(module, LEVEL_NONE)
    if action in ("view",):
        return _level_gte(level, LEVEL_VIEW)
    elif action in ("create", "edit"):
        return _level_gte(level, LEVEL_CREATE)
    elif action in ("delete", "import"):
        return _level_gte(level, LEVEL_FULL)
    return False

def is_admin():
    user = session.get("user")
    if not user: return False
    info = get_user_perms(user)
    return info.get("role") == "admin"

# ── Permission routes
@app.route("/api/admin/users", methods=["GET"])
def api_admin_get_users():
    if not is_admin():
        return jsonify({"error": "Sin permiso"}), 403
    users = users_load()
    for i in range(1, 10):
        val = _os.environ.get(f"USER{i}", "")
        if ":" in val:
            uname = val.split(":", 1)[0].strip()
            if uname not in users:
                role = "admin" if uname == ADMIN_USER else "viewer"
                users[uname] = {"role": role, "permissions": _default_perms(role)}
    return jsonify({"users": users, "modules": MODULES,
                    "current_user": session.get("user"), "admin_user": ADMIN_USER})

@app.route("/api/admin/users/<username>", methods=["PUT"])
def api_admin_update_user(username):
    if not is_admin():
        return jsonify({"error": "Sin permiso"}), 403
    data  = request.get_json()
    users = users_load()
    if username not in users:
        role = "admin" if username == ADMIN_USER else "viewer"
        users[username] = {"role": role, "permissions": _default_perms(role)}
    new_role = data.get("role", users[username]["role"])
    users[username]["role"] = new_role
    if "permissions" in data:
        # Merge single-module update into existing permissions
        existing = users[username].get("permissions", _default_perms(new_role))
        existing.update(data["permissions"])
        users[username]["permissions"] = existing
    elif new_role != users[username].get("role"):
        users[username]["permissions"] = _default_perms(new_role)
    users_save(users)
    return jsonify({"ok": True, "user": users[username]})

@app.route("/api/me/perms", methods=["GET"])
def api_me_perms():
    user = session.get("user")
    if not user: return jsonify({"error": "No autenticado"}), 401
    info = get_user_perms(user)
    return jsonify({"user": user, "role": info.get("role","viewer"),
                    "permissions": info.get("permissions", _default_perms("viewer")),
                    "is_admin": is_admin()})



# ══════════════════════════════════════════════════════════════════
#  SV NUMBERS
# ══════════════════════════════════════════════════════════════════
SV_FILE = _os.path.join(_DATA, "sv_numbers.json")

def sv_load():
    p = Path(SV_FILE)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except: return []
    return []

def sv_save(records):
    Path(SV_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(SV_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def sv_next_number():
    records = sv_load()
    if not records:
        return "SV0001"
    nums = []
    for r in records:
        try: nums.append(int(r["sv_number"].replace("SV","")))
        except: pass
    return f"SV{(max(nums)+1):04d}" if nums else "SV0001"

def pt_next_number():
    records = pt_load()
    if not records:
        return "PT0001"
    nums = []
    for r in records:
        try: nums.append(int(r["pt_number"].replace("PT","")))
        except: pass
    return f"PT{(max(nums)+1):04d}" if nums else "PT0001"

@app.route("/api/sv", methods=["GET"])
def api_get_sv():
    try:
        records = sv_load()
        q = request.args.get("q","").lower()
        if q:
            records = [r for r in records if q in json.dumps(r, ensure_ascii=False).lower()]
        return jsonify({"records": records, "total": len(records),
                        "next_number": sv_next_number()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/sv", methods=["POST"])
def api_create_sv():
    try:
        data = request.get_json()
        with lock:
            records = sv_load()
            sv_num = str(data.get("sv_number","")).strip().upper()
            if not sv_num:
                sv_num = sv_next_number()  # called inside lock — safe
            if any(r["sv_number"] == sv_num for r in records):
                return jsonify({"error": f"{sv_num} ya existe"}), 409
        rec = {
            "sv_number":        sv_num,
            "customer":         str(data.get("customer","")).strip(),
            "customer_program": str(data.get("customer_program","")).strip(),
            "pm":               str(data.get("pm","")).strip(),
            "jobs":             [j.strip().upper() for j in data.get("jobs",[]) if j.strip()],
            "notes":            str(data.get("notes","")).strip(),
            "created_at":       datetime.datetime.now().isoformat(),
        }
        records.append(rec)
        sv_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/sv/<sv_number>", methods=["GET"])
def api_get_sv_one(sv_number):
    try:
        records = sv_load()
        rec = next((r for r in records if r["sv_number"]==sv_number.upper()), None)
        if rec is None:
            return jsonify({"error": "SV no encontrado"}), 404
        jobs_info = []
        for jn in rec.get("jobs", []):
            info_path = job_folder(jn) / "job_info.json"
            if info_path.exists():
                try:
                    with open(info_path, "r", encoding="utf-8") as f:
                        ji = json.load(f)
                    jobs_info.append({"job_number": jn,
                                      "customer": ji.get("customer",""),
                                      "description": ji.get("description",""),
                                      "pm": ji.get("pm","")})
                except:
                    jobs_info.append({"job_number": jn, "customer":"","description":"","pm":""})
            else:
                jobs_info.append({"job_number": jn, "customer":"","description":"","pm":""})
        return jsonify({"sv": rec, "jobs": jobs_info})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/sv/<sv_number>", methods=["PUT"])
def api_update_sv(sv_number):
    try:
        data = request.get_json()
        records = sv_load()
        idx = next((i for i,r in enumerate(records) if r["sv_number"]==sv_number.upper()), None)
        if idx is None:
            return jsonify({"error": "SV no encontrado"}), 404
        rec = records[idx]
        for k in ["customer","customer_program","pm","notes"]:
            if k in data: rec[k] = str(data[k]).strip()
        if "jobs" in data:
            rec["jobs"] = [j.strip().upper() for j in data["jobs"] if j.strip()]
        rec["updated_at"] = datetime.datetime.now().isoformat()
        sv_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/sv/<sv_number>", methods=["DELETE"])
def api_delete_sv(sv_number):
    try:
        records = sv_load()
        new_records = [r for r in records if r["sv_number"] != sv_number.upper()]
        if len(new_records) == len(records):
            return jsonify({"error": "SV no encontrado"}), 404
        sv_save(new_records)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ══════════════════════════════════════════════════════════════════
#  WORKFLOW — Quote → CPO → Job → PT/SV
# ══════════════════════════════════════════════════════════════════

@app.route("/api/workflow/next-numbers", methods=["GET"])
def api_next_numbers():
    """Devuelve el siguiente PT y SV disponibles."""
    return jsonify({
        "next_pt": pt_next_number(),
        "next_sv": sv_next_number(),
        "pt_list": [{"pt_number": r["pt_number"], "customer": r.get("customer",""),
                     "customer_program": r.get("customer_program","")}
                    for r in pt_load()],
        "sv_list": [{"sv_number": r["sv_number"], "customer": r.get("customer",""),
                     "customer_program": r.get("customer_program","")}
                    for r in sv_load()],
    })

@app.route("/api/workflow/award", methods=["POST"])
def api_workflow_award():
    """
    Flujo AWARDED:
    1. Marca la cotización como awarded
    2. Crea la CPO
    3. Crea los Jobs
    4. Asigna o crea PT/SV
    """
    try:
        data    = request.get_json()
        q_row   = int(data.get("q_row", -1))
        year    = int(data.get("cpo_year", CURRENT_YEAR))

        # ── Paso 1: Marcar cotización como AWARDED
        quotes = _load_quotes()
        if q_row < 0 or q_row >= len(quotes):
            return jsonify({"error": "Cotización no encontrada"}), 404
        quote = quotes[q_row]
        quote["awarded"]        = True
        quote["award_date"]     = datetime.datetime.now().isoformat()
        quote["cpo_registered"] = True
        if data.get("refused_reason"):
            quote["refused"] = False
        _save_quotes(quotes)

        results = {"quote": quote["qnum"], "cpo": None, "jobs": [], "pt_sv": None}

        # ── Paso 2: Crear CPO (venta)
        cpo_data = data.get("cpo")
        if cpo_data:
            cpo_year = int(cpo_data.get("year", year))
            cpos = cpo_load(cpo_year)
            cpo_rec = {
                "id":                f"CPO-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                "type_id":           "CPO",
                "po_number":         cpo_to_str(cpo_data.get("po_number","")),
                "date":              cpo_to_str(cpo_data.get("date","")),
                "job":               "",  # se actualiza al crear jobs
                "customer_supplier": cpo_to_str(cpo_data.get("customer_supplier","")),
                "value":             cpo_to_float(cpo_data.get("value",0)),
                "type_name":         "01_REVENUE",
                "customer":          cpo_to_str(cpo_data.get("customer","")),
                "year":              cpo_year,
                "pm":                cpo_to_str(cpo_data.get("pm","")),
                "status":            "WIP",
                "est_finalize":      cpo_to_str(cpo_data.get("est_finalize","")),
                "q_number":          quote["qnum"],
                "created_at":        datetime.datetime.now().isoformat(),
            }
            cpos.append(cpo_rec)
            cpo_save(cpo_year, cpos)
            results["cpo"] = cpo_rec["id"]

        # ── Paso 3: Crear Jobs
        job_numbers_created = []
        jobs_data = data.get("jobs", [])
        with lock:
            base_main = None  # main index del primer job
            for jd in jobs_data:
                sub = str(jd.get("subindex","00")).zfill(2)
                use_base = jd.get("use_base_main_index", False) and base_main is not None
                if use_base:
                    main = base_main
                else:
                    main = next_main_index()
                if base_main is None:
                    base_main = main
                job_number = f"{main}-{sub}"
                folder = job_folder(job_number)
                folder.mkdir(parents=True, exist_ok=True)
                record = {
                    "job_number":      job_number,
                    "main_index":      main,
                    "subindex":        sub,
                    "subindex_label":  subindex_label(sub),
                    "customer":        jd.get("customer", quote.get("customer","")),
                    "pm":              jd.get("pm", quote.get("pm","")),
                    "description":     jd.get("description", quote.get("desc","")),
                    "product_group":   jd.get("product_group",""),
                    "product_subgroup":jd.get("product_subgroup",""),
                    "revenue":         cpo_to_float(jd.get("value")) if jd.get("value") not in (None, 0, "", "0") else cpo_to_float(data.get("cpo",{}).get("value",0)),
                    "estimated_cost":  0,
                    "po_number":       cpo_to_str(data.get("cpo",{}).get("po_number","")),
                    "ship_date":       "",
                    "approval_fc":     "ToApprove",
                    "status":          "Open",
                    "notes":           jd.get("notes",""),
                    "q_number":        quote["qnum"],
                    "cpo_id":          results.get("cpo",""),
                    "created_at":      datetime.datetime.now().isoformat(),
                }
                write_meta(job_number, record)
                job_numbers_created.append(job_number)
        results["jobs"] = job_numbers_created

        # Actualizar job en CPO
        if cpo_data and job_numbers_created and results["cpo"]:
            cpos = cpo_load(year)
            for c in cpos:
                if c.get("id") == results["cpo"]:
                    c["job"] = job_numbers_created[0]
            cpo_save(year, cpos)

        # ── Paso 4: Asignar PT o SV
        pt_sv_data = data.get("pt_sv")
        if pt_sv_data and job_numbers_created:
            kind = pt_sv_data.get("kind","pt")  # "pt" o "sv"
            mode = pt_sv_data.get("mode","new")  # "new" o "existing"
            if kind == "pt":
                records = pt_load()
                if mode == "new":
                    num = pt_sv_data.get("number", pt_next_number())
                    rec = {
                        "pt_number":        num,
                        "customer":         pt_sv_data.get("customer", quote.get("customer","")),
                        "customer_program": pt_sv_data.get("customer_program",""),
                        "pm":               pt_sv_data.get("pm",""),
                        "jobs":             job_numbers_created,
                        "notes":            pt_sv_data.get("notes",""),
                        "q_number":         quote["qnum"],
                        "created_at":       datetime.datetime.now().isoformat(),
                    }
                    records.append(rec)
                else:
                    num = pt_sv_data.get("number","")
                    for r in records:
                        if r["pt_number"] == num:
                            r["jobs"] = list(set(r.get("jobs",[]) + job_numbers_created))
                pt_save(records)
                results["pt_sv"] = {"kind":"pt","number":num}
                # Actualizar job_info con pt_number
                for jn in job_numbers_created:
                    m = read_meta(jn)
                    m["pt_number"] = num
                    write_meta(jn, m)
            else:  # sv
                records = sv_load()
                if mode == "new":
                    num = pt_sv_data.get("number", sv_next_number())
                    rec = {
                        "sv_number":        num,
                        "customer":         pt_sv_data.get("customer", quote.get("customer","")),
                        "customer_program": pt_sv_data.get("customer_program",""),
                        "pm":               pt_sv_data.get("pm",""),
                        "jobs":             job_numbers_created,
                        "notes":            pt_sv_data.get("notes",""),
                        "q_number":         quote["qnum"],
                        "created_at":       datetime.datetime.now().isoformat(),
                    }
                    records.append(rec)
                else:
                    num = pt_sv_data.get("number","")
                    for r in records:
                        if r["sv_number"] == num:
                            r["jobs"] = list(set(r.get("jobs",[]) + job_numbers_created))
                sv_save(records)
                results["pt_sv"] = {"kind":"sv","number":num}
                for jn in job_numbers_created:
                    m = read_meta(jn)
                    m["sv_number"] = num
                    write_meta(jn, m)

        return jsonify({"ok": True, "results": results})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/workflow/refuse", methods=["POST"])
def api_workflow_refuse():
    """Marca una cotización como REFUSED con motivo opcional."""
    try:
        data   = request.get_json()
        q_row  = int(data.get("q_row", -1))
        reason = str(data.get("reason","")).strip()
        quotes = _load_quotes()
        if q_row < 0 or q_row >= len(quotes):
            return jsonify({"error": "Cotización no encontrada"}), 404
        quotes[q_row]["awarded"]       = False
        quotes[q_row]["refused"]       = True
        quotes[q_row]["refuse_reason"] = reason
        quotes[q_row]["refuse_date"]   = datetime.datetime.now().isoformat()
        _save_quotes(quotes)
        return jsonify({"ok": True, "qnum": quotes[q_row]["qnum"]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/pt/next-number", methods=["GET"])
def api_pt_next():
    with lock:
        return jsonify({"next": pt_next_number()})

@app.route("/api/sv/next-number", methods=["GET"])
def api_sv_next():
    with lock:
        return jsonify({"next": sv_next_number()})



@app.route("/api/jobs/merge", methods=["POST"])
def api_merge_jobs():
    """Fusiona source_job hacia target_job. Solo admins."""
    if not is_admin():
        return jsonify({"error": "Sin permiso — solo administradores"}), 403
    try:
        data        = request.get_json()
        source      = str(data.get("source","")).strip().upper()
        target      = str(data.get("target","")).strip().upper()
        if not JOB_RE.match(source) or not JOB_RE.match(target):
            return jsonify({"error": "Job numbers inválidos"}), 400
        if source == target:
            return jsonify({"error": "Source y target no pueden ser iguales"}), 400

        source_folder = job_folder(source)
        target_folder = job_folder(target)

        if not source_folder.exists():
            return jsonify({"error": f"{source} no existe"}), 404
        if not target_folder.exists():
            return jsonify({"error": f"{target} no existe"}), 404

        moved_files = []
        skipped_files = []

        # Mover todos los archivos (excepto job_info.json) de source a target
        for f in source_folder.iterdir():
            if f.name == "job_info.json":
                continue
            dest = target_folder / f.name
            if dest.exists():
                # Renombrar con sufijo para no sobreescribir
                stem = f.stem; suffix = f.suffix; i = 1
                while dest.exists():
                    dest = target_folder / f"{stem}_from_{source}_{i}{suffix}"
                    i += 1
                skipped_files.append(f.name)
            f.rename(dest)
            moved_files.append(f.name)

        # Migrar Work Hours
        wh_year = int(data.get("wh_year", CURRENT_YEAR))
        wh_path = Path(WH_FOLDER) / f"wh_{wh_year}.json"
        wh_updated = 0
        if wh_path.exists():
            with open(wh_path, "r", encoding="utf-8") as f:
                wh_data = json.load(f)
            for rec in wh_data:
                if rec.get("work_code","").upper() == source:
                    rec["work_code"] = target
                    wh_updated += 1
            with open(wh_path, "w", encoding="utf-8") as f:
                json.dump(wh_data, f, ensure_ascii=False, indent=2)

        # Migrar IPOs
        po_year = int(data.get("po_year", CURRENT_YEAR))
        po_path = Path(PO_FOLDER) / f"po_{po_year}.json"
        po_updated = 0
        if po_path.exists():
            with open(po_path, "r", encoding="utf-8") as f:
                po_data = json.load(f)
            for rec in po_data:
                if str(rec.get("job","")).upper() == source:
                    rec["job"] = target
                    po_updated += 1
            with open(po_path, "w", encoding="utf-8") as f:
                json.dump(po_data, f, ensure_ascii=False, indent=2)

        # Migrar CPOs
        cpo_year = int(data.get("cpo_year", CURRENT_YEAR))
        cpo_path = Path(CPO_FOLDER) / f"cpo_{cpo_year}.json"
        cpo_updated = 0
        if cpo_path.exists():
            with open(cpo_path, "r", encoding="utf-8") as f:
                cpo_data = json.load(f)
            for rec in cpo_data:
                if str(rec.get("job","")).upper() == source:
                    rec["job"] = target
                    cpo_updated += 1
            with open(cpo_path, "w", encoding="utf-8") as f:
                json.dump(cpo_data, f, ensure_ascii=False, indent=2)

        # Actualizar PT Numbers
        pt_records = pt_load()
        for rec in pt_records:
            if source in rec.get("jobs",[]):
                rec["jobs"] = [target if j==source else j for j in rec["jobs"]]
        pt_save(pt_records)

        # Actualizar SV Numbers
        sv_records = sv_load()
        for rec in sv_records:
            if source in rec.get("jobs",[]):
                rec["jobs"] = [target if j==source else j for j in rec["jobs"]]
        sv_save(sv_records)

        # Eliminar carpeta source
        import shutil as _shutil
        _shutil.rmtree(str(source_folder))

        return jsonify({
            "ok": True,
            "source": source, "target": target,
            "files_moved": moved_files,
            "files_renamed": skipped_files,
            "wh_updated": wh_updated,
            "po_updated": po_updated,
            "cpo_updated": cpo_updated,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/jobs/<job_number>/renumber", methods=["POST"])
def api_renumber_job(job_number):
    """Cambia el número de un job. Solo admins."""
    if not is_admin():
        return jsonify({"error": "Sin permiso — solo administradores"}), 403
    if not JOB_RE.match(job_number):
        return jsonify({"error": "Job number inválido"}), 400
    try:
        data       = request.get_json()
        new_number = str(data.get("new_number","")).strip().upper()
        if not JOB_RE.match(new_number):
            return jsonify({"error": f"Nuevo número '{new_number}' inválido"}), 400
        if new_number == job_number:
            return jsonify({"error": "El nuevo número es igual al actual"}), 400
        if new_number in all_job_numbers():
            return jsonify({"error": f"{new_number} ya existe"}), 409

        old_folder = job_folder(job_number)
        new_folder = job_folder(new_number)
        if not old_folder.exists():
            return jsonify({"error": f"{job_number} no existe"}), 404

        # Renombrar carpeta
        old_folder.rename(new_folder)

        # Actualizar job_info.json
        meta = new_folder / "job_info.json"
        if meta.exists():
            with open(meta, "r", encoding="utf-8") as f:
                ji = json.load(f)
            parts = new_number.split("-")
            ji["job_number"]  = new_number
            ji["main_index"]  = int(parts[0])
            ji["subindex"]    = parts[1]
            ji["updated_at"]  = datetime.datetime.now().isoformat()
            with open(meta, "w", encoding="utf-8") as f:
                json.dump(ji, f, ensure_ascii=False, indent=2)

        # Actualizar PT Numbers
        pt_records = pt_load()
        for rec in pt_records:
            if job_number in rec.get("jobs",[]):
                rec["jobs"] = [new_number if j==job_number else j for j in rec["jobs"]]
        pt_save(pt_records)

        # Actualizar SV Numbers
        sv_records = sv_load()
        for rec in sv_records:
            if job_number in rec.get("jobs",[]):
                rec["jobs"] = [new_number if j==job_number else j for j in rec["jobs"]]
        sv_save(sv_records)

        return jsonify({"ok": True, "old": job_number, "new": new_number})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/wh/clear", methods=["POST"])
def api_clear_wh():
    """Borra todos los registros de WH de un año. Solo admins."""
    if not is_admin():
        return jsonify({"error": "Sin permiso — solo administradores"}), 403
    try:
        data = request.get_json()
        year = int(data.get("year", CURRENT_YEAR))
        wh_save(year, [])
        return jsonify({"ok": True, "year": year, "message": f"Work Hours {year} eliminados"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  SV NUMBER FILES
# ══════════════════════════════════════════════════════════════════
SV_DOCS_BASE = _os.path.join(_DATA, "SV_DOCS")

def sv_folder(sv_number):
    return Path(SV_DOCS_BASE) / sv_number

@app.route("/api/sv/<sv_number>/files", methods=["GET"])
def api_list_sv_files(sv_number):
    folder = sv_folder(sv_number)
    if not folder.exists(): return jsonify([])
    files = []
    for f in sorted(folder.iterdir()):
        if f.is_file():
            st = f.stat()
            files.append({
                "name": f.name, "size": st.st_size,
                "modified": datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
    return jsonify(files)

@app.route("/api/sv/<sv_number>/files", methods=["POST"])
def api_upload_sv_file(sv_number):
    folder = sv_folder(sv_number)
    try: folder.mkdir(parents=True, exist_ok=True)
    except Exception as e: return jsonify({"error": str(e)}), 500
    saved = []
    for f in request.files.getlist("files"):
        dest = folder / f.filename
        f.save(str(dest))
        saved.append({"name": f.filename, "size": dest.stat().st_size})
    return jsonify({"saved": saved})

@app.route("/api/sv/<sv_number>/files/<filename>", methods=["GET"])
def api_download_sv_file(sv_number, filename):
    folder = sv_folder(sv_number)
    if not (folder / filename).exists():
        return jsonify({"error": "Archivo no encontrado"}), 404
    return send_from_directory(str(folder), filename, as_attachment=True)

@app.route("/api/sv/<sv_number>/files/<filename>", methods=["DELETE"])
def api_delete_sv_file(sv_number, filename):
    target = sv_folder(sv_number) / filename
    if target.exists() and target.is_file():
        target.unlink()
        return jsonify({"ok": True})
    return jsonify({"error": "Archivo no encontrado"}), 404


# ══════════════════════════════════════════════════════════════════
#  USER MANAGEMENT
# ══════════════════════════════════════════════════════════════════
@app.route("/api/admin/users/list", methods=["GET"])
def api_admin_list_users():
    if not is_admin(): return jsonify({"error": "Sin permiso"}), 403
    try:
        auth  = _load_auth()
        perms = users_load()
        result = []
        for uname, info in auth.items():
            p = perms.get(uname, {})
            result.append({
                "username":   uname,
                "active":     info.get("active", True),
                "role":       p.get("role", "admin" if uname==ADMIN_USER else "viewer"),
                "created_at": info.get("created_at",""),
                "is_admin_user": uname == ADMIN_USER,
            })
        return jsonify({"users": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/users/create", methods=["POST"])
def api_admin_create_user():
    if not is_admin(): return jsonify({"error": "Sin permiso"}), 403
    try:
        data     = request.get_json()
        username = str(data.get("username","")).strip().lower()
        password = str(data.get("password","")).strip()
        role     = data.get("role","viewer")
        if not username or not password:
            return jsonify({"error": "Usuario y contraseña son requeridos"}), 400
        if len(password) < 6:
            return jsonify({"error": "La contraseña debe tener al menos 6 caracteres"}), 400
        auth = _load_auth()
        if username in auth:
            return jsonify({"error": f"El usuario '{username}' ya existe"}), 409
        auth[username] = {
            "password_hash": _hash(password),
            "active": True,
            "created_at": datetime.datetime.now().isoformat()
        }
        _save_auth(auth)
        # Crear permisos
        perms = users_load()
        perms[username] = {"role": role, "permissions": _default_perms(role)}
        users_save(perms)
        return jsonify({"ok": True, "username": username})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/users/<username>/password", methods=["PUT"])
def api_admin_change_password(username):
    if not is_admin(): return jsonify({"error": "Sin permiso"}), 403
    try:
        data     = request.get_json()
        password = str(data.get("password","")).strip()
        if len(password) < 6:
            return jsonify({"error": "Mínimo 6 caracteres"}), 400
        auth = _load_auth()
        if username not in auth:
            return jsonify({"error": "Usuario no encontrado"}), 404
        auth[username]["password_hash"] = _hash(password)
        _save_auth(auth)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/users/<username>/toggle", methods=["PUT"])
def api_admin_toggle_user(username):
    if not is_admin(): return jsonify({"error": "Sin permiso"}), 403
    if username == ADMIN_USER:
        return jsonify({"error": "No se puede desactivar al super admin"}), 400
    try:
        auth = _load_auth()
        if username not in auth:
            return jsonify({"error": "Usuario no encontrado"}), 404
        auth[username]["active"] = not auth[username].get("active", True)
        _save_auth(auth)
        return jsonify({"ok": True, "active": auth[username]["active"]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/users/<username>", methods=["DELETE"])
def api_admin_delete_user(username):
    if not is_admin(): return jsonify({"error": "Sin permiso"}), 403
    if username == ADMIN_USER:
        return jsonify({"error": "No se puede eliminar al super admin"}), 400
    try:
        auth = _load_auth()
        if username in auth: del auth[username]
        _save_auth(auth)
        perms = users_load()
        if username in perms: del perms[username]
        users_save(perms)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ══════════════════════════════════════════════════════════════════
#  BACKUPS / RESPALDOS
# ══════════════════════════════════════════════════════════════════
import csv as _csv

def _db_to_rows(name, year=None):
    """Devuelve (headers, rows) para cada base de datos."""
    y = year or CURRENT_YEAR
    if name == "quotes":
        data = _load_quotes()
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[r.get(h,"") for h in headers] for r in data]
    elif name == "jobs":
        data = scan_jobs()
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[r.get(h,"") for h in headers] for r in data]
    elif name == "rates":
        p = rates_root() / f"rates_{y}.json"
        data = json.load(open(p)) if p.exists() else []
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[r.get(h,"") for h in headers] for r in data]
    elif name == "wh":
        data = wh_load(y)
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[r.get(h,"") for h in headers] for r in data]
    elif name == "po":
        data = po_load(y)
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[r.get(h,"") for h in headers] for r in data]
    elif name == "ivp":
        data = ivp_load(y)
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[r.get(h,"") for h in headers] for r in data]
    elif name == "cpo":
        data = cpo_load(y)
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[r.get(h,"") for h in headers] for r in data]
    elif name == "fx":
        data = fx_load(y)
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[r.get(h,"") for h in headers] for r in data]
    elif name == "pt":
        data = pt_load()
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[str(r.get(h,"")) for h in headers] for r in data]
    elif name == "sv":
        data = sv_load()
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[str(r.get(h,"")) for h in headers] for r in data]
    elif name == "stock":
        data = stock_load()
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[str(r.get(h,"")) for h in headers] for r in data]
    elif name == "reassign":
        orders = reassign_load()
        # Flatten orders to items
        rows = []
        for o in orders:
            for item in o.get("items",[]):
                rows.append({
                    "order_number": o["order_number"],
                    "created_at":   o.get("created_at",""),
                    **item
                })
        if not rows: return [], []
        headers = list(rows[0].keys())
        return headers, [[str(r.get(h,"")) for h in headers] for r in rows]
    elif name == "recovery":
        data = recovery_load()
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[str(r.get(h,"")) for h in headers] for r in data]
    elif name == "proveedores":
        data = prov_load()
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[str(r.get(h,"")) for h in headers] for r in data]
    elif name in ("catalogo_electrico","catalogo_mecanico","catalogo_servicios"):
        tipo_map = {"catalogo_electrico":"electrico","catalogo_mecanico":"mecanico","catalogo_servicios":"servicios"}
        data = cat_load(tipo_map[name])
        if not data: return [], []
        headers = list(data[0].keys())
        return headers, [[str(r.get(h,"")) for h in headers] for r in data]
    return [], []

@app.route("/api/backup/<name>", methods=["GET"])
def api_backup(name):
    fmt  = request.args.get("fmt","xlsx")
    year = request.args.get("year", str(CURRENT_YEAR))
    try: year = int(year)
    except: year = CURRENT_YEAR
    try:
        headers, rows = _db_to_rows(name, year)
        if not headers:
            return jsonify({"error": "Sin datos para respaldar"}), 404
        fname = f"{name}_{year}_backup"
        if fmt == "csv":
            import io as _io
            buf = _io.StringIO()
            w = _csv.writer(buf)
            w.writerow(headers)
            w.writerows(rows)
            buf.seek(0)
            return Response(buf.getvalue(), mimetype="text/csv",
                headers={"Content-Disposition": f"attachment;filename={fname}.csv"})
        else:  # xlsx
            import io as _io
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = name[:31]
            ws.append(headers)
            for row in rows:
                ws.append([str(v) if isinstance(v, list) else v for v in row])
            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(buf.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment;filename={fname}.xlsx"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/backup/all", methods=["GET"])
def api_backup_all():
    if not is_admin(): return jsonify({"error": "Sin permiso"}), 403
    fmt  = request.args.get("fmt","xlsx")
    year = request.args.get("year", str(CURRENT_YEAR))
    try: year = int(year)
    except: year = CURRENT_YEAR
    import io as _io, zipfile as _zip
    buf = _io.BytesIO()
    dbs = ["quotes","jobs","rates","wh","po","ivp","cpo","fx","pt","sv","stock","reassign","recovery","proveedores","catalogo_electrico","catalogo_mecanico","catalogo_servicios"]
    with _zip.ZipFile(buf, "w", _zip.ZIP_DEFLATED) as zf:
        for name in dbs:
            try:
                headers, rows = _db_to_rows(name, year)
                if not headers: continue
                if fmt == "csv":
                    sb = _io.StringIO()
                    w = _csv.writer(sb)
                    w.writerow(headers)
                    w.writerows(rows)
                    zf.writestr(f"{name}_{year}.csv", sb.getvalue())
                else:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = name[:31]
                    ws.append(headers)
                    for row in rows:
                        ws.append([str(v) if isinstance(v,list) else v for v in row])
                    xb = _io.BytesIO()
                    wb.save(xb)
                    zf.writestr(f"{name}_{year}.xlsx", xb.getvalue())
            except: pass
    buf.seek(0)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    return Response(buf.getvalue(), mimetype="application/zip",
        headers={"Content-Disposition": f"attachment;filename=backup_completo_{ts}.zip"})

# ══════════════════════════════════════════════════════════════════
#  USER PREFERENCES — Idioma
# ══════════════════════════════════════════════════════════════════
@app.route("/api/me/lang", methods=["GET"])
def api_get_lang():
    return jsonify({"lang": session.get("lang", "es")})

@app.route("/api/me/lang", methods=["POST"])
def api_set_lang():
    data = request.get_json()
    lang = data.get("lang", "es")
    if lang not in ("es", "en", "it"):
        return jsonify({"error": "Idioma no válido"}), 400
    session["lang"] = lang
    session.modified = True
    return jsonify({"ok": True, "lang": lang})


# ══════════════════════════════════════════════════════════════════
#  STOCK & REASIGNACION
# ══════════════════════════════════════════════════════════════════
STOCK_FILE    = _os.path.join(_DATA, "stock.json")
REASSIGN_FILE  = _os.path.join(_DATA, "reassign_orders.json")
RECOVERY_FILE  = _os.path.join(_DATA, "recovery.json")

def recovery_load():
    _h = _cache_get("recovery")
    if _h is not None: return _h
    p = Path(RECOVERY_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def recovery_save(records):
    Path(RECOVERY_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(RECOVERY_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("recovery", records)

def stock_load():
    _h = _cache_get("stock")
    if _h is not None: return _h
    p = Path(STOCK_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def stock_save(records):
    Path(STOCK_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(STOCK_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("stock", records)

def reassign_load():
    _h = _cache_get("reassign")
    if _h is not None: return _h
    p = Path(REASSIGN_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def reassign_save(records):
    Path(REASSIGN_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(REASSIGN_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("reassign", records)

def reassign_next_number():
    return _doc_next_number("RA")

@app.route("/api/stock", methods=["GET"])
def api_get_stock():
    try:
        q = request.args.get("q","").lower()
        records = stock_load()
        if q:
            records = [r for r in records if
                q in (r.get("part_number","")).lower() or
                q in (r.get("manufacturer","")).lower() or
                q in (r.get("description","")).lower() or
                q in (r.get("label_code","")).lower()]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/stock", methods=["POST"])
def api_create_stock():
    try:
        data = request.get_json()
        with lock:
            records      = stock_load()
            mfr          = str(data.get("manufacturer","")).strip().upper()
            pnum         = str(data.get("part_number","")).strip().upper()
            new_qty      = int(data.get("new_quantity", 0))   # nuevos ingresos
            new_cost     = float(data.get("last_cost", 0))
            recovery_job = str(data.get("recovery_job","")).strip()
            unit         = str(data.get("unit","Pieza"))
            description  = str(data.get("description","")).strip()
            section      = str(data.get("section","")).strip()
            box          = str(data.get("box","")).strip()
            label_code   = str(data.get("label_code","")).strip().upper()

            existing = next((r for r in records
                if r.get("manufacturer","").upper()==mfr
                and r.get("part_number","").upper()==pnum), None)

            if existing:
                prev_qty           = int(existing.get("quantity", 0))
                existing["quantity"]   = prev_qty + new_qty
                existing["last_cost"]  = new_cost if new_cost > 0 else existing["last_cost"]
                existing["updated_at"] = datetime.datetime.now().isoformat()
                if section: existing["section"] = section
                if box:     existing["box"]     = box
                if recovery_job: existing["recovery_job"] = recovery_job
                if label_code: existing["label_code"] = label_code
                stock_save(records)
                # Recovery entry only for new ingress
                if new_qty > 0 and recovery_job:
                    cost  = new_cost if new_cost > 0 else float(existing.get("last_cost",0))
                    total = round(new_qty * cost * -1, 2)
                    rec_recovery = {
                        "id":           f"RCV-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                        "manufacturer": mfr, "part_number": pnum,
                        "description":  existing.get("description",""),
                        "last_cost":    cost, "quantity": new_qty,
                        "unit":         existing.get("unit", unit),
                        "section":      existing.get("section",""),
                        "box":          existing.get("box",""),
                        "label_code":   existing.get("label_code",""),
                        "job":          recovery_job,
                        "total_value":  total,
                        "stock_id":     existing["id"],
                        "created_at":   datetime.datetime.now().isoformat(),
                    }
                    recoveries = recovery_load()
                    recoveries.append(rec_recovery)
                    recovery_save(recoveries)
                return jsonify({"ok": True, "action":"updated", "record": existing,
                                "new_qty": new_qty, "prev_qty": prev_qty})
            else:
                rec = {
                    "id":           f"STK-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                    "manufacturer": mfr, "part_number": pnum,
                    "description":  description,
                    "last_cost":    new_cost,
                    "quantity":     new_qty,
                    "unit":         unit,
                    "section":      section,
                    "box":          box,
                    "label_code":   label_code,
                    "recovery_job": recovery_job,
                    "created_at":   datetime.datetime.now().isoformat(),
                }
                records.append(rec)
                stock_save(records)
                # Recovery for first ingress
                if new_qty > 0 and recovery_job:
                    total = round(new_qty * new_cost * -1, 2)
                    rec_recovery = {
                        "id":           f"RCV-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                        "manufacturer": mfr, "part_number": pnum,
                        "description":  description,
                        "last_cost":    new_cost, "quantity": new_qty,
                        "unit":         unit, "section": section, "box": box,
                        "job":          recovery_job,
                        "total_value":  total,
                        "stock_id":     rec["id"],
                        "created_at":   datetime.datetime.now().isoformat(),
                    }
                    recoveries = recovery_load()
                    recoveries.append(rec_recovery)
                    recovery_save(recoveries)
                return jsonify({"ok": True, "action":"created", "record": rec,
                                "new_qty": new_qty})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/stock/<item_id>", methods=["PUT"])
def api_update_stock(item_id):
    try:
        data = request.get_json()
        with lock:
            records = stock_load()
            rec = next((r for r in records if r.get("id")==item_id), None)
            if not rec: return jsonify({"error":"Item no encontrado"}), 404
            for k in ["manufacturer","part_number","description","unit","section","box","recovery_job","label_code"]:
                if k in data: rec[k] = str(data[k]).strip()
            rec["label_code"] = rec.get("label_code","").upper()
            if "last_cost" in data: rec["last_cost"] = float(data["last_cost"])
            if "quantity"  in data: rec["quantity"]  = int(data["quantity"])
            rec["updated_at"] = datetime.datetime.now().isoformat()
            stock_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# (delete stock moved to admin endpoint below)

@app.route("/api/stock/import", methods=["POST"])
def api_import_stock():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error":"No se recibio archivo"}), 400
        mode = request.form.get("mode","append")
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in list(ws.iter_rows(min_row=1,max_row=1))[0]:
            if cell.value:
                headers[str(cell.value).strip().upper()] = cell.column-1
        def col(*aliases):
            for a in aliases:
                if a.upper() in headers: return headers[a.upper()]
        return None
        ci_mfr  = col("FABRICANTE","MANUFACTURER","MARCA")
        ci_pnum = col("NUMERO DE PARTE","PART NUMBER","PART_NUMBER","NO. PARTE")
        ci_desc = col("DESCRIPCION","DESCRIPTION","DESC")
        ci_cost = col("ULTIMO COSTO","LAST COST","COSTO","COST")
        ci_qty  = col("EXISTENCIA","QUANTITY","CANTIDAD","QTY")
        ci_unit = col("UNIDAD","UNIT")
        ci_sec  = col("SECCION","SECTION")
        ci_box  = col("CAJA","BOX")
        ci_rec  = col("RECUPERACION","RECOVERY","RECOVERY_JOB")
        ci_label = col("ETIQUETA","LABEL","QR","CODIGO DE BARRAS","BARCODE","COD. ETIQUETA")
        imported = 0
        with lock:
            records = stock_load() if mode=="append" else []
            for row in ws.iter_rows(min_row=2, values_only=True):
                pnum = str(row[ci_pnum]).strip().upper() if ci_pnum is not None and row[ci_pnum] else ""
                if not pnum or pnum in ("NONE","","#N/A"): continue
                mfr = str(row[ci_mfr]).strip().upper() if ci_mfr is not None and row[ci_mfr] else ""
                existing = next((r for r in records
                    if r.get("part_number","")==pnum and r.get("manufacturer","")==mfr), None)
                try: cost = float(row[ci_cost]) if ci_cost is not None and row[ci_cost] else 0.0
                except: cost = 0.0
                try: qty = int(float(str(row[ci_qty]))) if ci_qty is not None and row[ci_qty] else 0
                except: qty = 0
                label = str(row[ci_label]).strip().upper() if ci_label is not None and row[ci_label] else ""
                if existing:
                    existing["quantity"]   = qty
                    existing["last_cost"]  = cost
                    if label: existing["label_code"] = label
                    existing["updated_at"] = datetime.datetime.now().isoformat()
                else:
                    records.append({
                        "id":           f"STK-imp-{imported}",
                        "manufacturer": mfr, "part_number": pnum,
                        "description":  str(row[ci_desc]).strip() if ci_desc is not None and row[ci_desc] else "",
                        "last_cost": cost, "quantity": qty,
                        "unit":         str(row[ci_unit]).strip() if ci_unit is not None and row[ci_unit] else "Pieza",
                        "section":      str(row[ci_sec]).strip() if ci_sec is not None and row[ci_sec] else "",
                        "box":          str(row[ci_box]).strip() if ci_box is not None and row[ci_box] else "",
                        "recovery_job": str(row[ci_rec]).strip() if ci_rec is not None and row[ci_rec] else "",
                        "label_code":   label,
                        "created_at":   datetime.datetime.now().isoformat(),
                    })
                imported += 1
            stock_save(records)
        return jsonify({"ok":True,"imported":imported,"total":len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reassign", methods=["GET"])
def api_get_reassign():
    try:
        job = request.args.get("job","").upper()
        orders = reassign_load()
        if job:
            orders = [o for o in orders if any(
                i.get("job","").upper()==job for i in o.get("items",[]))]
        return jsonify({"orders": orders, "total": len(orders),
                        "next_number": reassign_next_number()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reassign", methods=["POST"])
def api_create_reassign():
    try:
        data = request.get_json()
        order_number = str(data.get("order_number","")).strip().upper()
        is_new = data.get("is_new", True)
        items  = data.get("items", [])
        if not items: return jsonify({"error":"Sin items"}), 400
        with lock:
            orders  = reassign_load()
            records = stock_load()
            if is_new:
                if not order_number:
                    order_number = reassign_next_number()
                if any(o["order_number"]==order_number for o in orders):
                    return jsonify({"error":f"{order_number} ya existe"}), 409
                order = {"order_number": order_number,
                         "created_at": datetime.datetime.now().isoformat(), "items": []}
                orders.append(order)
            else:
                order = next((o for o in orders if o["order_number"]==order_number), None)
                if not order: return jsonify({"error":"Orden no encontrada"}), 404
            for item in items:
                pnum = str(item.get("part_number","")).strip().upper()
                mfr  = str(item.get("manufacturer","")).strip().upper()
                qty  = int(item.get("quantity",0))
                cost = float(item.get("unit_cost",0))
                stk  = next((r for r in records
                    if r.get("part_number","")==pnum and r.get("manufacturer","")==mfr), None)
                if stk:
                    stk["quantity"]   = max(0, stk["quantity"] - qty)
                    stk["updated_at"] = datetime.datetime.now().isoformat()
                order["items"].append({
                    "part_number": pnum, "manufacturer": mfr,
                    "description": str(item.get("description","")).strip(),
                    "label_code":  str(item.get("label_code") or (stk.get("label_code","") if stk else "")).strip().upper(),
                    "job":         str(item.get("job","")).strip().upper(),
                    "unit_cost":   cost, "quantity": qty,
                    "total_cost":  round(cost*qty, 2),
                    "added_at":    datetime.datetime.now().isoformat(),
                })
            order["updated_at"] = datetime.datetime.now().isoformat()
            reassign_save(orders)
            stock_save(records)
        return jsonify({"ok":True,"order_number":order_number,"order":order})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reassign/<job_number>/total")
def api_reassign_total(job_number):
    try:
        orders = reassign_load()
        total, matched = 0.0, []
        for o in orders:
            for item in o.get("items",[]):
                if item.get("job","").upper() == job_number.upper():
                    total += float(item.get("total_cost",0))
                    matched.append({"order_number": o["order_number"], **item})
        return jsonify({"job": job_number, "total": round(total,2), "items": matched})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ── Recovery routes
@app.route("/api/recovery", methods=["GET"])
def api_get_recovery():
    try:
        job = request.args.get("job","").upper()
        records = recovery_load()
        if job:
            records = [r for r in records if r.get("job","").upper()==job]
        q = request.args.get("q","").lower()
        if q:
            records = [r for r in records if
                q in (r.get("part_number","")).lower() or
                q in (r.get("manufacturer","")).lower() or
                q in (r.get("description","")).lower() or
                q in (r.get("label_code","")).lower()]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/recovery/<rec_id>", methods=["DELETE"])
def api_delete_recovery(rec_id):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            records = recovery_load()
            new = [r for r in records if r.get("id")!=rec_id]
            if len(new)==len(records): return jsonify({"error":"No encontrado"}), 404
            recovery_save(new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/stock/<item_id>", methods=["DELETE"])
def api_delete_stock_admin(item_id):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            records = stock_load()
            new = [r for r in records if r.get("id")!=item_id]
            if len(new)==len(records): return jsonify({"error":"No encontrado"}), 404
            stock_save(new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reassign/order/<order_number>", methods=["DELETE"])
def api_delete_reassign_order(order_number):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            orders = reassign_load()
            new = [o for o in orders if o.get("order_number")!=order_number.upper()]
            if len(new)==len(orders): return jsonify({"error":"Orden no encontrada"}), 404
            reassign_save(new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reassign/order/<order_number>/pdf")
def api_reassign_pdf(order_number):
    """Genera PDF de la orden de reasignación."""
    try:
        orders = reassign_load()
        order  = next((o for o in orders if o["order_number"]==order_number.upper()), None)
        if not order: return jsonify({"error":"Orden no encontrada"}), 404
        import io as _io
        html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
        <style>
          body{{font-family:Arial,sans-serif;font-size:11px;color:#222;margin:30px}}
          h1{{font-size:18px;color:#c8102e;margin-bottom:4px}}
          .sub{{font-size:11px;color:#666;margin-bottom:20px}}
          table{{width:100%;border-collapse:collapse;margin-top:16px}}
          th{{background:#1f3864;color:#fff;padding:7px 10px;text-align:left;font-size:10px;text-transform:uppercase}}
          td{{padding:7px 10px;border-bottom:1px solid #e0e0e0;font-size:11px}}
          tr:nth-child(even){{background:#f5f5f5}}
          .total{{text-align:right;font-weight:bold;font-size:13px;color:#1a7a1a;margin-top:12px}}
          .footer{{margin-top:30px;font-size:10px;color:#999;border-top:1px solid #ddd;padding-top:8px}}
        </style></head><body>
        <h1>Orden de Reasignación: {order['order_number']}</h1>
        <div class="sub">Fecha: {order.get('created_at','')[:10]} &nbsp;|&nbsp; Persico México</div>
        <table>
          <tr><th>No. Parte</th><th>Fabricante</th><th>Descripción</th><th>Job</th><th style="text-align:right">Cant.</th><th style="text-align:right">Costo Unit.</th><th style="text-align:right">Total USD</th></tr>
        """
        total = 0.0
        for item in order.get("items",[]):
            t = float(item.get("total_cost",0))
            total += t
            html += f"""<tr>
              <td>{item.get('part_number','')}</td>
              <td>{item.get('manufacturer','')}</td>
              <td>{item.get('description','')}</td>
              <td>{item.get('job','')}</td>
              <td style="text-align:right">{item.get('quantity',0)}</td>
              <td style="text-align:right">${item.get('unit_cost',0):,.2f}</td>
              <td style="text-align:right">${t:,.2f}</td>
            </tr>"""
        html += f"""</table>
        <div class="total">Total: ${total:,.2f} USD</div>
        <div class="footer">Generado por Persico Suite · {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
        </body></html>"""
        return Response(html, mimetype="text/html",
            headers={"Content-Disposition": f"inline;filename={order_number}.html"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ══════════════════════════════════════════════════════════════════
#  PROVEEDORES
# ══════════════════════════════════════════════════════════════════
PROV_FILE     = _os.path.join(_DATA, "proveedores.json")
PROV_DOCS_DIR = _os.path.join(_DATA, "PROV_DOCS")

def prov_load():
    p = Path(PROV_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def prov_save(records):
    Path(PROV_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(PROV_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def prov_docs_folder(clave):
    p = Path(PROV_DOCS_DIR) / str(clave)
    p.mkdir(parents=True, exist_ok=True)
    return p

@app.route("/api/proveedores", methods=["GET"])
def api_get_proveedores():
    try:
        q = request.args.get("q","").lower()
        records = prov_load()
        if q:
            records = [r for r in records if
                q in (r.get("nombre","")).lower() or
                q in (r.get("rfc","")).lower() or
                q in str(r.get("clave","")).lower()]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/proveedores", methods=["POST"])
def api_create_proveedor():
    try:
        data = request.get_json()
        with lock:
            records = prov_load()
            # Auto clave
            clave = data.get("clave")
            if not clave:
                existing_claves = [int(r.get("clave",0)) for r in records if str(r.get("clave","")).isdigit()]
                clave = max(existing_claves)+1 if existing_claves else 1
            clave = int(clave)
            if any(int(r.get("clave",0))==clave for r in records):
                return jsonify({"error": f"Clave {clave} ya existe"}), 409
            rec = {
                "clave":             clave,
                "estatus":           str(data.get("estatus","Activo")),
                "nombre":            str(data.get("nombre","")).strip().upper(),
                "rfc":               str(data.get("rfc","")).strip().upper(),
                "calle":             str(data.get("calle","")).strip(),
                "num_interior":      str(data.get("num_interior","")).strip(),
                "num_exterior":      str(data.get("num_exterior","")).strip(),
                "telefono":          str(data.get("telefono","")).strip(),
                "clasificacion":     str(data.get("clasificacion","")).strip(),
                "forma_pago":        str(data.get("forma_pago","Transferencia")),
                "terminos_pago":     str(data.get("terminos_pago","")).strip(),
                "moneda":            str(data.get("moneda","MXN")),
                "fecha_ultima_compra": str(data.get("fecha_ultima_compra","")).strip(),
                "created_at":        datetime.datetime.now().isoformat(),
            }
            records.append(rec)
            prov_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/proveedores/<int:clave>", methods=["PUT"])
def api_update_proveedor(clave):
    try:
        data = request.get_json()
        with lock:
            records = prov_load()
            rec = next((r for r in records if int(r.get("clave",0))==clave), None)
            if not rec: return jsonify({"error":"Proveedor no encontrado"}), 404
            for k in ["estatus","nombre","rfc","calle","num_interior","num_exterior",
                      "telefono","clasificacion","forma_pago","terminos_pago","moneda","fecha_ultima_compra"]:
                if k in data: rec[k] = str(data[k]).strip()
            rec["nombre"] = rec["nombre"].upper()
            rec["rfc"]    = rec["rfc"].upper()
            rec["updated_at"] = datetime.datetime.now().isoformat()
            prov_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/proveedores/<int:clave>", methods=["DELETE"])
def api_delete_proveedor(clave):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            records = prov_load()
            new = [r for r in records if int(r.get("clave",0))!=clave]
            if len(new)==len(records): return jsonify({"error":"No encontrado"}), 404
            prov_save(new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/proveedores/import", methods=["POST"])
def api_import_proveedores():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error":"No se recibió archivo"}), 400
        mode = request.form.get("mode","append")
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in list(ws.iter_rows(min_row=1,max_row=1))[0]:
            if cell.value:
                headers[str(cell.value).strip().upper()] = cell.column-1
        def col(*aliases):
            for a in aliases:
                if a.upper() in headers: return headers[a.upper()]
            return None
        ci_clave   = col("CLAVE","ID")
        ci_est     = col("ESTATUS","STATUS")
        ci_nombre  = col("NOMBRE","NAME","PROVEEDOR")
        ci_rfc     = col("RFC")
        ci_calle   = col("CALLE","DIRECCIÓN","DIRECCION","ADDRESS")
        ci_ni      = col("NÚMERO INTERIOR","NUM_INTERIOR","N. INT")
        ci_ne      = col("NÚMERO EXTERIOR","NUM_EXTERIOR","N. EXT")
        ci_tel     = col("TELÉFONO","TELEFONO","TEL","PHONE")
        ci_clas    = col("CLASIFICACIÓN","CLASIFICACION","CATEGORY")
        ci_fult    = col("FECHA DE ÚLTIMA COMPRA","FECHA ULTIMA COMPRA","LAST PURCHASE")
        ci_fpago   = col("FORMA DE PAGO","PAYMENT METHOD","PAYMENT")
        ci_tpago   = col("TÉRMINOS DE PAGO","TERMINOS DE PAGO","PAYMENT TERMS")
        ci_moneda  = col("MONEDA","CURRENCY")
        if ci_nombre is None:
            return jsonify({"error":"No se encontró columna NOMBRE"}), 400
        imported = 0
        with lock:
            records = prov_load() if mode=="append" else []
            existing_claves = {int(r.get("clave",0)) for r in records}
            max_clave = max(existing_claves) if existing_claves else 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                nombre = str(row[ci_nombre]).strip().upper() if ci_nombre is not None and row[ci_nombre] else ""
                if not nombre or nombre in ("NONE","","#N/A"): continue
                if ci_clave is not None and row[ci_clave]:
                    try: clave = int(float(str(row[ci_clave])))
                    except: max_clave+=1; clave=max_clave
                else:
                    max_clave+=1; clave=max_clave
                if clave in existing_claves: continue
                existing_claves.add(clave)
                def g(ci): return str(row[ci]).strip() if ci is not None and row[ci] and str(row[ci]).strip() not in ("None","nan","#N/A") else ""
                fult = g(ci_fult)
                if fult and len(fult)>10: fult=fult[:10]
                records.append({
                    "clave": clave, "estatus": g(ci_est) or "Activo",
                    "nombre": nombre, "rfc": g(ci_rfc).upper(),
                    "calle": g(ci_calle), "num_interior": g(ci_ni), "num_exterior": g(ci_ne),
                    "telefono": g(ci_tel), "clasificacion": g(ci_clas),
                    "fecha_ultima_compra": fult,
                    "forma_pago": g(ci_fpago) or "Transferencia",
                    "terminos_pago": g(ci_tpago), "moneda": g(ci_moneda) or "MXN",
                    "created_at": datetime.datetime.now().isoformat(),
                })
                imported += 1
            prov_save(records)
        return jsonify({"ok":True,"imported":imported,"total":len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── Proveedor document routes
@app.route("/api/proveedores/<int:clave>/files", methods=["GET"])
def api_list_prov_files(clave):
    folder = prov_docs_folder(clave)
    files = [{"name":f.name,"size":f.stat().st_size,
              "modified":datetime.datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M")}
             for f in sorted(folder.iterdir()) if f.is_file()]
    return jsonify(files)

@app.route("/api/proveedores/<int:clave>/files", methods=["POST"])
def api_upload_prov_file(clave):
    folder = prov_docs_folder(clave)
    saved = []
    for f in request.files.getlist("files"):
        dest = folder / f.filename
        f.save(str(dest))
        saved.append({"name":f.filename,"size":dest.stat().st_size})
    return jsonify({"saved":saved})

@app.route("/api/proveedores/<int:clave>/files/<filename>", methods=["GET"])
def api_download_prov_file(clave, filename):
    folder = prov_docs_folder(clave)
    if not (folder/filename).exists(): return jsonify({"error":"No encontrado"}), 404
    return send_from_directory(str(folder), filename, as_attachment=True)

@app.route("/api/proveedores/<int:clave>/files/<filename>", methods=["DELETE"])
def api_delete_prov_file(clave, filename):
    target = prov_docs_folder(clave) / filename
    if target.exists(): target.unlink()
    return jsonify({"ok":True})



# ══════════════════════════════════════════════════════════════════
#  CATÁLOGOS (ELÉCTRICO / MECÁNICO / SERVICIOS)
# ══════════════════════════════════════════════════════════════════
CAT_CONFIG = {
    "electrico":  {"file": "catalogo_electrico.json",  "prefix": "CE"},
    "mecanico":   {"file": "catalogo_mecanico.json",   "prefix": "CM"},
    "servicios":  {"file": "catalogo_servicios.json",  "prefix": "CS"},
}

def cat_path(tipo):
    cfg = CAT_CONFIG.get(tipo)
    if not cfg: return None
    return _os.path.join(_DATA, cfg["file"])

def cat_load(tipo):
    p = cat_path(tipo)
    if not p: return []
    pp = Path(p)
    if pp.exists():
        try:
            with open(pp,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def cat_save(tipo, records):
    p = cat_path(tipo)
    if not p: return
    Path(p).parent.mkdir(parents=True, exist_ok=True)
    with open(p,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def cat_next_code(tipo):
    cfg = CAT_CONFIG.get(tipo)
    if not cfg: return None
    prefix = cfg["prefix"]
    records = cat_load(tipo)
    nums = []
    for r in records:
        code = str(r.get("code",""))
        if code.startswith(prefix+"-"):
            try: nums.append(int(code.replace(prefix+"-","")))
            except: pass
    n = (max(nums)+1) if nums else 1
    return f"{prefix}-{n:05d}"

@app.route("/api/catalogo/<tipo>", methods=["GET"])
def api_get_catalogo(tipo):
    if tipo not in CAT_CONFIG: return jsonify({"error":"Tipo de catálogo inválido"}), 400
    try:
        q = request.args.get("q","").lower()
        records = cat_load(tipo)
        if q:
            records = [r for r in records if
                q in (r.get("brand","")).lower() or
                q in (r.get("part_number","")).lower() or
                q in (r.get("description","")).lower() or
                q in (r.get("code","")).lower() or
                q in (r.get("label_code","")).lower()]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/catalogo/<tipo>", methods=["POST"])
def api_create_catalogo_item(tipo):
    if tipo not in CAT_CONFIG: return jsonify({"error":"Tipo de catálogo inválido"}), 400
    try:
        data = request.get_json()
        with lock:
            records = cat_load(tipo)
            code = str(data.get("code","")).strip().upper()
            if not code:
                code = cat_next_code(tipo)
            if any(r.get("code","")==code for r in records):
                return jsonify({"error": f"{code} ya existe"}), 409
            rec = {
                "code":         code,
                "brand":        str(data.get("brand","")).strip().upper(),
                "part_number":  str(data.get("part_number","")).strip().upper(),
                "description":  str(data.get("description","")).strip(),
                "last_price":   float(data.get("last_price",0) or 0),
                "label_code":   str(data.get("label_code","")).strip().upper(),
                "created_at":   datetime.datetime.now().isoformat(),
            }
            records.append(rec)
            cat_save(tipo, records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/catalogo/<tipo>/<code>", methods=["PUT"])
def api_update_catalogo_item(tipo, code):
    if tipo not in CAT_CONFIG: return jsonify({"error":"Tipo de catálogo inválido"}), 400
    try:
        data = request.get_json()
        with lock:
            records = cat_load(tipo)
            rec = next((r for r in records if r.get("code","")==code.upper()), None)
            if not rec: return jsonify({"error":"Item no encontrado"}), 404
            for k in ["brand","part_number","description","label_code"]:
                if k in data: rec[k] = str(data[k]).strip()
            rec["brand"] = rec.get("brand","").upper()
            rec["part_number"] = rec.get("part_number","").upper()
            rec["label_code"] = rec.get("label_code","").upper()
            if "last_price" in data:
                rec["last_price"] = float(data["last_price"] or 0)
            rec["updated_at"] = datetime.datetime.now().isoformat()
            cat_save(tipo, records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/catalogo/<tipo>/<code>", methods=["DELETE"])
def api_delete_catalogo_item(tipo, code):
    if tipo not in CAT_CONFIG: return jsonify({"error":"Tipo de catálogo inválido"}), 400
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            records = cat_load(tipo)
            new = [r for r in records if r.get("code","")!=code.upper()]
            if len(new)==len(records): return jsonify({"error":"No encontrado"}), 404
            cat_save(tipo, new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/catalogo/<tipo>/import", methods=["POST"])
def api_import_catalogo(tipo):
    if tipo not in CAT_CONFIG: return jsonify({"error":"Tipo de catálogo inválido"}), 400
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error":"No se recibió archivo"}), 400
        mode = request.form.get("mode","append")
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in list(ws.iter_rows(min_row=1,max_row=1))[0]:
            if cell.value:
                headers[str(cell.value).strip().upper()] = cell.column-1
        def col(*aliases):
            for a in aliases:
                if a.upper() in headers: return headers[a.upper()]
            return None
        ci_code  = col("CE-CODE","CM-CODE","CS-CODE","CODE","CODIGO")
        ci_brand = col("BRAND","MARCA","FABRICANTE")
        ci_pnum  = col("PART NUMBER","NUMERO DE PARTE","NO. PARTE")
        ci_desc  = col("DESCRIPTION","DESCRIPCION")
        ci_price = col("ULTIMO PRECIO","LAST PRICE","PRECIO")
        ci_label = col("ETIQUETA","LABEL","QR","CODIGO DE BARRAS","BARCODE","COD. ETIQUETA")
        if ci_pnum is None:
            return jsonify({"error":"No se encontró columna PART NUMBER"}), 400
        imported = 0
        with lock:
            records = cat_load(tipo) if mode=="append" else []
            existing_codes = {r.get("code","") for r in records}
            for row in ws.iter_rows(min_row=2, values_only=True):
                pn = str(row[ci_pnum]).strip().upper() if ci_pnum is not None and row[ci_pnum] else ""
                if not pn or pn in ("NONE","","#N/A"): continue
                code = str(row[ci_code]).strip().upper() if ci_code is not None and row[ci_code] else ""
                if not code or code in existing_codes:
                    code = cat_next_code(tipo)
                    # update local counter manually for batch import
                    prefix = CAT_CONFIG[tipo]["prefix"]
                    n = int(code.replace(prefix+"-","")) + 1
                existing_codes.add(code)
                try: price = float(row[ci_price]) if ci_price is not None and row[ci_price] else 0.0
                except: price = 0.0
                records.append({
                    "code": code,
                    "brand": str(row[ci_brand]).strip().upper() if ci_brand is not None and row[ci_brand] else "",
                    "part_number": pn,
                    "description": str(row[ci_desc]).strip() if ci_desc is not None and row[ci_desc] else "",
                    "last_price": price,
                    "label_code": str(row[ci_label]).strip().upper() if ci_label is not None and row[ci_label] else "",
                    "created_at": datetime.datetime.now().isoformat(),
                })
                imported += 1
            cat_save(tipo, records)
        return jsonify({"ok":True,"imported":imported,"total":len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════
#  GENERATED PURCHASE ORDERS  (PO-000000001 sequence)
# ══════════════════════════════════════════════════════════════════
GPO_FILE = _os.path.join(_DATA, "generated_pos.json")

def gpo_load():
    _h = _cache_get("gpo")
    if _h is not None: return _h
    p = Path(GPO_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def gpo_save(records):
    Path(GPO_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(GPO_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("gpo", records)

def gpo_next_number():
    records = gpo_load()
    used = set()
    for r in records:
        n = r.get("po_number","")
        if n.startswith("PO-"):
            try: used.add(int(n.replace("PO-","")))
            except: pass
    n = 1
    while n in used: n += 1
    return f"PO-{n:09d}"

@app.route("/api/gpo", methods=["GET"])
def api_get_gpo():
    try:
        q = request.args.get("q","").lower()
        records = gpo_load()
        if q:
            records = [r for r in records if
                q in (r.get("po_number","")).lower() or
                q in (r.get("supplier_name","")).lower() or
                q in (r.get("pt_sv","")).lower() or
                q in (r.get("job_type","")).lower() or
                q in (r.get("job","")).lower()]
        # Compute display_total: respect effective_total for cancelled/closed
        for r in records:
            st = r.get("status","")
            if st == "Cancelada":
                r["display_total"] = 0.0
            elif "effective_total" in r:
                r["display_total"] = r["effective_total"]
            else:
                r["display_total"] = r.get("total", r.get("subtotal",0))
        return jsonify({"records": records, "total": len(records),
                        "next_number": gpo_next_number()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/gpo", methods=["POST"])
def api_create_gpo():
    try:
        data = request.get_json()
        supplier  = data.get("supplier", {})
        pt_sv     = str(data.get("pt_sv","")).strip().upper()
        cpo       = str(data.get("cpo","")).strip().upper()
        job_type  = str(data.get("job_type","Unico")).strip()
        job_main  = str(data.get("job","")).strip().upper()
        items     = data.get("items", [])
        notes     = str(data.get("notes","")).strip()
        iva_pct   = float(data.get("iva_pct", 0) or 0)
        moneda    = str(data.get("moneda","USD")).upper()
        fx_rate   = float(data.get("fx_rate") or 1.0) if moneda=="MXN" else 1.0
        if not items:
            return jsonify({"error": "La PO debe tener al menos un item"}), 400
        with lock:
            po_number = gpo_next_number()
            now = datetime.datetime.now().isoformat()
            po_items = []
            for idx, it in enumerate(items):
                if job_type in ("Shopfloor","Fix Asset"):
                    item_job = job_type
                elif job_type == "Unico":
                    item_job = job_main
                else:
                    item_job = str(it.get("job","")).strip().upper() or job_main
                qty = max(1, int(it.get("quantity",1)))
                up  = float(it.get("unit_price",0) or 0)
                po_items.append({
                    "line":        idx+1,
                    "cat_type":    str(it.get("cat_type","")).strip(),
                    "cat_code":    str(it.get("cat_code","")).strip().upper(),
                    "part_number": str(it.get("part_number","")).strip().upper(),
                    "brand":       str(it.get("brand","")).strip().upper(),
                    "description": str(it.get("description","")).strip(),
                    "label_code":  str(it.get("label_code","")).strip().upper(),
                    "quantity":    qty,
                    "unit_price":  up,
                    "total":       round(qty * up, 2),
                    "job":         item_job,
                    "notes":       str(it.get("notes","")).strip(),
                })
            subtotal     = round(sum(i["total"] for i in po_items), 2)
            iva_amt      = round(subtotal * iva_pct / 100, 2)
            total        = round(subtotal + iva_amt, 2)
            total_usd    = round(total / fx_rate, 2) if moneda=="MXN" and fx_rate else total
            rec = {
                "po_number":     po_number,
                "supplier":      supplier,
                "supplier_name": str(supplier.get("nombre","")).strip().upper(),
                "pt_sv":         pt_sv,
                "cpo":           cpo,
                "job_type":      job_type,
                "job":           job_main,
                "items":         po_items,
                "subtotal":      subtotal,
                "moneda":        moneda,
                "fx_rate":       fx_rate,
                "iva_pct":       iva_pct,
                "iva_amt":       iva_amt,
                "total":         total,
                "total_usd":     total_usd,
                "notes":         notes,
                "status":        "Emitida",
                "created_by":    session.get("user",""),
                "created_at":    now,
            }
            records = gpo_load()
            records.append(rec)
            gpo_save(records)
            # Registrar en IPOs por item
            year = datetime.datetime.now().year
            ipo_records = po_load(year)
            existing_claves = [int(r.get("clave",0)) for r in ipo_records
                               if str(r.get("clave","")).isdigit()]
            next_clave = max(existing_claves)+1 if existing_claves else 1
            for i, it in enumerate(po_items):
                ipo_records.append({
                    "clave":                po_number,
                    "fecha_doc":            now[:10],
                    "entregar_a":           it["job"],
                    "nombre":               rec["supplier_name"],
                    "subtotal":             it["total"],
                    "moneda":               moneda,          # ← CRITICAL: store currency
                    "tipo_cambio":          fx_rate if moneda=="MXN" else 1.0,
                    "subtotal_mxn":         round(it["total"] * fx_rate, 2) if moneda=="MXN" else it["total"],
                    "estatus":              "Emitida",
                    "descuento_financiero": 0,
                    "pct_descuento":        0,
                    "fecha_recepcion":      "",
                    "gpo_number":           po_number,
                    "gpo_pdf":              i==0,
                    "cpo":                  cpo,
                    "part_number":          it["part_number"],
                    "description":          it["description"],
                    "quantity":             it["quantity"],
                    "unit_price":           it["unit_price"],
                })
            po_save(year, ipo_records)
            # Actualizar último precio en catálogo
            for it in po_items:
                ct = it.get("cat_type","")
                cc = it.get("cat_code","")
                up = it.get("unit_price",0)
                if ct in CAT_CONFIG and cc and up > 0:
                    cat_records = cat_load(ct)
                    for cr in cat_records:
                        if cr.get("code","").upper() == cc.upper():
                            cr["last_price"] = up
                            cr["updated_at"] = now
                            break
                    cat_save(ct, cat_records)
        return jsonify({"ok": True, "po_number": po_number, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _numero_a_letras(num):
    u  = ['','uno','dos','tres','cuatro','cinco','seis','siete','ocho','nueve',
          'diez','once','doce','trece','catorce','quince','dieciséis',
          'diecisiete','dieciocho','diecinueve']
    d  = ['','','veinte','treinta','cuarenta','cincuenta','sesenta',
          'setenta','ochenta','noventa']
    c  = ['','cien','doscientos','trescientos','cuatrocientos','quinientos',
          'seiscientos','setecientos','ochocientos','novecientos']
    def _c(n):
        if n < 20:  return u[n]
        if n < 100: return d[n//10] + (' y '+u[n%10] if n%10 else '')
        if n == 100: return 'cien'
        return c[n//100] + (' '+_c(n%100) if n%100 else '')
    def _m(n):
        if n == 0:     return 'cero'
        if n < 1000:   return _c(n)
        if n < 2000:   return 'mil' + (' '+_c(n%1000) if n%1000 else '')
        if n < 1000000:return _c(n//1000)+' mil'+(' '+_c(n%1000) if n%1000 else '')
        if n < 2000000:return 'un millón'+(' '+_m(n%1000000) if n%1000000 else '')
        return _c(n//1000000)+' millones'+(' '+_m(n%1000000) if n%1000000 else '')
    entero  = int(abs(num))
    centavos= round((abs(num)-entero)*100)
    letras  = _m(entero).upper()
    return f"{letras} {centavos:02d}/100"

def _gpo_update_ipos(po_number, fields: dict):
    """Update all IPO records linked to this GPO with the given fields."""
    po_upper = str(po_number).upper()
    # Check GPO record for creation year hint
    try:
        gpo_records = gpo_load()
        gpo_rec = next((r for r in gpo_records if str(r.get("po_number","")).upper() == po_upper), None)
        creation_year = int(gpo_rec.get("created_at","")[:4]) if gpo_rec and gpo_rec.get("created_at") else None
    except:
        creation_year = None

    years_to_check = list(range(CURRENT_YEAR - 3, CURRENT_YEAR + 2))
    if creation_year and creation_year not in years_to_check:
        years_to_check.append(creation_year)

    for year in years_to_check:
        records = po_load(year)
        changed = False
        for r in records:
            if (str(r.get("gpo_number","")).upper() == po_upper or
                str(r.get("clave","")).upper() == po_upper):
                r.update(fields)
                changed = True
        if changed:
            po_save(year, records)

@app.route("/api/gpo/<po_number>/modificar", methods=["POST"])
def api_gpo_modificar(po_number):
    """
    Modifica una GPO existente.
    tipo: 'cancelar' | 'cierre_anticipado' | 'nueva_version'
    """
    try:
        data = request.get_json()
        tipo = str(data.get("tipo","")).strip()
        if tipo not in ("cancelar","cierre_anticipado","nueva_version"):
            return jsonify({"error":"Tipo de modificación inválido"}), 400

        with lock:
            records = gpo_load()
            rec = next((r for r in records
                if str(r.get("po_number","")).upper() == po_number.upper()), None)
            if not rec:
                return jsonify({"error":f"Orden {po_number} no encontrada"}), 404

            status = rec.get("status","")
            if status == "Entregada":
                return jsonify({"error":"Las órdenes entregadas no se pueden modificar"}), 400

            now = datetime.datetime.now().isoformat()
            user = session.get("user","")

            if tipo == "cancelar":
                if status not in ("Emitida",):
                    return jsonify({"error":"Solo se puede cancelar una orden Emitida sin entregas"}), 400
                rec["status"]            = "Cancelada"
                rec["effective_total"]   = 0.0
                rec["effective_total_usd"] = 0.0
                rec["modificacion_tipo"] = "Cancelar"
                rec["modificacion_by"]   = user
                rec["modificacion_at"]   = now
                rec["modificacion_nota"] = str(data.get("nota","")).strip()
                # Propagate to IPO records: set subtotal to 0 and status to Cancelada
                _gpo_update_ipos(po_number, {"estatus": "Cancelada", "subtotal": 0.0, "subtotal_mxn": 0.0})

            elif tipo == "cierre_anticipado":
                if status not in ("Parcial",):
                    return jsonify({"error":"El cierre anticipado aplica solo a órdenes con entrega parcial"}), 400

                # Calculate effective total from actually delivered quantities
                moneda  = rec.get("moneda","USD")
                fx_rate = float(rec.get("fx_rate",1) or 1)
                val_recibido = 0.0
                for it in rec.get("items",[]):
                    qty_del   = float(it.get("quantity_delivered",0) or 0)
                    unit_price= float(it.get("unit_price",0) or 0)
                    val_recibido += qty_del * unit_price
                val_recibido = round(val_recibido, 2)

                rec["status"]              = "Cierre Anticipado"
                rec["effective_total"]     = val_recibido
                rec["effective_total_usd"] = round(val_recibido / fx_rate, 2) if moneda=="MXN" else val_recibido
                rec["modificacion_tipo"]   = "Cierre Anticipado"
                rec["modificacion_by"]     = user
                rec["modificacion_at"]     = now
                rec["modificacion_nota"]   = str(data.get("nota","")).strip()

                # Update IPO records: set non-delivered items to 0, keep delivered amounts
                po_upper = po_number.upper()
                try:
                    gpo_created_yr = int(rec.get("created_at","")[:4])
                except: gpo_created_yr = CURRENT_YEAR
                for yr in list({gpo_created_yr, CURRENT_YEAR}):
                    ipo_recs = po_load(yr)
                    changed  = False
                    for ir in ipo_recs:
                        if (str(ir.get("gpo_number","")).upper() == po_upper or
                            str(ir.get("clave","")).upper() == po_upper):
                            # Find matching GPO item by part number
                            pnum = str(ir.get("part_number","")).upper()
                            gpo_item = next((it for it in rec.get("items",[])
                                if str(it.get("part_number","")).upper() == pnum), None)
                            if gpo_item:
                                qty_del = float(gpo_item.get("quantity_delivered",0) or 0)
                                up      = float(gpo_item.get("unit_price",0) or 0)
                                new_sub = round(qty_del * up, 2)
                            else:
                                new_sub = 0.0
                            ir["subtotal"]     = new_sub
                            ir["subtotal_mxn"] = round(new_sub * fx_rate, 2) if moneda=="MXN" else new_sub
                            ir["estatus"]      = "Cierre Anticipado"
                            changed = True
                    if changed:
                        po_save(yr, ipo_recs)
                        break

            elif tipo == "nueva_version":
                if status not in ("Emitida",):
                    return jsonify({"error":"Solo se puede crear nueva versión en órdenes Emitidas sin entregas"}), 400
                new_items = data.get("items",[])
                if not new_items:
                    return jsonify({"error":"La nueva versión debe tener al menos un item"}), 400
                iva_pct  = float(rec.get("iva_pct",0) or 0)
                moneda   = rec.get("moneda","USD")
                fx_rate  = float(rec.get("fx_rate",1) or 1)
                new_subtotal = 0.0
                parsed_items = []
                for idx, it in enumerate(new_items):
                    qty  = float(it.get("quantity",0) or 0)
                    price= float(it.get("unit_price",0) or 0)
                    tot  = round(qty * price, 2)
                    new_subtotal += tot
                    parsed_items.append({
                        "line":        it.get("line", idx+1),
                        "cat_type":    it.get("cat_type",""),
                        "cat_code":    it.get("cat_code",""),
                        "part_number": str(it.get("part_number","")).strip().upper(),
                        "brand":       str(it.get("brand","")).strip().upper(),
                        "description": str(it.get("description","")).strip(),
                        "label_code":  it.get("label_code",""),
                        "quantity":    qty,
                        "unit_price":  price,
                        "total":       tot,
                        "job":         it.get("job", rec.get("job","")),
                        "notes":       str(it.get("notes","")).strip(),
                    })
                new_subtotal = round(new_subtotal, 2)
                new_iva      = round(new_subtotal * iva_pct / 100, 2)
                new_total    = round(new_subtotal + new_iva, 2)
                new_total_usd= round(new_total / fx_rate, 2) if moneda=="MXN" and fx_rate else new_total
                # Archive old version
                version = int(rec.get("version",1))
                rec["version_history"] = rec.get("version_history",[]) + [{
                    "version":  version,
                    "items":    rec.get("items",[]),
                    "subtotal": rec.get("subtotal",0),
                    "total":    rec.get("total",0),
                    "archived_at": now,
                    "archived_by": user,
                }]
                rec["version"]     = version + 1
                rec["items"]       = parsed_items
                rec["subtotal"]    = new_subtotal
                rec["iva_amt"]     = new_iva
                rec["total"]       = new_total
                rec["total_usd"]   = new_total_usd
                rec["modificacion_tipo"] = f"Nueva Versión v{version+1}"
                rec["modificacion_by"]   = user
                rec["modificacion_at"]   = now
                rec["modificacion_nota"] = str(data.get("nota","")).strip()

                # ── Replace IPO records with new items ──────────────────
                po_upper = po_number.upper()
                supplier_name = rec.get("supplier_name","")
                # Find which year holds the existing records
                target_year = now[:4]
                try:
                    gpo_created = rec.get("created_at","")[:4]
                    if gpo_created: target_year = int(gpo_created)
                    else: target_year = CURRENT_YEAR
                except: target_year = CURRENT_YEAR

                for yr in list({int(target_year), CURRENT_YEAR}):
                    existing = po_load(yr)
                    # Remove old records for this PO
                    filtered = [r for r in existing
                                if str(r.get("gpo_number","")).upper() != po_upper
                                and str(r.get("clave","")).upper() != po_upper]
                    if len(filtered) != len(existing):  # records were found and removed
                        # Add new records per item
                        for i, it in enumerate(parsed_items):
                            item_job = it.get("job", "") or rec.get("job","")
                            filtered.append({
                                "clave":                po_number,
                                "fecha_doc":            now[:10],
                                "entregar_a":           item_job,
                                "nombre":               supplier_name,
                                "subtotal":             it["total"],
                                "moneda":               moneda,
                                "tipo_cambio":          fx_rate if moneda=="MXN" else 1.0,
                                "subtotal_mxn":         round(it["total"]*fx_rate,2) if moneda=="MXN" else it["total"],
                                "estatus":              "Emitida",
                                "descuento_financiero": 0,
                                "pct_descuento":        0,
                                "fecha_recepcion":      "",
                                "gpo_number":           po_number,
                                "gpo_pdf":              i==0,
                                "part_number":          it["part_number"],
                                "description":          it["description"],
                                "quantity":             it["quantity"],
                                "unit_price":           it["unit_price"],
                                "version":              version+1,
                            })
                        po_save(yr, filtered)
                        break  # Found the right year, stop

            gpo_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/gpo/<po_number>/lookup", methods=["GET"])
def api_gpo_lookup(po_number):
    """Returns a single GPO record with delivered quantity info."""
    try:
        records = gpo_load()
        rec = next((r for r in records
            if str(r.get("po_number","")).upper() == po_number.upper()), None)
        if not rec:
            return jsonify({"error":f"Orden {po_number} no encontrada"}), 404
        # Calculate cierre breakdown
        moneda  = rec.get("moneda","USD")
        fx_rate = float(rec.get("fx_rate",1) or 1)
        delivered_total = 0.0
        pending_total   = 0.0
        for it in rec.get("items",[]):
            qty     = float(it.get("quantity",0) or 0)
            qty_del = float(it.get("quantity_delivered",0) or 0)
            up      = float(it.get("unit_price",0) or 0)
            delivered_total += qty_del * up
            pending_total   += (qty - qty_del) * up
        rec["_delivered_total"] = round(delivered_total, 2)
        rec["_pending_total"]   = round(pending_total, 2)
        return jsonify({"record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/gpo/<po_number>/pdf")
def api_gpo_pdf(po_number):
    try:
        records = gpo_load()
        rec = next((r for r in records if r["po_number"]==po_number.upper()), None)
        if not rec: return jsonify({"error":"PO no encontrada"}), 404
        sup      = rec.get("supplier",{})
        items    = rec.get("items",[])
        subtotal = rec.get("subtotal",0)
        iva_pct  = float(rec.get("iva_pct", 0) or 0)
        iva_amt  = rec.get("iva_amt", round(float(subtotal) * iva_pct / 100, 2))
        total    = rec.get("total",   round(float(subtotal) + float(iva_amt), 2))
        moneda   = rec.get("moneda","USD")
        fx_rate  = rec.get("fx_rate", 1.0) or 1.0
        iva_label = f"IVA / VAT ({iva_pct:.0f}%)"
        moneda_label = "PESOS MEXICANOS" if moneda=="MXN" else "DÓLARES AMERICANOS"
        moneda_sym   = "$" if moneda=="MXN" else "USD $"
        total_letras = _numero_a_letras(total)
        total_leyenda = f"{total_letras} {moneda_label}"
        fmt      = lambda v: f"{moneda_sym}{float(v):,.2f}"
        logo_path = _os.path.join(_BASE, "static", "persico_logo.webp")
        logo_b64  = ""
        if _os.path.exists(logo_path):
            import base64 as _b64
            with open(logo_path,"rb") as lf:
                logo_b64 = _b64.b64encode(lf.read()).decode()
        logo_tag = f'<img src="data:image/webp;base64,{logo_b64}" style="height:54px">' if logo_b64 else '<span style="font-size:22px;font-weight:900;color:#c8102e">PERSICO</span>'
        rows_html = ""
        for i, it in enumerate(items, 1):
            rows_html += f"""
            <tr>
              <td style="text-align:center">{it.get('line', i)}</td>
              <td style="font-family:monospace">{it.get('cat_code','—')}</td>
              <td style="font-family:monospace">{it.get('part_number','—')}</td>
              <td>{it.get('brand','—')}</td>
              <td>{it.get('description','—')}</td>
              <td style="font-family:monospace;color:#555">{it.get('label_code','—')}</td>
              <td style="text-align:center">{it.get('job','—')}</td>
              <td style="text-align:right">{it.get('quantity',0)}</td>
              <td style="text-align:right">{fmt(it.get('unit_price',0))}</td>
              <td style="text-align:right;font-weight:600">{fmt(it.get('total',0))}</td>
              <td style="font-size:9px;color:#777">{it.get('notes','')}</td>
            </tr>"""
        html = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>{rec['po_number']}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Arial,sans-serif;font-size:11px;color:#111;padding:28px 32px}}
  .header{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:18px;border-bottom:3px solid #c8102e;padding-bottom:14px}}
  .po-num{{font-size:28px;font-weight:900;color:#c8102e;letter-spacing:1px;text-align:right}}
  .po-num-label{{font-size:10px;color:#888;text-transform:uppercase;letter-spacing:1px;text-align:right}}
  .company-block{{font-size:10px;line-height:1.6;color:#444}}
  .company-name{{font-size:13px;font-weight:700;color:#111;margin-bottom:2px}}
  .section-title{{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;color:#888;margin-bottom:4px;margin-top:12px}}
  .sup-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px 20px;margin-bottom:14px;padding:10px 14px;background:#f7f7f7;border-radius:6px;border:1px solid #e8e8e8}}
  .sup-field{{font-size:10px}}.sup-field b{{color:#444;display:block;font-size:9px;text-transform:uppercase}}
  .meta-bar{{display:flex;gap:24px;padding:8px 14px;background:#1f3864;color:#fff;border-radius:6px;margin-bottom:14px;font-size:11px;flex-wrap:wrap}}
  .meta-bar span{{opacity:.75}}.meta-bar b{{opacity:1}}
  table{{width:100%;border-collapse:collapse;margin-bottom:12px;font-size:10px}}
  thead th{{background:#1f3864;color:#fff;padding:7px 8px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.8px;white-space:nowrap}}
  tbody tr:nth-child(even){{background:#f7f7f7}}
  tbody td{{padding:6px 8px;border-bottom:1px solid #ebebeb;vertical-align:top}}
  .subtotal-row{{display:flex;justify-content:flex-end;gap:20px;margin-top:4px;padding:10px 14px;background:#1f3864;color:#fff;border-radius:6px;font-size:13px;font-weight:700}}
  .footer{{margin-top:28px;border-top:1px solid #ddd;padding-top:10px;font-size:9px;color:#aaa;display:flex;justify-content:space-between}}
  .barcode-wrap{{text-align:right;margin-top:6px}}
  .barcode-wrap svg{{max-width:220px}}
  @media print{{body{{padding:10px 14px}}}}
</style>
<script src="https://cdn.jsdelivr.net/npm/jsbarcode@3.11.5/dist/barcodes/JsBarcode.code128.min.js"></script>
</head>
<body>
<div class="header">
  <div>
    {logo_tag}
    <div class="company-block" style="margin-top:8px">
      <div class="company-name">PERSICO MÉXICO S.A. DE C.V.</div>
      RFC: PME200101AB1<br>
      Blvd. Forjadores de Puebla 3120, Parque Industrial<br>
      Puebla, Pue. C.P. 72070 · Tel. (222) 000-0000
    </div>
  </div>
  <div>
    <div class="po-num-label">Purchase Order</div>
    <div class="po-num">{rec['po_number'].replace('-','_')}</div>
    <div style="font-size:10px;color:#888;text-align:right;margin-top:4px">{rec['created_at'][:10]}</div>
    {f'<div style="font-size:10px;font-weight:700;text-align:right;color:#f5a623;margin-top:2px;letter-spacing:.5px">Versión {rec["version"]}</div>' if rec.get("version") and int(rec.get("version",1)) > 1 else '<div style="font-size:10px;color:#aaa;text-align:right;margin-top:2px">v1 — Original</div>'}
    <div class="barcode-wrap">
      <svg id="barcode"></svg>
    </div>
  </div>
</div>
<script>
  JsBarcode("#barcode", "{rec['po_number'].split('-')[-1]}", {{
    format: "CODE128",
    width: 1.6,
    height: 40,
    displayValue: true,
    fontSize: 11,
    margin: 4,
    lineColor: "#111",
    background: "transparent"
  }});
</script>
<div class="section-title">Datos del Proveedor</div>
<div class="sup-grid">
  <div class="sup-field"><b>Nombre</b>{sup.get('nombre','—')}</div>
  <div class="sup-field"><b>RFC</b>{sup.get('rfc','—')}</div>
  <div class="sup-field"><b>Dirección</b>{sup.get('calle','')} {sup.get('num_exterior','')} {sup.get('num_interior','')}</div>
  <div class="sup-field"><b>Teléfono</b>{sup.get('telefono','—')}</div>
  <div class="sup-field"><b>Forma de Pago</b>{sup.get('forma_pago','—')}</div>
  <div class="sup-field"><b>Moneda</b>{sup.get('moneda','MXN')}</div>
</div>
<div class="meta-bar">
  {f'<div><span>PT/SV</span> <b>{rec["pt_sv"]}</b></div>' if rec.get("pt_sv") else ''}
  <div><span>Tipo Job</span> <b>{rec.get('job_type','')}</b></div>
  {f'<div><span>Job</span> <b>{rec["job"]}</b></div>' if rec.get("job") and rec.get("job_type") in ("Unico","Shopfloor","Fix Asset") else ''}
  {f'<div><span>CPO</span> <b>{rec["cpo"]}</b></div>' if rec.get("cpo") else ''}
  <div><span>Solicitado por</span> <b>{rec.get('created_by','')}</b></div>
  {f'<div><span>Notas</span> <b>{rec["notes"]}</b></div>' if rec.get("notes") else ''}
</div>
<table>
  <thead><tr>
    <th>#</th><th>Código</th><th>No. Parte</th><th>Marca</th>
    <th>Descripción</th><th>Etiqueta</th><th>Job</th>
    <th style="text-align:right">Cant.</th>
    <th style="text-align:right">P. Unit.</th>
    <th style="text-align:right">Total</th>
    <th>Notas</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<div class="subtotal-row">
  <div style="display:flex;flex-direction:column;gap:4px;align-items:flex-end;min-width:260px">
    <div style="display:flex;justify-content:space-between;width:100%;font-size:11px;font-weight:400;opacity:.85">
      <span>SUBTOTAL {moneda}</span><span>{fmt(subtotal)}</span>
    </div>
    <div style="display:flex;justify-content:space-between;width:100%;font-size:11px;font-weight:400;opacity:.85">
      <span>{iva_label}</span><span>{fmt(iva_amt)}</span>
    </div>
    <div style="display:flex;justify-content:space-between;width:100%;border-top:1px solid rgba(255,255,255,.3);padding-top:6px;margin-top:2px">
      <span>TOTAL {moneda}</span><span>{fmt(total)}</span>
    </div>
  </div>
</div>
<div style="margin-top:10px;padding:10px 14px;background:#f7f7f7;border:1px solid #e0e0e0;border-radius:6px;font-size:10px;color:#444;font-style:italic;text-align:center;letter-spacing:.3px">
  <b>{total_leyenda}</b>
</div>
<div class="footer">
  <span>Generado por Persico Suite · {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</span>
  <span>{rec['po_number']} · {rec.get('supplier_name','')}</span>
</div>
</body></html>"""
        return Response(html, mimetype="text/html",
            headers={"Content-Disposition": f"inline;filename={rec['po_number']}.html"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  PT / SV IMPORT FROM EXCEL
# ══════════════════════════════════════════════════════════════════
@app.route("/api/pt/import", methods=["POST"])
def api_import_pt():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error": "No se recibió archivo"}), 400
        mode = request.form.get("mode", "append")
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            if cell.value:
                headers[str(cell.value).strip().lower()] = cell.column - 1
        def col(*aliases):
            for a in aliases:
                if a.lower() in headers: return headers[a.lower()]
        return None
        ci_pt   = col("pt_number", "pt number", "pt")
        ci_cust = col("customer", "cliente")
        ci_prog = col("customer_program", "customer program", "programa")
        ci_pm   = col("pm")
        ci_jobs = col("jobs")
        ci_note = col("notes", "notas")
        if ci_pt is None:
            return jsonify({"error": "No se encontró columna PT Number"}), 400
        imported = 0
        with lock:
            records = pt_load() if mode == "append" else []
            existing = {r["pt_number"] for r in records}
            for row in ws.iter_rows(min_row=2, values_only=True):
                pt_num = str(row[ci_pt]).strip().upper() if ci_pt is not None and row[ci_pt] else ""
                if not pt_num or pt_num in ("NONE", "", "NAN"): continue
                if pt_num in existing: continue
                # Parse jobs — can be comma-separated or list-like string
                raw_jobs = str(row[ci_jobs]).strip() if ci_jobs is not None and row[ci_jobs] else ""
                jobs_list = [j.strip().upper() for j in
                             raw_jobs.replace("[","").replace("]","").replace("'","").split(",")
                             if j.strip() and j.strip().upper() not in ("NONE","NAN","")]
                rec = {
                    "pt_number":        pt_num,
                    "customer":         str(row[ci_cust]).strip() if ci_cust is not None and row[ci_cust] else "",
                    "customer_program": str(row[ci_prog]).strip() if ci_prog is not None and row[ci_prog] else "",
                    "pm":               str(row[ci_pm]).strip() if ci_pm is not None and row[ci_pm] else "",
                    "jobs":             jobs_list,
                    "notes":            str(row[ci_note]).strip() if ci_note is not None and row[ci_note] else "",
                    "created_at":       datetime.datetime.now().isoformat(),
                }
                records.append(rec)
                existing.add(pt_num)
                imported += 1
            pt_save(records)
        return jsonify({"ok": True, "imported": imported, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/sv/import", methods=["POST"])
def api_import_sv():
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error": "No se recibió archivo"}), 400
        mode = request.form.get("mode", "append")
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        headers = {}
        for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
            if cell.value:
                headers[str(cell.value).strip().lower()] = cell.column - 1
        def col(*aliases):
            for a in aliases:
                if a.lower() in headers: return headers[a.lower()]
            return None
        ci_sv   = col("sv_number", "sv number", "sv")
        ci_cust = col("customer", "cliente")
        ci_prog = col("customer_program", "customer program", "programa")
        ci_pm   = col("pm")
        ci_jobs = col("jobs")
        ci_note = col("notes", "notas")
        ci_q    = col("q_number", "q number", "quote")
        if ci_sv is None:
            return jsonify({"error": "No se encontró columna SV Number"}), 400
        imported = 0
        with lock:
            records = sv_load() if mode == "append" else []
            existing = {r["sv_number"] for r in records}
            for row in ws.iter_rows(min_row=2, values_only=True):
                sv_num = str(row[ci_sv]).strip().upper() if ci_sv is not None and row[ci_sv] else ""
                if not sv_num or sv_num in ("NONE", "", "NAN"): continue
                if sv_num in existing: continue
                raw_jobs = str(row[ci_jobs]).strip() if ci_jobs is not None and row[ci_jobs] else ""
                jobs_list = [j.strip().upper() for j in
                             raw_jobs.replace("[","").replace("]","").replace("'","").split(",")
                             if j.strip() and j.strip().upper() not in ("NONE","NAN","")]
                rec = {
                    "sv_number":        sv_num,
                    "customer":         str(row[ci_cust]).strip() if ci_cust is not None and row[ci_cust] else "",
                    "customer_program": str(row[ci_prog]).strip() if ci_prog is not None and row[ci_prog] else "",
                    "pm":               str(row[ci_pm]).strip() if ci_pm is not None and row[ci_pm] else "",
                    "jobs":             jobs_list,
                    "notes":            str(row[ci_note]).strip() if ci_note is not None and row[ci_note] else "",
                    "q_number":         str(row[ci_q]).strip() if ci_q is not None and row[ci_q] else "",
                    "created_at":       datetime.datetime.now().isoformat(),
                }
                records.append(rec)
                existing.add(sv_num)
                imported += 1
            sv_save(records)
        return jsonify({"ok": True, "imported": imported, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ══════════════════════════════════════════════════════════════════
#  PROJECT CONFIG  (Configurar Proyecto)
# ══════════════════════════════════════════════════════════════════
PROJCFG_FILE = _os.path.join(_DATA, "project_configs.json")

def projcfg_load():
    p = Path(PROJCFG_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def projcfg_save(records):
    Path(PROJCFG_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(PROJCFG_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

@app.route("/api/projconfig", methods=["GET"])
def api_get_projconfig():
    try:
        records = projcfg_load()
        q = request.args.get("q","").upper()
        if q:
            records = [r for r in records if q in r.get("ptsv","").upper()]
        return jsonify({"records": records})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/projconfig", methods=["POST"])
def api_create_projconfig():
    try:
        data = request.get_json()
        ptsv = str(data.get("ptsv","")).strip().upper()
        if not ptsv: return jsonify({"error":"PT/SV requerido"}), 400
        with lock:
            records = projcfg_load()
            # Remove existing config for same PT/SV (overwrite)
            records = [r for r in records if r.get("ptsv","").upper() != ptsv]
            rec = {
                "id":         f"PC-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}",
                "ptsv":       ptsv,
                "jobs":       data.get("jobs", []),
                "timing":     data.get("timing", []),
                "created_by": session.get("user",""),
                "created_at": datetime.datetime.now().isoformat(),
                "updated_at": datetime.datetime.now().isoformat(),
            }
            records.append(rec)
            projcfg_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/projconfig/<cfg_id>", methods=["DELETE"])
def api_delete_projconfig(cfg_id):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            records = projcfg_load()
            new = [r for r in records if r.get("id","") != cfg_id]
            projcfg_save(new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ══════════════════════════════════════════════════════════════════
#  INGRESO DE MATERIAL / APARTADOS
# ══════════════════════════════════════════════════════════════════
INGRESO_FILE  = _os.path.join(_DATA, "ingresos.json")
APARTADO_FILE = _os.path.join(_DATA, "apartados.json")

def ingreso_load():
    _h = _cache_get("ingreso")
    if _h is not None: return _h
    p = Path(INGRESO_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def ingreso_save(records):
    Path(INGRESO_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(INGRESO_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("ingreso", records)

def apartado_load():
    _h = _cache_get("apartado")
    if _h is not None: return _h
    p = Path(APARTADO_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def apartado_save(records):
    Path(APARTADO_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(APARTADO_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("apartado", records)

@app.route("/api/ingreso", methods=["GET"])
def api_get_ingreso():
    try:
        records = ingreso_load()
        q    = request.args.get("q","").lower()
        tipo = request.args.get("tipo","")
        if q:
            records = [r for r in records if
                q in (r.get("po_number","")).lower() or
                q in (r.get("job","")).lower() or
                any(q in (i.get("part_number","")).lower() or
                    q in (i.get("description","")).lower()
                    for i in r.get("items",[]))]
        if tipo:
            records = [r for r in records if r.get("tipo","") == tipo]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/ingreso", methods=["POST"])
def api_create_ingreso():
    """
    Procesa un ingreso (manual o por OC).
    Body: { tipo, po_number, recibe, items: [{part_number, description, brand,
            cat_code, label_code, quantity_ordered, quantity_delivered,
            unit_cost, job, notes}] }
    """
    try:
        data    = request.get_json()
        tipo    = str(data.get("tipo","manual"))
        po_num  = str(data.get("po_number","")).strip().upper()
        recibe  = str(data.get("recibe","")).strip()
        items   = data.get("items", [])
        if not items:
            return jsonify({"error":"Debe incluir al menos un item"}), 400

        with lock:
            now = datetime.datetime.now().isoformat()
            # Build ingreso record
            ing_items = []
            for it in items:
                qty_ord = float(it.get("quantity_ordered", it.get("quantity_delivered", 0)) or 0)
                # For manual ingress, everything is considered delivered
                qty_del = float(it.get("quantity_delivered", 0) or 0)
                if tipo == "manual":
                    qty_del = qty_ord  # manual = all ordered quantities are delivered
                uc = float(it.get("unit_cost", 0) or 0)
                if qty_del <= 0: continue   # skip zero-qty rows
                job_val = str(it.get("job","")).strip().upper()
                ing_items.append({
                    "part_number":       str(it.get("part_number","")).strip().upper(),
                    "brand":             str(it.get("brand","")).strip().upper(),
                    "description":       str(it.get("description","")).strip(),
                    "cat_code":          str(it.get("cat_code","")).strip().upper(),
                    "label_code":        str(it.get("label_code","")).strip().upper(),
                    "quantity_ordered":  qty_ord,
                    "quantity_delivered":qty_del,
                    "unit_cost":         uc,
                    "total":             round(qty_del * uc, 2),
                    "job":               job_val,
                    "notes":             str(it.get("notes","")).strip(),
                })
            if not ing_items:
                return jsonify({"error": "Ningún item con cantidad > 0"}), 400

            rec = {
                "id":         _doc_next_number("WI"),
                "tipo":       tipo,
                "po_number":  po_num,
                "recibe":     session.get("user",""),   # always the logged user
                "items":      ing_items,
                "created_by": session.get("user",""),
                "fecha":      now[:10],
                "created_at": now,
            }
            ingresos = ingreso_load()
            ingresos.append(rec)
            ingreso_save(ingresos)

            # ── Registrar en Apartados (estructura consolidada por No. Parte)
            apartados = apartado_load()
            for it in ing_items:
                # qty_del > 0 is guaranteed by the filter above
                pnum = it["part_number"]
                job  = it["job"]
                qty  = it["quantity_delivered"]
                uc   = it["unit_cost"]
                now2 = datetime.datetime.now().isoformat()
                # Find existing part_number record
                existing = next((a for a in apartados
                    if a.get("part_number","").upper() == pnum.upper()), None)
                if not existing:
                    # New part → create consolidated record
                    apartados.append({
                        "part_number":   pnum,
                        "brand":         it["brand"],
                        "description":   it["description"],
                        "cat_code":      it["cat_code"],
                        "label_code":    it["label_code"],
                        "total_quantity": qty,
                        "jobs": [{
                            "job":       job,
                            "quantity":  qty,
                            "unit_cost": uc,
                            "ingresos":  [rec["id"]],
                        }],
                        "created_at":  now2,
                        "updated_at":  now2,
                    })
                else:
                    # Part exists → find or add job
                    existing_job = next((j for j in existing.get("jobs",[])
                        if (j.get("job","")).upper() == job.upper()), None)
                    if not existing_job:
                        existing.setdefault("jobs",[]).append({
                            "job":       job,
                            "quantity":  qty,
                            "unit_cost": uc,
                            "ingresos":  [rec["id"]],
                        })
                    else:
                        existing_job["quantity"] = existing_job.get("quantity",0) + qty
                        existing_job.setdefault("ingresos",[]).append(rec["id"])
                    # Update total
                    existing["total_quantity"] = sum(
                        j.get("quantity",0) for j in existing.get("jobs",[]))
                    existing["updated_at"] = now2
                    # Update description/brand if blank
                    if not existing.get("brand") and it["brand"]:
                        existing["brand"] = it["brand"]
                    if not existing.get("description") and it["description"]:
                        existing["description"] = it["description"]
            apartado_save(apartados)

            # ── Actualizar estatus de la GPO si es por OC
            if tipo == "gpo" and po_num:
                gpos = gpo_load()
                gpo_rec = next((g for g in gpos if g.get("po_number","").upper()==po_num), None)
                if gpo_rec:
                    for gpo_item in gpo_rec.get("items",[]):
                        match = next((it for it in ing_items
                            if it["part_number"]==gpo_item.get("part_number","").upper()), None)
                        if match:
                            prev_del = float(gpo_item.get("quantity_delivered",0))
                            gpo_item["quantity_delivered"] = prev_del + match["quantity_delivered"]
                    all_delivered = all(
                        float(gi.get("quantity_delivered",0)) >= float(gi.get("quantity",1))
                        for gi in gpo_rec.get("items",[])
                    )
                    any_delivered = any(
                        float(gi.get("quantity_delivered",0)) > 0
                        for gi in gpo_rec.get("items",[])
                    )
                    gpo_rec["status"] = "Entregada" if all_delivered else ("Parcial" if any_delivered else "Emitida")
                    gpo_save(gpos)
                    year = datetime.datetime.now().year
                    ipo_recs = po_load(year)
                    for ir in ipo_recs:
                        if ir.get("gpo_number","").upper() == po_num:
                            ir["estatus"] = gpo_rec["status"]
                    po_save(year, ipo_recs)

            # ── Actualizar quantity_delivered en IPO para ingresos SAE
            elif tipo == "sae" and po_num:
                po_upper = po_num.upper()
                for yr in range(CURRENT_YEAR - 2, CURRENT_YEAR + 2):
                    ipo_recs = po_load(yr)
                    changed = False
                    for ir in ipo_recs:
                        clave = str(ir.get("clave","") or ir.get("gpo_number","")).upper()
                        if clave != po_upper and \
                           clave.replace("PO-","").lstrip("0") != po_upper.replace("PO-","").lstrip("0"):
                            continue
                        pnum = str(ir.get("part_number","")).strip().upper()
                        match = next((it for it in ing_items
                            if str(it.get("part_number","")).strip().upper() == pnum), None)
                        if match:
                            prev = float(ir.get("quantity_delivered",0))
                            ir["quantity_delivered"] = prev + match["quantity_delivered"]
                            qty_ord = float(ir.get("quantity",ir.get("quantity_ordered",1)) or 1)
                            ir["estatus"] = "Entregada" if ir["quantity_delivered"] >= qty_ord else "Parcial"
                            changed = True
                    if changed:
                        po_save(yr, ipo_recs)

        return jsonify({"ok": True, "record": rec,
                        "apartados_created": sum(1 for it in ing_items if it["quantity_delivered"]>0)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/ingreso/<ing_id>", methods=["DELETE"])
def api_delete_ingreso(ing_id):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            ingresos = ingreso_load()
            new = [r for r in ingresos if r.get("id","") != ing_id]
            ingreso_save(new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ══════════════════════════════════════════════════════════════════
#  FINANZAS — RECEPCIONES  (/api/recepciones/*)
# ══════════════════════════════════════════════════════════════════
RECEPCIONES_FILE = _os.path.join(_DATA, "FINANZAS", "recepciones.json")

def rec_load():
    _h = _cache_get("recepciones")
    if _h is not None: return _h
    p = Path(RECEPCIONES_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def rec_save(records):
    Path(RECEPCIONES_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(RECEPCIONES_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("recepciones", records)

def _ingresos_fiscalizados_map() -> dict:
    """Devuelve {ingreso_id: rec_number} para todo ingreso físico ya ligado a una
    Recepción fiscal previa (para no volver a facturarlo dos veces)."""
    used = {}
    for r in rec_load():
        for ing_id in r.get("ingreso_ids", []) or []:
            used[ing_id] = r.get("rec_number","")
    return used

@app.route("/api/recepciones/oc/<po_number>", methods=["GET"])
def api_recepciones_oc_lookup(po_number):
    """
    Devuelve los datos de una GPO (proveedor, folio, estatus físico de referencia)
    junto con los INGRESOS FÍSICOS (Almacenes) ya registrados para esa OC, marcando
    cuáles ya fueron fiscalizados en una Recepción previa.
    """
    try:
        po_upper = po_number.upper()
        gpos = gpo_load()
        gpo_rec = next((g for g in gpos if str(g.get("po_number","")).upper()==po_upper), None)
        if not gpo_rec:
            return jsonify({"error": f"Orden {po_number} no encontrada"}), 404

        fiscalizados = _ingresos_fiscalizados_map()
        ingresos = [i for i in ingreso_load()
                    if i.get("tipo") == "gpo" and str(i.get("po_number","")).upper() == po_upper]
        ingresos.sort(key=lambda i: i.get("created_at",""))
        ingresos_out = []
        for ing in ingresos:
            ing_copy = dict(ing)
            ing_copy["fiscalizado"] = ing["id"] in fiscalizados
            ing_copy["fiscalizado_en"] = fiscalizados.get(ing["id"], "")
            ingresos_out.append(ing_copy)

        out = dict(gpo_rec)
        out["ingresos"] = ingresos_out
        return jsonify({"record": out})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/recepciones", methods=["GET"])
def api_get_recepciones():
    try:
        records = rec_load()
        q = request.args.get("q","").lower()
        if q:
            records = [r for r in records if
                q in (r.get("rec_number","")).lower() or
                q in (r.get("po_number","")).lower() or
                q in (r.get("job","")).lower() or
                q in (r.get("cpo","")).lower() or
                q in (r.get("factura","")).lower() or
                q in (r.get("supplier_name","")).lower()]
        records = sorted(records, key=lambda r: r.get("created_at",""), reverse=True)
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/recepciones", methods=["POST"])
def api_create_recepcion():
    """
    Registra una recepción fiscal (Finanzas), ligada a uno o más ingresos
    físicos (Almacenes) ya existentes para la OC — o, si es Compra Directa,
    a items capturados manualmente (no hay ingreso físico previo que ligar).
    Body: {
      po_number, compra_directa (bool),
      ingreso_ids: [ "WI-0000000001", ... ]   (flujo normal, ligado a Almacenes)
      supplier {..}, items: [...]              (sólo Compra Directa)
      factura, notes
    }
    """
    try:
        data          = request.get_json() or {}
        po_number     = str(data.get("po_number","")).strip().upper()
        compra_directa= bool(data.get("compra_directa", False))
        factura       = str(data.get("factura","")).strip()
        notes         = str(data.get("notes","")).strip()
        supplier      = data.get("supplier", {}) or {}
        ingreso_ids   = data.get("ingreso_ids", []) or []
        job           = str(data.get("job","")).strip().upper()
        cpo           = str(data.get("cpo","")).strip().upper()

        if not factura:
            return jsonify({"error":"Debes capturar el número de factura o folio"}), 400
        if not compra_directa and not po_number:
            return jsonify({"error":"Debes indicar el número de Orden de Compra o marcar Compra Directa"}), 400

        with lock:
            now = datetime.datetime.now().isoformat()
            rec_items = []

            if compra_directa:
                items_in = data.get("items", [])
                if not items_in:
                    return jsonify({"error":"Debe incluir al menos un item recibido"}), 400
                for it in items_in:
                    qty = float(it.get("quantity_received", 0) or 0)
                    if qty <= 0: continue
                    uc  = float(it.get("unit_cost", 0) or 0)
                    rec_items.append({
                        "part_number": str(it.get("part_number","")).strip().upper(),
                        "description": str(it.get("description","")).strip(),
                        "brand":       str(it.get("brand","")).strip().upper(),
                        "job":         str(it.get("job","")).strip().upper(),
                        "quantity_received": qty,
                        "unit_cost":   uc,
                        "total":       round(qty*uc, 2),
                    })
                if not rec_items:
                    return jsonify({"error":"Ningún item con cantidad recibida > 0"}), 400
            else:
                gpos = gpo_load()
                gpo_rec = next((g for g in gpos if str(g.get("po_number","")).upper()==po_number), None)
                if not gpo_rec:
                    return jsonify({"error": f"Orden {po_number} no encontrada"}), 404
                supplier = gpo_rec.get("supplier", {})
                job = gpo_rec.get("job","")
                cpo = gpo_rec.get("cpo","")

                if not ingreso_ids:
                    return jsonify({"error":"Selecciona al menos un ingreso de almacén a fiscalizar"}), 400

                fiscalizados = _ingresos_fiscalizados_map()
                ya_usados = [i for i in ingreso_ids if i in fiscalizados]
                if ya_usados:
                    return jsonify({"error":
                        f"El ingreso {ya_usados[0]} ya fue fiscalizado en {fiscalizados[ya_usados[0]]}"}), 400

                todos_ingresos = ingreso_load()
                seleccionados = [i for i in todos_ingresos if i.get("id") in ingreso_ids]
                encontrados_ids = {i.get("id") for i in seleccionados}
                faltantes = [i for i in ingreso_ids if i not in encontrados_ids]
                if faltantes:
                    return jsonify({"error": f"Ingreso no encontrado: {faltantes[0]}"}), 404
                for ing in seleccionados:
                    if str(ing.get("po_number","")).upper() != po_number:
                        return jsonify({"error": f"El ingreso {ing.get('id')} no pertenece a la OC {po_number}"}), 400

                # Agregar items de todos los ingresos seleccionados (agrupando por No. Parte + Job)
                grouped = {}
                for ing in seleccionados:
                    for it in ing.get("items", []):
                        key = (str(it.get("part_number","")).strip().upper(), str(it.get("job","")).strip().upper())
                        qty = float(it.get("quantity_delivered", 0) or 0)
                        if qty <= 0: continue
                        uc = float(it.get("unit_cost", 0) or 0)
                        if key not in grouped:
                            grouped[key] = {
                                "part_number": key[0], "job": key[1],
                                "description": str(it.get("description","")).strip(),
                                "brand":       str(it.get("brand","")).strip().upper(),
                                "quantity_received": 0.0, "unit_cost": uc,
                            }
                        grouped[key]["quantity_received"] += qty
                for g in grouped.values():
                    g["total"] = round(g["quantity_received"] * g["unit_cost"], 2)
                    rec_items.append(g)
                if not rec_items:
                    return jsonify({"error":"Los ingresos seleccionados no tienen items con cantidad > 0"}), 400

            rec = {
                "rec_number":    _doc_next_number("REC"),
                "po_number":     po_number if not compra_directa else "COMPRA DIRECTA",
                "compra_directa":compra_directa,
                "ingreso_ids":   ingreso_ids if not compra_directa else [],
                "job":           job,
                "cpo":           cpo,
                "supplier":      supplier,
                "supplier_name": str(supplier.get("nombre","")).strip().upper(),
                "factura":       factura,
                "notes":         notes,
                "items":         rec_items,
                "total":         round(sum(i["total"] for i in rec_items), 2),
                "fecha":         now[:10],
                "created_by":    session.get("user",""),
                "created_at":    now,
            }
            recs = rec_load()
            recs.append(rec)
            rec_save(recs)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/recepciones/<rec_number>", methods=["PUT"])
def api_edit_recepcion(rec_number):
    """Edición ligera: sólo número de factura y notas (no reprocesa cantidades)."""
    try:
        data = request.get_json() or {}
        with lock:
            recs = rec_load()
            rec = next((r for r in recs if r.get("rec_number","").upper()==rec_number.upper()), None)
            if not rec:
                return jsonify({"error":"Recepción no encontrada"}), 404
            if "factura" in data:
                factura = str(data.get("factura","")).strip()
                if not factura:
                    return jsonify({"error":"El número de factura no puede quedar vacío"}), 400
                rec["factura"] = factura
            if "notes" in data:
                rec["notes"] = str(data.get("notes","")).strip()
            rec["updated_by"] = session.get("user","")
            rec["updated_at"] = datetime.datetime.now().isoformat()
            rec_save(recs)
        return jsonify({"ok": True, "record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/recepciones/<rec_number>", methods=["DELETE"])
def api_delete_recepcion(rec_number):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            recs = rec_load()
            new = [r for r in recs if r.get("rec_number","").upper() != rec_number.upper()]
            rec_save(new)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/recepciones/<rec_number>/pdf")
def api_recepcion_pdf(rec_number):
    try:
        recs = rec_load()
        rec = next((r for r in recs if r.get("rec_number","").upper()==rec_number.upper()), None)
        if not rec: return jsonify({"error":"Recepción no encontrada"}), 404
        sup   = rec.get("supplier",{})
        items = rec.get("items",[])
        fmt   = lambda v: f"${float(v):,.2f}"
        logo_path = _os.path.join(_BASE, "static", "persico_logo.webp")
        logo_b64  = ""
        if _os.path.exists(logo_path):
            import base64 as _b64
            with open(logo_path,"rb") as lf:
                logo_b64 = _b64.b64encode(lf.read()).decode()
        logo_tag = f'<img src="data:image/webp;base64,{logo_b64}" style="height:54px">' if logo_b64 else '<span style="font-size:22px;font-weight:900;color:#c8102e">PERSICO</span>'
        rows_html = ""
        for i, it in enumerate(items, 1):
            rows_html += f"""
            <tr>
              <td style="text-align:center">{i}</td>
              <td style="font-family:monospace">{it.get('part_number','—')}</td>
              <td>{it.get('brand','—')}</td>
              <td>{it.get('description','—')}</td>
              <td style="text-align:center;font-family:monospace;color:#555">{it.get('job','—')}</td>
              <td style="text-align:right">{it.get('quantity_received',0)}</td>
              <td style="text-align:right">{fmt(it.get('unit_cost',0))}</td>
              <td style="text-align:right;font-weight:600">{fmt(it.get('total',0))}</td>
            </tr>"""
        html = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>{rec['rec_number']}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Arial,sans-serif;font-size:11px;color:#111;padding:28px 32px}}
  .header{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:18px;border-bottom:3px solid #c8102e;padding-bottom:14px}}
  .rec-num{{font-size:26px;font-weight:900;color:#c8102e;letter-spacing:1px;text-align:right}}
  .rec-num-label{{font-size:10px;color:#888;text-transform:uppercase;letter-spacing:1px;text-align:right}}
  .company-block{{font-size:10px;line-height:1.6;color:#444}}
  .company-name{{font-size:13px;font-weight:700;color:#111;margin-bottom:2px}}
  .section-title{{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;color:#888;margin-bottom:4px;margin-top:12px}}
  .sup-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px 20px;margin-bottom:14px;padding:10px 14px;background:#f7f7f7;border-radius:6px;border:1px solid #e8e8e8}}
  .sup-field{{font-size:10px}}.sup-field b{{color:#444;display:block;font-size:9px;text-transform:uppercase}}
  .meta-bar{{display:flex;gap:24px;padding:8px 14px;background:#1f3864;color:#fff;border-radius:6px;margin-bottom:14px;font-size:11px;flex-wrap:wrap}}
  .meta-bar span{{opacity:.75}}.meta-bar b{{opacity:1}}
  table{{width:100%;border-collapse:collapse;margin-bottom:12px;font-size:10px}}
  thead th{{background:#1f3864;color:#fff;padding:7px 8px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.8px;white-space:nowrap}}
  tbody tr:nth-child(even){{background:#f7f7f7}}
  tbody td{{padding:6px 8px;border-bottom:1px solid #ebebeb;vertical-align:top}}
  .total-row{{display:flex;justify-content:flex-end;margin-top:4px;padding:10px 14px;background:#1f3864;color:#fff;border-radius:6px;font-size:13px;font-weight:700}}
  .footer{{margin-top:28px;border-top:1px solid #ddd;padding-top:10px;font-size:9px;color:#aaa;display:flex;justify-content:space-between}}
  @media print{{body{{padding:10px 14px}}}}
</style>
</head>
<body>
<div class="header">
  <div>
    {logo_tag}
    <div class="company-block" style="margin-top:8px">
      <div class="company-name">PERSICO MÉXICO S.A. DE C.V.</div>
      RFC: PME200101AB1<br>
      Blvd. Forjadores de Puebla 3120, Parque Industrial<br>
      Puebla, Pue. C.P. 72070 · Tel. (222) 000-0000
    </div>
  </div>
  <div>
    <div class="rec-num-label">Recepción</div>
    <div class="rec-num">{rec['rec_number']}</div>
    <div style="font-size:10px;color:#888;text-align:right;margin-top:4px">{rec['fecha']}</div>
  </div>
</div>
<div class="section-title">Datos del Proveedor</div>
<div class="sup-grid">
  <div class="sup-field"><b>Nombre</b>{sup.get('nombre','—')}</div>
  <div class="sup-field"><b>RFC</b>{sup.get('rfc','—')}</div>
  <div class="sup-field"><b>Dirección</b>{sup.get('calle','')} {sup.get('num_exterior','')} {sup.get('num_interior','')}</div>
  <div class="sup-field"><b>Teléfono</b>{sup.get('telefono','—')}</div>
</div>
<div class="meta-bar">
  <div><span>Orden de Compra</span> <b>{rec.get('po_number','—')}</b></div>
  {f'<div><span>Job</span> <b>{rec["job"]}</b></div>' if rec.get('job') else ''}
  {f'<div><span>CPO</span> <b>{rec["cpo"]}</b></div>' if rec.get('cpo') else ''}
  <div><span>No. Factura</span> <b>{rec.get('factura','—')}</b></div>
  <div><span>Tipo</span> <b>{'Compra Directa' if rec.get('compra_directa') else 'Ligada a OC'}</b></div>
  {f'<div><span>Ingresos Fiscalizados</span> <b>{", ".join(rec.get("ingreso_ids",[]))}</b></div>' if rec.get('ingreso_ids') else ''}
  <div><span>Recibido por</span> <b>{rec.get('created_by','')}</b></div>
</div>
<table>
  <thead><tr>
    <th>#</th><th>No. Parte</th><th>Marca</th><th>Descripción</th><th>Job</th>
    <th style="text-align:right">Cant. Recibida</th>
    <th style="text-align:right">Costo Unit.</th>
    <th style="text-align:right">Total</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<div class="total-row">
  <span>TOTAL&nbsp;&nbsp;{fmt(rec.get('total',0))}</span>
</div>
<div class="footer">
  <span>Generado por Persico Suite · {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</span>
  <span>{rec['rec_number']} · {rec.get('supplier_name','')}</span>
</div>
</body></html>"""
        return Response(html, mimetype="text/html",
            headers={"Content-Disposition": f"inline;filename={rec['rec_number']}.html"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  FINANZAS — PROCESAR COMPRA  (/api/procesar-compra/*)  +  CPP
# ══════════════════════════════════════════════════════════════════
PUR_FILE = _os.path.join(_DATA, "FINANZAS", "procesar_compra.json")
CPP_FILE = _os.path.join(_DATA, "FINANZAS", "cpp.json")

def pur_load():
    _h = _cache_get("pur")
    if _h is not None: return _h
    p = Path(PUR_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def pur_save(records):
    Path(PUR_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(PUR_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("pur", records)

def cpp_load():
    _h = _cache_get("cpp")
    if _h is not None: return _h
    p = Path(CPP_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def cpp_save(records):
    Path(CPP_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(CPP_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("cpp", records)

@app.route("/api/procesar-compra", methods=["GET"])
def api_get_procesar_compra():
    try:
        records = pur_load()
        q = request.args.get("q","").lower()
        if q:
            records = [r for r in records if
                q in (r.get("pur_number","")).lower() or
                q in (r.get("rec_number","")).lower() or
                q in (r.get("po_number","")).lower() or
                q in (r.get("job","")).lower() or
                q in (r.get("cpo","")).lower() or
                q in (r.get("factura","")).lower() or
                q in (r.get("usuario","")).lower()]
        records = sorted(records, key=lambda r: r.get("created_at",""), reverse=True)
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/procesar-compra/buscar/<q>", methods=["GET"])
def api_procesar_compra_buscar(q):
    """Busca una Recepción (REC) por su número de factura o su REC number."""
    try:
        q_upper = q.strip().upper()
        recs = rec_load()
        rec = next((r for r in recs if
            r.get("rec_number","").upper() == q_upper or
            r.get("factura","").upper() == q_upper), None)
        if not rec:
            return jsonify({"error": f"No se encontró ninguna Recepción con factura/folio '{q}'"}), 404

        purs = pur_load()
        ya_procesada = next((p for p in purs if p.get("rec_number","")==rec.get("rec_number")), None)
        out = dict(rec)
        out["_ya_procesada"] = bool(ya_procesada)
        out["_pur_number"] = ya_procesada.get("pur_number","") if ya_procesada else ""
        return jsonify({"record": out})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/procesar-compra", methods=["POST"])
def api_create_procesar_compra():
    """
    Registra el procesamiento de una Recepción como compra en firme.
    Genera folio PUR-000000000x y, en espejo, un registro en CPP con el
    mismo monto pero en NEGATIVO.
    Body: { rec_number }
    """
    try:
        data = request.get_json() or {}
        rec_number = str(data.get("rec_number","")).strip().upper()
        if not rec_number:
            return jsonify({"error":"Debes indicar el número de Recepción"}), 400

        with lock:
            recs = rec_load()
            rec = next((r for r in recs if r.get("rec_number","").upper()==rec_number), None)
            if not rec:
                return jsonify({"error": f"Recepción {rec_number} no encontrada"}), 404

            purs = pur_load()
            if any(p.get("rec_number","")==rec["rec_number"] for p in purs):
                return jsonify({"error": f"La Recepción {rec['rec_number']} ya fue procesada como compra"}), 400

            now = datetime.datetime.now().isoformat()
            usuario = session.get("user","")
            monto = float(rec.get("total", 0) or 0)

            pur = {
                "pur_number": _doc_next_number("PUR"),
                "rec_number": rec["rec_number"],
                "po_number":  rec.get("po_number",""),
                "job":        rec.get("job",""),
                "cpo":        rec.get("cpo",""),
                "factura":    rec.get("factura",""),
                "monto":      monto,
                "supplier_name": rec.get("supplier_name",""),
                "fecha":      now[:10],
                "usuario":    usuario,
                "created_at": now,
            }
            purs.append(pur)
            pur_save(purs)

            cpps = cpp_load()
            cpp = {
                "cpp_number": _doc_next_number("CPP"),
                "pur_number": pur["pur_number"],
                "rec_number": rec["rec_number"],
                "po_number":  rec.get("po_number",""),
                "job":        rec.get("job",""),
                "cpo":        rec.get("cpo",""),
                "factura":    rec.get("factura",""),
                "monto":      round(-monto, 2),
                "supplier_name": rec.get("supplier_name",""),
                "estatus":    "Pendiente",
                "fecha":      now[:10],
                "usuario":    usuario,
                "created_at": now,
            }
            cpps.append(cpp)
            cpp_save(cpps)

        return jsonify({"ok": True, "record": pur, "cpp_record": cpp})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _ledger_pdf_html(titulo, folio, fecha, meta_fields, monto, footer_ref):
    """Genera un HTML tipo comprobante para folios de Finanzas (PUR/CPP/PAY)."""
    fmt = lambda v: ("-" if float(v) < 0 else "") + f"${abs(float(v)):,.2f}"
    meta_html = "".join(
        f'<div class="meta-field"><b>{label}</b>{value or "—"}</div>'
        for label, value in meta_fields
    )
    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>{folio}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Arial,sans-serif;font-size:11px;color:#111;padding:28px 32px}}
  .header{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:18px;border-bottom:3px solid #c8102e;padding-bottom:14px}}
  .pur-num{{font-size:26px;font-weight:900;color:#c8102e;letter-spacing:1px;text-align:right}}
  .pur-num-label{{font-size:10px;color:#888;text-transform:uppercase;letter-spacing:1px;text-align:right}}
  .company-name{{font-size:13px;font-weight:700;color:#111}}
  .meta-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px 20px;margin-bottom:14px;padding:14px 16px;background:#f7f7f7;border-radius:6px;border:1px solid #e8e8e8}}
  .meta-field{{font-size:11px}}.meta-field b{{color:#444;display:block;font-size:9px;text-transform:uppercase}}
  .total-row{{display:flex;justify-content:flex-end;margin-top:10px;padding:12px 16px;background:#1f3864;color:#fff;border-radius:6px;font-size:15px;font-weight:700}}
  .footer{{margin-top:28px;border-top:1px solid #ddd;padding-top:10px;font-size:9px;color:#aaa;display:flex;justify-content:space-between}}
  @media print{{body{{padding:10px 14px}}}}
</style>
</head>
<body>
<div class="header">
  <div class="company-name">PERSICO MÉXICO S.A. DE C.V.</div>
  <div>
    <div class="pur-num-label">{titulo}</div>
    <div class="pur-num">{folio}</div>
    <div style="font-size:10px;color:#888;text-align:right;margin-top:4px">{fecha}</div>
  </div>
</div>
<div class="meta-grid">
  {meta_html}
</div>
<div class="total-row">
  <span>MONTO&nbsp;&nbsp;{fmt(monto)}</span>
</div>
<div class="footer">
  <span>Generado por Persico Suite · {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</span>
  <span>{footer_ref}</span>
</div>
</body></html>"""

@app.route("/api/procesar-compra/<pur_number>/pdf")
def api_procesar_compra_pdf(pur_number):
    try:
        purs = pur_load()
        pur = next((p for p in purs if p.get("pur_number","").upper()==pur_number.upper()), None)
        if not pur: return jsonify({"error":"Registro no encontrado"}), 404
        html = _ledger_pdf_html(
            "Procesar Compra", pur['pur_number'], pur.get('fecha',''),
            [("Recepción (REC)", pur.get('rec_number','')),
             ("Orden de Compra", pur.get('po_number','')),
             ("Job", pur.get('job','')),
             ("CPO", pur.get('cpo','')),
             ("No. Factura", pur.get('factura','')),
             ("Proveedor", pur.get('supplier_name','')),
             ("Procesado por", pur.get('usuario',''))],
            pur.get('monto',0),
            f"{pur['pur_number']} · {pur.get('supplier_name','')}"
        )
        return Response(html, mimetype="text/html",
            headers={"Content-Disposition": f"inline;filename={pur['pur_number']}.html"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cpp", methods=["GET"])
def api_get_cpp():
    try:
        records = cpp_load()
        q = request.args.get("q","").lower()
        if q:
            records = [r for r in records if
                q in (r.get("cpp_number","")).lower() or
                q in (r.get("pur_number","")).lower() or
                q in (r.get("po_number","")).lower() or
                q in (r.get("job","")).lower() or
                q in (r.get("cpo","")).lower() or
                q in (r.get("factura","")).lower() or
                q in (r.get("supplier_name","")).lower()]
        records = sorted(records, key=lambda r: r.get("created_at",""), reverse=True)
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cpp/<cpp_number>/lookup", methods=["GET"])
def api_cpp_lookup(cpp_number):
    try:
        cpps = cpp_load()
        rec = next((c for c in cpps if c.get("cpp_number","").upper()==cpp_number.upper()), None)
        if not rec:
            return jsonify({"error": f"CPP {cpp_number} no encontrada"}), 404
        return jsonify({"record": rec})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cpp/pendientes", methods=["GET"])
def api_cpp_pendientes():
    """Lista las CPP sin pago registrado, opcionalmente filtradas por proveedor."""
    try:
        q = request.args.get("proveedor","").lower()
        cpps = [c for c in cpp_load() if c.get("estatus","Pendiente") != "Pagado"]
        if q:
            cpps = [c for c in cpps if q in (c.get("supplier_name","")).lower()]
        cpps = sorted(cpps, key=lambda c: c.get("created_at",""))
        return jsonify({"records": cpps, "total": len(cpps)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cpp/<cpp_number>/pdf")
def api_cpp_pdf(cpp_number):
    try:
        cpps = cpp_load()
        cpp = next((c for c in cpps if c.get("cpp_number","").upper()==cpp_number.upper()), None)
        if not cpp: return jsonify({"error":"CPP no encontrada"}), 404
        html = _ledger_pdf_html(
            "CPP — Cuenta por Pagar", cpp['cpp_number'], cpp.get('fecha',''),
            [("Pur Number", cpp.get('pur_number','')),
             ("Rec Number", cpp.get('rec_number','')),
             ("PO Number", cpp.get('po_number','')),
             ("Job", cpp.get('job','')),
             ("CPO", cpp.get('cpo','')),
             ("No. Factura", cpp.get('factura','')),
             ("Proveedor", cpp.get('supplier_name','')),
             ("Estatus", cpp.get('estatus','Pendiente')),
             ("Usuario", cpp.get('usuario',''))],
            cpp.get('monto',0),
            f"{cpp['cpp_number']} · {cpp.get('supplier_name','')}"
        )
        return Response(html, mimetype="text/html",
            headers={"Content-Disposition": f"inline;filename={cpp['cpp_number']}.html"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  FINANZAS — PAGOS  (/api/pagos/*)
# ══════════════════════════════════════════════════════════════════
PAGOS_FILE = _os.path.join(_DATA, "FINANZAS", "pagos.json")

def pago_load():
    _h = _cache_get("pagos")
    if _h is not None: return _h
    p = Path(PAGOS_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def pago_save(records):
    Path(PAGOS_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(PAGOS_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("pagos", records)

@app.route("/api/pagos", methods=["GET"])
def api_get_pagos():
    try:
        records = pago_load()
        q = request.args.get("q","").lower()
        if q:
            records = [r for r in records if
                q in (r.get("pago_number","")).lower() or
                q in (r.get("cpp_number","")).lower() or
                q in (r.get("pur_number","")).lower() or
                q in (r.get("po_number","")).lower() or
                q in (r.get("job","")).lower() or
                q in (r.get("cpo","")).lower() or
                q in (r.get("factura","")).lower() or
                q in (r.get("supplier_name","")).lower()]
        records = sorted(records, key=lambda r: r.get("created_at",""), reverse=True)
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/pagos", methods=["POST"])
def api_procesar_pago():
    """
    Procesa el pago de una o varias CPP. Cada CPP genera su propio
    registro independiente en Pagos (heredando sus datos), y queda marcada
    como 'Pagado' para excluirse del Saldo Total CPP.
    Body: { cpp_numbers: ["CPP-0000000001", ...] }
    """
    try:
        data = request.get_json() or {}
        cpp_numbers = data.get("cpp_numbers", []) or []
        if not cpp_numbers:
            return jsonify({"error":"Selecciona al menos una CPP a pagar"}), 400

        with lock:
            now = datetime.datetime.now().isoformat()
            usuario = session.get("user","")
            cpps = cpp_load()
            pagos = pago_load()
            creados = []

            for cpp_number in cpp_numbers:
                cpp_number = str(cpp_number).strip().upper()
                cpp = next((c for c in cpps if c.get("cpp_number","").upper()==cpp_number), None)
                if not cpp:
                    return jsonify({"error": f"CPP {cpp_number} no encontrada"}), 404
                if cpp.get("estatus","Pendiente") == "Pagado":
                    return jsonify({"error": f"La CPP {cpp_number} ya tiene un pago registrado"}), 400

                cpp["estatus"] = "Pagado"

                pago = {
                    "pago_number":   _doc_next_number("PAY"),
                    "cpp_number":    cpp["cpp_number"],
                    "pur_number":    cpp.get("pur_number",""),
                    "rec_number":    cpp.get("rec_number",""),
                    "po_number":     cpp.get("po_number",""),
                    "job":           cpp.get("job",""),
                    "cpo":           cpp.get("cpo",""),
                    "factura":       cpp.get("factura",""),
                    "monto":         cpp.get("monto",0),
                    "supplier_name": cpp.get("supplier_name",""),
                    "confirmado":    False,
                    "fecha":         now[:10],
                    "usuario":       usuario,
                    "created_at":    now,
                }
                pagos.append(pago)
                creados.append(pago)

            cpp_save(cpps)
            pago_save(pagos)

        return jsonify({"ok": True, "records": creados})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/pagos/<pago_number>/confirmar", methods=["POST"])
def api_confirmar_pago(pago_number):
    try:
        with lock:
            pagos = pago_load()
            pago = next((p for p in pagos if p.get("pago_number","").upper()==pago_number.upper()), None)
            if not pago:
                return jsonify({"error": f"Pago {pago_number} no encontrado"}), 404
            if pago.get("confirmado"):
                return jsonify({"error": "Este pago ya está confirmado"}), 400
            pago["confirmado"] = True
            pago["confirmado_por"] = session.get("user","")
            pago["confirmado_at"] = datetime.datetime.now().isoformat()
            pago_save(pagos)
        return jsonify({"ok": True, "record": pago})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/pagos/<pago_number>/pdf")
def api_pago_pdf(pago_number):
    try:
        pagos = pago_load()
        pago = next((p for p in pagos if p.get("pago_number","").upper()==pago_number.upper()), None)
        if not pago: return jsonify({"error":"Pago no encontrado"}), 404
        html = _ledger_pdf_html(
            "Comprobante de Pago", pago['pago_number'], pago.get('fecha',''),
            [("CPP Number", pago.get('cpp_number','')),
             ("Pur Number", pago.get('pur_number','')),
             ("Rec Number", pago.get('rec_number','')),
             ("PO Number", pago.get('po_number','')),
             ("Job", pago.get('job','')),
             ("CPO", pago.get('cpo','')),
             ("No. Factura", pago.get('factura','')),
             ("Proveedor", pago.get('supplier_name','')),
             ("Confirmado", "Sí" if pago.get('confirmado') else "No"),
             ("Usuario", pago.get('usuario',''))],
            pago.get('monto',0),
            f"{pago['pago_number']} · {pago.get('supplier_name','')}"
        )
        return Response(html, mimetype="text/html",
            headers={"Content-Disposition": f"inline;filename={pago['pago_number']}.html"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route("/api/apartados", methods=["GET"])
def api_get_apartados():
    try:
        records = apartado_load()
        q   = request.args.get("q","").lower()
        job = request.args.get("job","").upper()
        if q:
            records = [r for r in records if
                q in (r.get("part_number","")).lower() or
                q in (r.get("brand","")).lower() or
                q in (r.get("description","")).lower() or
                q in (r.get("cat_code","")).lower() or
                q in (r.get("label_code","")).lower() or
                any(q in (j.get("job","")).lower() for j in r.get("jobs",[]))]
        if job:
            records = [r for r in records if
                any(j.get("job","").upper()==job for j in r.get("jobs",[]))]
        # Only return records with total_quantity > 0
        records = [r for r in records if r.get("total_quantity",0) > 0]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/disponibilidad")
def api_disponibilidad():
    """
    Calcula disponibilidad real por part_number + job:
    DISPONIBLES = INGRESOS - (WO_PENDIENTES + SURTIDOS)
    """
    try:
        job_filter = str(request.args.get("job","")).strip().upper()

        # ── 1. Sumar ingresos por (part_number, job)
        ingresos_map = {}   # (pnum, job) → qty
        for ing in ingreso_load():
            for it in ing.get("items", []):
                pnum = str(it.get("part_number","")).strip().upper()
                job  = str(it.get("job","")).strip().upper()
                if job_filter and job != job_filter: continue
                qty  = float(it.get("quantity_delivered", it.get("quantity_ordered", 0)) or 0)
                key  = (pnum, job)
                ingresos_map[key] = ingresos_map.get(key, 0) + qty

        # ── 2. Sumar salidas (pendientes + surtidas) por (part_number, job)
        salidas_map = {}    # (pnum, job) → qty total solicitada
        for sal in salida_load():
            job = str(sal.get("job","")).strip().upper()
            if job_filter and job != job_filter: continue
            for it in sal.get("items", []):
                pnum = str(it.get("part_number","")).strip().upper()
                qty  = float(it.get("quantity", 0) or 0)
                key  = (pnum, job)
                salidas_map[key] = salidas_map.get(key, 0) + qty

        # ── 3. Disponible = ingresos - salidas
        # Enrich with meta from apartados for description / cost
        apt_meta = {}
        for apt in apartado_load():
            pnum = str(apt.get("part_number","")).strip().upper()
            apt_meta[pnum] = {
                "brand":       apt.get("brand",""),
                "description": apt.get("description",""),
                "cat_code":    apt.get("cat_code",""),
                "label_code":  apt.get("label_code",""),
            }
            for j in apt.get("jobs", []):
                if j.get("unit_cost"):
                    apt_meta[pnum]["unit_cost"] = j.get("unit_cost", 0)

        result = []
        all_keys = set(ingresos_map.keys()) | set(salidas_map.keys())
        for (pnum, job) in sorted(all_keys):
            ingresos = ingresos_map.get((pnum, job), 0)
            salidas  = salidas_map.get((pnum, job), 0)
            disp     = max(0, ingresos - salidas)
            meta     = apt_meta.get(pnum, {})
            result.append({
                "part_number": pnum,
                "job":         job,
                "ingresos":    ingresos,
                "salidas":     salidas,
                "disponible":  disp,
                "brand":       meta.get("brand",""),
                "description": meta.get("description",""),
                "cat_code":    meta.get("cat_code",""),
                "label_code":  meta.get("label_code",""),
                "unit_cost":   meta.get("unit_cost", 0),
            })

        return jsonify({"records": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def api_apartado_edit_qty():
    """Super admin: correct quantity for a specific part_number + job in apartados."""
    if not is_admin():
        return jsonify({"error": "Sin permiso — solo administradores"}), 403
    try:
        data       = request.get_json()
        pnum       = str(data.get("part_number","")).strip().upper()
        job        = str(data.get("job","")).strip().upper()
        new_qty    = float(data.get("quantity", 0))
        unit_cost  = float(data.get("unit_cost", 0) or 0)
        if not pnum: return jsonify({"error": "part_number requerido"}), 400
        if new_qty < 0: return jsonify({"error": "Cantidad no puede ser negativa"}), 400
        with lock:
            records = apartado_load()
            rec = next((r for r in records if r.get("part_number","").upper() == pnum), None)
            if not rec:
                return jsonify({"error": f"No se encontró apartado para {pnum}"}), 404
            job_entry = next((j for j in rec.get("jobs",[]) if j.get("job","").upper() == job), None)
            if not job_entry:
                return jsonify({"error": f"No se encontró Job {job} en {pnum}"}), 404
            old_qty = job_entry.get("quantity", 0)
            job_entry["quantity"] = new_qty
            if unit_cost > 0:
                job_entry["unit_cost"] = unit_cost
            # Recalc total_quantity
            rec["total_quantity"] = sum(float(j.get("quantity",0)) for j in rec.get("jobs",[]))
            rec["updated_at"]     = datetime.datetime.now().isoformat()
            rec["_corrected_by"]  = session.get("user","")
            apartado_save(records)
        return jsonify({"ok": True, "part_number": pnum, "job": job,
                        "old_qty": old_qty, "new_qty": new_qty})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def api_delete_apartado(part_number):
    """Delete an entire apartado record by part_number (admin only)."""
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        job = request.args.get("job","").upper()
        with lock:
            records = apartado_load()
            if job:
                # Delete only the specific job entry from the part
                for r in records:
                    if r.get("part_number","").upper() == part_number.upper():
                        r["jobs"] = [j for j in r.get("jobs",[])
                                     if j.get("job","").upper() != job]
                        r["total_quantity"] = sum(
                            float(j.get("quantity",0)) for j in r["jobs"])
                        break
            else:
                # Delete entire part record
                records = [r for r in records
                           if r.get("part_number","").upper() != part_number.upper()]
            apartado_save(records)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500




# ══════════════════════════════════════════════════════════════════
#  NUMERACIÓN SECUENCIAL DE DOCUMENTOS
# ══════════════════════════════════════════════════════════════════
DOC_COUNTERS_FILE = _os.path.join(_DATA, "doc_counters.json")
_doc_lock = __import__('threading').Lock()   # separate lock — avoids deadlock with main `lock`

def _doc_counter_load():
    p = Path(DOC_COUNTERS_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: pass
    return {}

def _doc_next_number(prefix: str) -> str:
    """Returns next sequential number like WI-0000000001. Uses its own lock."""
    with _doc_lock:
        counters = _doc_counter_load()
        n = counters.get(prefix, 0) + 1
        counters[prefix] = n
        Path(DOC_COUNTERS_FILE).parent.mkdir(parents=True, exist_ok=True)
        with open(DOC_COUNTERS_FILE,"w",encoding="utf-8") as f:
            json.dump(counters, f)
    return f"{prefix}-{str(n).zfill(10)}"


SALIDA_FILE = _os.path.join(_DATA, "salidas.json")

def salida_load():
    _h = _cache_get("salida")
    if _h is not None: return _h
    p = Path(SALIDA_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def salida_save(records):
    Path(SALIDA_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(SALIDA_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    _cache_set("salida", records)

@app.route("/api/salida", methods=["GET"])
def api_get_salida():
    try:
        records = salida_load()
        q      = request.args.get("q","").lower()
        status = request.args.get("status","")
        if q:
            records = [r for r in records if
                q in (r.get("job","")).lower() or
                q in (r.get("solicitante","")).lower() or
                any(q in (i.get("part_number","")).lower() or
                    q in (i.get("brand","")).lower()
                    for i in r.get("items",[]))]
        if status:
            records = [r for r in records if r.get("status","") == status]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/salida", methods=["POST"])
def api_create_salida():
    try:
        data       = request.get_json()
        job        = str(data.get("job","")).strip().upper()
        solicitante= str(data.get("solicitante","")).strip()
        items      = data.get("items", [])
        if not job:   return jsonify({"error":"Job requerido"}), 400
        if not items: return jsonify({"error":"Selecciona al menos un item"}), 400
        with lock:
            now = datetime.datetime.now().isoformat()
            sal_items = []
            for it in items:
                qty = float(it.get("quantity",0))
                uc  = float(it.get("unit_cost",0) or 0)
                if qty <= 0: continue
                sal_items.append({
                    "part_number": str(it.get("part_number","")).strip().upper(),
                    "brand":       str(it.get("brand","")).strip().upper(),
                    "description": str(it.get("description","")).strip(),
                    "cat_code":    str(it.get("cat_code","")).strip().upper(),
                    "label_code":  str(it.get("label_code","")).strip().upper(),
                    "quantity":    qty, "unit_cost": uc, "total": round(qty*uc,2),
                })
            if not sal_items: return jsonify({"error":"Ningún item con cantidad > 0"}), 400
            rec = {
                "id": _doc_next_number("WO"),
                "job": job, "solicitante": solicitante, "items": sal_items,
                "status": "Pendiente", "fecha": now[:10], "created_at": now,
                "surtido_at": None, "surtido_by": None,
            }
            records = salida_load(); records.append(rec); salida_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/salida/<sal_id>/surtir", methods=["POST"])
def api_surtir_salida(sal_id):
    try:
        with lock:
            records = salida_load()
            rec = next((r for r in records if r["id"]==sal_id), None)
            if not rec: return jsonify({"error":"Salida no encontrada"}), 404
            if rec.get("status")=="Surtida": return jsonify({"error":"Ya fue surtida"}), 400
            now = datetime.datetime.now().isoformat()
            apartados = apartado_load()
            for it in rec.get("items",[]):
                pnum=it["part_number"].upper(); job=rec["job"].upper(); qty=float(it.get("quantity",0))
                apt=next((a for a in apartados if a.get("part_number","").upper()==pnum),None)
                if apt:
                    je=next((j for j in apt.get("jobs",[]) if j.get("job","").upper()==job),None)
                    if je:
                        je["quantity"]=max(0,float(je.get("quantity",0))-qty)
                        if je["quantity"]<=0: apt["jobs"]=[j for j in apt.get("jobs",[]) if j.get("job","").upper()!=job]
                    apt["total_quantity"]=max(0,sum(float(j.get("quantity",0)) for j in apt.get("jobs",[])))
                    apt["updated_at"]=now
            apartado_save(apartados)
            rec["status"]="Surtida"; rec["surtido_at"]=now; rec["surtido_by"]=session.get("user","")
            salida_save(records)
        return jsonify({"ok": True, "record": rec})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/salida/<sal_id>", methods=["DELETE"])
def api_delete_salida(sal_id):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            records=salida_load(); new=[r for r in records if r["id"]!=sal_id]
            if len(new)==len(records): return jsonify({"error":"Salida no encontrada"}), 404
            salida_save(new)
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/salida/<sal_id>/pdf")
def api_salida_pdf(sal_id):
    try:
        records=salida_load()
        rec=next((r for r in records if r["id"]==sal_id),None)
        if not rec: return jsonify({"error":"Salida no encontrada"}),404
        logo_path=_os.path.join(_BASE,"static","persico_logo.webp"); logo_b64=""
        if _os.path.exists(logo_path):
            import base64 as _b64
            with open(logo_path,"rb") as lf: logo_b64=_b64.b64encode(lf.read()).decode()
        logo_tag=f'<img src="data:image/webp;base64,{logo_b64}" style="height:48px">' if logo_b64 else '<b style="color:#c8102e">PERSICO</b>'
        fmt=lambda v:f"${float(v):,.2f}"
        items=rec.get("items",[]); subtotal=sum(float(it.get("total",0)) for it in items)
        sc="#48c78e" if rec.get("status")=="Surtida" else "#ffdd57"
        si=f'<div style="display:flex;gap:24px;padding:8px 14px;background:#1f3864;color:#fff;border-radius:6px;margin-bottom:14px;font-size:11px"><div>Surtido por <b>{rec.get("surtido_by","—")}</b></div><div>Fecha <b>{(rec.get("surtido_at","") or "")[:10]}</b></div></div>' if rec.get("status")=="Surtida" else ""
        rows="".join(f"<tr><td style='text-align:center'>{i+1}</td><td style='font-family:monospace'>{it.get('cat_code','—')}</td><td style='font-family:monospace'>{it.get('part_number','—')}</td><td>{it.get('brand','—')}</td><td>{it.get('description','—')}</td><td style='text-align:right'>{int(it.get('quantity',0))}</td><td style='text-align:right'>{fmt(it.get('unit_cost',0))}</td><td style='text-align:right;font-weight:600'>{fmt(it.get('total',0))}</td></tr>" for i,it in enumerate(items))
        html=f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>{rec['id']}</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:Arial,sans-serif;font-size:11px;padding:28px}}
table{{width:100%;border-collapse:collapse;margin-bottom:12px}}thead th{{background:#1f3864;color:#fff;padding:7px 8px;font-size:9px;text-transform:uppercase;text-align:left}}
tbody tr:nth-child(even){{background:#f7f7f7}}td{{padding:6px 8px;border-bottom:1px solid #eee}}</style></head>
<body>
<div style="display:flex;justify-content:space-between;border-bottom:3px solid #c8102e;padding-bottom:14px;margin-bottom:16px">
  <div>{logo_tag}</div>
  <div style="text-align:right"><div style="font-size:20px;font-weight:900;color:#c8102e">{rec['id']}</div>
  <div style="font-size:10px;color:#888">{rec.get('fecha','')} · <span style="background:{sc}20;color:{sc};padding:2px 8px;border-radius:4px;font-weight:700">{rec.get('status','—')}</span></div></div>
</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px;padding:10px;background:#f7f7f7;border-radius:6px;font-size:10px">
<div><b>Job:</b> {rec.get('job','—')}</div><div><b>Solicitante:</b> {rec.get('solicitante','—')}</div>
</div>
{si}
<table><thead><tr><th>#</th><th>Código</th><th>No. Parte</th><th>Marca</th><th>Descripción</th><th style="text-align:right">Cant.</th><th style="text-align:right">Costo Unit.</th><th style="text-align:right">Total</th></tr></thead>
<tbody>{rows}</tbody></table>
<div style="display:flex;justify-content:flex-end;gap:20px;padding:10px 14px;background:#1f3864;color:#fff;border-radius:6px;font-size:13px;font-weight:700;margin-bottom:28px"><span>TOTAL USD</span><span>{fmt(subtotal)}</span></div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:40px;margin-top:20px">
<div style="border-top:1px solid #aaa;margin-top:50px;padding-top:6px;font-size:10px;color:#555;text-align:center"><b>{rec.get('solicitante','—')}</b><br>Solicitante</div>
<div style="border-top:1px solid #aaa;margin-top:50px;padding-top:6px;font-size:10px;color:#555;text-align:center"><b>{rec.get('surtido_by','—') if rec.get('status')=='Surtida' else '____________________'}</b><br>Almacenista</div>
</div>
</body></html>"""
        return Response(html,mimetype="text/html",headers={"Content-Disposition":f"inline;filename={rec['id']}.html"})
    except Exception as e: return jsonify({"error":str(e)}),500

import _io as _io_mod

VIATICOS_FILE = _os.path.join(_DATA, "viaticos.json")
GASTOS_FILE   = _os.path.join(_DATA, "gastos_viaje.json")
ENVIOS_FILE   = _os.path.join(_DATA, "envios.json")
PODS_FOLDER   = _os.path.join(_DATA, "envio_pods")

def _svc_load(path):
    _h = _cache_get("svc_"+str(path))
    if _h is not None: return _h
    p = Path(path)
    if p.exists():
        try:
            with open(p, "r", encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def _svc_save(path, data):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _get_fx_for_date(date_str):
    try:
        d    = datetime.datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
        year = d.year
        for yr in [year, year-1]:
            fx = fx_load(yr)
            if not fx: continue
            if str(date_str)[:10] in fx: return fx[str(date_str)[:10]]
            closest = max((k for k in fx if k <= str(date_str)[:10]), default=None)
            if closest: return fx[closest]
    except: pass
    _cache_set("ingreso", None)
    _cache_set("salida", None)
    _cache_set("apartado", None)
    _cache_set("svc_"+str(path), None)
    return None

def _parse_money(val):
    """Parse a value that may contain $, commas, or be a plain float."""
    if val is None: return 0.0
    try: return float(val)
    except:
        try: return float(str(val).replace('$','').replace(',','').strip())
        except: return 0.0

# ── VIÁTICOS
@app.route("/api/viaticos", methods=["GET"])
def api_get_viaticos():
    try:
        records = _svc_load(VIATICOS_FILE)
        q   = request.args.get("q","").lower()
        job = request.args.get("job","").upper()
        if q:
            records = [r for r in records if
                q in (r.get("tipo_movimiento","")).lower() or
                q in (r.get("job","")).lower() or
                q in (r.get("fecha","")).lower()]
        if job:
            records = [r for r in records if r.get("job","").upper() == job]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/viaticos", methods=["POST"])
def api_create_viatico():
    try:
        data = request.get_json()
        rows = data if isinstance(data, list) else [data]
        with lock:
            records = _svc_load(VIATICOS_FILE)
            for d in rows:
                fecha = str(d.get("fecha","")).strip()[:10]
                monto = _parse_money(d.get("monto", 0))
                tc    = _parse_money(d.get("tipo_cambio", 0)) or _get_fx_for_date(fecha) or 1.0
                usd   = round(monto / tc, 4) if tc else monto
                records.append({
                    "id":              f"VIA-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                    "id_externo":      str(d.get("id_externo","")).strip(),
                    "fecha":           fecha,
                    "tipo_movimiento": str(d.get("tipo_movimiento","")).strip(),
                    "monto":           monto,
                    "tipo_cambio":     tc,
                    "valor_usd":       usd,
                    "job":             str(d.get("job","")).strip().upper(),
                    "notas":           str(d.get("notas","")).strip(),
                    "created_by":      session.get("user",""),
                    "created_at":      datetime.datetime.now().isoformat(),
                })
            _svc_save(VIATICOS_FILE, records)
        return jsonify({"ok": True, "added": len(rows)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/viaticos/<vid>", methods=["DELETE"])
def api_delete_viatico(vid):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            recs = _svc_load(VIATICOS_FILE)
            _svc_save(VIATICOS_FILE, [r for r in recs if r.get("id") != vid])
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/viaticos/import", methods=["POST"])
def api_import_viaticos():
    try:
        import io, openpyxl
        f = request.files.get("file")
        if not f: return jsonify({"error":"No se recibió archivo"}), 400
        wb   = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws   = wb.active
        rows_data = list(ws.iter_rows(values_only=True))
        if not rows_data: return jsonify({"error":"Archivo vacío"}), 400
        hdr = [str(c).lower().strip() if c else '' for c in rows_data[0]]
        def ci(kw): return next((i for i,h in enumerate(hdr) if kw in h), None)
        ci_id   = ci('id')
        ci_fecha= ci('fecha')
        ci_tipo = ci('tipo')
        ci_monto= ci('monto')
        ci_tc   = ci('cambio')
        ci_usd  = ci('usd')
        ci_job  = ci('job')
        with lock:
            records = _svc_load(VIATICOS_FILE)
            added = 0
            for row in rows_data[1:]:
                def cv(i): return row[i] if i is not None and i < len(row) else None
                fecha_raw = cv(ci_fecha)
                if not fecha_raw: continue
                try:
                    fecha = datetime.datetime.strptime(str(fecha_raw).split('T')[0].split(' ')[0], "%Y-%m-%d").strftime("%Y-%m-%d")
                except:
                    try: fecha = datetime.datetime.strptime(str(fecha_raw), "%d/%m/%Y").strftime("%Y-%m-%d")
                    except: continue
                monto = _parse_money(cv(ci_monto))
                if monto == 0: continue
                job_xl = str(cv(ci_job) or '').strip().upper()
                if not job_xl: continue   # skip rows with no job
                # TC: from Excel or auto-fetch
                tc_raw = cv(ci_tc)
                if tc_raw and str(tc_raw).strip() not in ('', 'nan', 'NaN', 'None'):
                    tc = _parse_money(tc_raw) or _get_fx_for_date(fecha) or 1.0
                else:
                    tc = _get_fx_for_date(fecha) or 1.0
                # USD: from Excel or calculate
                usd_raw = cv(ci_usd)
                if usd_raw and str(usd_raw).strip() not in ('', 'nan', 'NaN', 'None'):
                    usd = _parse_money(usd_raw) or round(monto / tc, 4)
                else:
                    usd = round(monto / tc, 4)
                records.append({
                    "id":              f"VIA-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                    "id_externo":      str(cv(ci_id) or '').strip(),
                    "fecha":           fecha,
                    "tipo_movimiento": str(cv(ci_tipo) or '').strip(),
                    "monto":           monto,
                    "tipo_cambio":     tc,
                    "valor_usd":       usd,
                    "job":             job_xl,
                    "notas":           "",
                    "created_by":      session.get("user",""),
                    "created_at":      datetime.datetime.now().isoformat(),
                })
                added += 1
            _svc_save(VIATICOS_FILE, records)
        return jsonify({"ok": True, "added": added})
    except Exception as e: return jsonify({"error": str(e)}), 500

# ── GASTOS DE VIAJE
@app.route("/api/gastos-viaje", methods=["GET"])
def api_get_gastos():
    try:
        records = _svc_load(GASTOS_FILE)
        q   = request.args.get("q","").lower()
        job = request.args.get("job","").upper()
        if q:
            records = [r for r in records if
                q in (r.get("tipo_gasto","")).lower() or q in (r.get("job","")).lower()]
        if job:
            records = [r for r in records if r.get("job","").upper() == job]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/gastos-viaje", methods=["POST"])
def api_create_gasto():
    try:
        data   = request.get_json()
        fecha  = str(data.get("fecha","")).strip()[:10]
        moneda = str(data.get("moneda","USD")).upper()
        costo  = _parse_money(data.get("costo", 0))
        tc     = _parse_money(data.get("tipo_cambio", 0)) or (_get_fx_for_date(fecha) if moneda=="MXN" else 1.0) or 1.0
        usd    = round(costo / tc, 4) if moneda == "MXN" else costo
        with lock:
            records = _svc_load(GASTOS_FILE)
            records.append({
                "id":         f"GV-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                "tipo_gasto": str(data.get("tipo_gasto","")).strip(),
                "fecha":      fecha,
                "moneda":     moneda,
                "tipo_cambio":tc,
                "costo":      costo,
                "valor_usd":  usd,
                "job":        str(data.get("job","")).strip().upper(),
                "notas":      str(data.get("notas","")).strip(),
                "created_by": session.get("user",""),
                "created_at": datetime.datetime.now().isoformat(),
            })
            _svc_save(GASTOS_FILE, records)
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/gastos-viaje/<gid>", methods=["DELETE"])
def api_delete_gasto(gid):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            recs = _svc_load(GASTOS_FILE)
            _svc_save(GASTOS_FILE, [r for r in recs if r.get("id") != gid])
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 500

# ── ENVÍOS DE MENSAJERÍA
@app.route("/api/envios", methods=["GET"])
def api_get_envios():
    try:
        records = _svc_load(ENVIOS_FILE)
        q   = request.args.get("q","").lower()
        job = request.args.get("job","").upper()
        if q:
            records = [r for r in records if
                q in (r.get("tracking","")).lower() or q in (r.get("job","")).lower()]
        if job:
            records = [r for r in records if r.get("job","").upper() == job]
        return jsonify({"records": records, "total": len(records)})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/envios", methods=["POST"])
def api_create_envio():
    try:
        data   = request.get_json()
        fecha  = str(data.get("fecha","")).strip()[:10]
        moneda = str(data.get("moneda","USD")).upper()
        costo  = _parse_money(data.get("costo", 0))
        tc     = _parse_money(data.get("tipo_cambio", 0)) or (_get_fx_for_date(fecha) if moneda=="MXN" else 1.0) or 1.0
        usd    = round(costo / tc, 4) if moneda == "MXN" else costo
        rec_id = f"ENV-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}"
        with lock:
            records = _svc_load(ENVIOS_FILE)
            records.append({
                "id":         rec_id,
                "fecha":      fecha,
                "tracking":   str(data.get("tracking","")).strip().upper(),
                "moneda":     moneda,
                "tipo_cambio":tc,
                "costo":      costo,
                "valor_usd":  usd,
                "job":        str(data.get("job","")).strip().upper(),
                "notas":      str(data.get("notas","")).strip(),
                "pod_file":   "",
                "created_by": session.get("user",""),
                "created_at": datetime.datetime.now().isoformat(),
            })
            _svc_save(ENVIOS_FILE, records)
        return jsonify({"ok": True, "record": {"id": rec_id}})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/envios/<eid>/pod", methods=["POST"])
def api_upload_pod(eid):
    try:
        f = request.files.get("file")
        if not f: return jsonify({"error":"Sin archivo"}), 400
        _os.makedirs(PODS_FOLDER, exist_ok=True)
        ext      = _os.path.splitext(f.filename)[1].lower()
        filename = f"{eid}{ext}"
        f.save(_os.path.join(PODS_FOLDER, filename))
        with lock:
            records = _svc_load(ENVIOS_FILE)
            for r in records:
                if r.get("id") == eid: r["pod_file"] = filename
            _svc_save(ENVIOS_FILE, records)
        return jsonify({"ok": True, "filename": filename})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/envios/<eid>/pod/view")
def api_view_pod(eid):
    try:
        records = _svc_load(ENVIOS_FILE)
        rec = next((r for r in records if r.get("id") == eid), None)
        if not rec or not rec.get("pod_file"): return jsonify({"error":"Sin POD"}), 404
        from flask import send_file as _sf
        return _sf(_os.path.join(PODS_FOLDER, rec["pod_file"]))
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/envios/<eid>", methods=["DELETE"])
def api_delete_envio(eid):
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        with lock:
            recs = _svc_load(ENVIOS_FILE)
            _svc_save(ENVIOS_FILE, [r for r in recs if r.get("id") != eid])
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/servicio/por-job/<job_number>")
def api_servicio_por_job(job_number):
    try:
        job = job_number.upper()
        via = round(sum(r.get("valor_usd",0) for r in _svc_load(VIATICOS_FILE) if r.get("job","").upper()==job), 4)
        gas = round(sum(r.get("valor_usd",0) for r in _svc_load(GASTOS_FILE)   if r.get("job","").upper()==job), 4)
        env = round(sum(r.get("valor_usd",0) for r in _svc_load(ENVIOS_FILE)   if r.get("job","").upper()==job), 4)
        return jsonify({"job":job,"viaticos_usd":via,"gastos_usd":gas,"envios_usd":env,"total_usd":round(via+gas+env,4)})
    except Exception as e: return jsonify({"error": str(e)}), 500

def _build_gantt_svg(timing):
    """Build SVG Gantt from timing data list."""
    import datetime as _dt2
    entries, end_map = [], {}
    for t in timing:
        act  = t.get("actividad","")
        prev = t.get("actividad_previa","")
        fini = t.get("fecha_inicial","")
        dias = int(t.get("dias_estimados",0) or 0)
        cumpl= bool(t.get("cumplido",False))
        mile = bool(t.get("milestone_facturacion",False))
        if not act: continue
        start = None
        if fini:
            try: start = _dt2.datetime.strptime(fini[:10], "%Y-%m-%d")
            except: pass
        if not start and prev in end_map:
            start = end_map[prev] + _dt2.timedelta(days=1)
        if not start: continue
        end = start + _dt2.timedelta(days=max(dias,1))
        end_map[act] = end
        entries.append((act, start, end, dias, cumpl, mile))
    if not entries: return ""
    today  = _dt2.datetime.now().replace(hour=0,minute=0,second=0,microsecond=0)
    min_d  = min(e[1] for e in entries)
    max_d  = max(e[2] for e in entries)
    total_s= max(1,(max_d-min_d).total_seconds())
    W,ROW,PAD,LABEL = 580,18,4,160
    def px(d): return (d-min_d).total_seconds()/total_s*W
    rows=""
    for i,(act,start,end,dias,cumpl,mile) in enumerate(entries):
        x=px(start);w=max(4,px(end)-x);y=PAD+i*(ROW+2)
        color='#48c78e' if cumpl else ('#c8102e' if end<today else '#f5a623')
        label=(act[:22]+'…') if len(act)>23 else act
        micon=' ★' if mile else ''
        rows+=f'<rect x="{LABEL+x:.1f}" y="{y+2}" width="{w:.1f}" height="{ROW-4}" rx="3" fill="{color}" opacity=".9"/>'
        rows+=f'<text x="{LABEL+x+4:.1f}" y="{y+ROW-5}" font-size="8" fill="white" font-family="Arial">{dias}d</text>'
        rows+=f'<text x="{LABEL-3:.1f}" y="{y+ROW-5}" font-size="9" fill="#333" text-anchor="end" font-family="Arial">{label}{micon}</text>'
    todayX=px(today)
    ht=PAD+len(entries)*(ROW+2)
    if 0<=todayX<=W:
        rows+=f'<line x1="{LABEL+todayX:.1f}" y1="0" x2="{LABEL+todayX:.1f}" y2="{ht}" stroke="#c8102e" stroke-width="1.5" stroke-dasharray="4,2"/>'
        rows+=f'<text x="{LABEL+todayX+3:.1f}" y="10" font-size="8" fill="#c8102e" font-family="Arial">Hoy</text>'
    return f'<svg width="{LABEL+W+10}" height="{ht+8}" xmlns="http://www.w3.org/2000/svg" style="max-width:100%">{rows}</svg>'

@app.route("/api/report/executive-pdf")
def api_executive_pdf():
    """Executive HTML report for a single job — opens print dialog."""
    try:
        job_number = request.args.get("job","").strip().upper()
        rate_year  = int(request.args.get("rate_year", CURRENT_YEAR))
        wh_year    = int(request.args.get("wh_year",   CURRENT_YEAR))
        po_year    = int(request.args.get("po_year",   CURRENT_YEAR))
        if not job_number: return jsonify({"error":"Job requerido"}), 400

        d = _build_report_data(job_number, rate_year, wh_year, po_year)

        # Project config
        proj_cfg, timing = None, []
        for cfg in projcfg_load():
            jc = next((j for j in cfg.get("jobs",[])
                if j.get("job_number","").upper()==job_number), None)
            if jc: proj_cfg=jc; timing=cfg.get("timing",[]); break

        fmt = lambda v: f"${float(v):,.2f}"
        logo_path = _os.path.join(_BASE,"static","persico_logo.webp")
        logo_b64=""
        if _os.path.exists(logo_path):
            import base64 as _b64
            with open(logo_path,"rb") as lf: logo_b64=_b64.b64encode(lf.read()).decode()
        logo_tag=f'<img src="data:image/webp;base64,{logo_b64}" style="height:44px">' if logo_b64 else '<b style="color:#c8102e;font-size:20px">PERSICO</b>'

        gantt_svg = _build_gantt_svg(timing)
        revenue   = d["revenue"]
        wh_cost   = d["amount_wh"]
        pur_total = d["purchasing_total"]
        svc_via   = d.get("svc_viaticos",0)
        svc_gas   = d.get("svc_gastos",0)
        svc_env   = d.get("svc_envios",0)
        svc_total = d.get("svc_total",0)
        reas_tot  = d.get("reassign_total",0)
        recov_tot = d.get("recovery_total",0)
        gm        = d["gross_margin"]
        gm_pct    = d["gm_pct"]
        presDisp  = float(proj_cfg.get("presupuesto_disponible",0)) if proj_cfg else revenue
        gm_op     = presDisp - wh_cost - pur_total - svc_total - reas_tot + recov_tot
        gm_op_pct = (gm_op/presDisp*100) if presDisp else 0
        pct       = lambda v: f"{(v/presDisp*100 if presDisp else 0):.1f}%"

        po_rows="".join(f"<tr><td style='padding:5px 8px;font-family:monospace;font-size:10px'>{r.get('clave','—')}</td><td style='padding:5px 8px;font-size:11px'>{r.get('nombre','—')}</td><td style='padding:5px 8px;text-align:right;font-family:monospace'>{fmt(r.get('subtotal_usd',r.get('subtotal',0)))}</td></tr>" for r in d["po_items"])

        targ_html=""
        if proj_cfg:
            targ_html=f"""
            <div class="st">🎯 Targets del Proyecto</div>
            <table><thead><tr><th>Campo</th><th style="text-align:right">Valor</th><th>Campo</th><th style="text-align:right">Valor</th></tr></thead>
            <tbody>
              <tr><td>Presupuesto A</td><td style="text-align:right">{fmt(proj_cfg.get('presupuesto_a',0))}</td>
                  <td>Presupuesto Disponible</td><td style="text-align:right;font-weight:700;color:#1f3864">{fmt(presDisp)}</td></tr>
              <tr><td>Target Compras</td><td style="text-align:right">{fmt(proj_cfg.get('target_compras',0))}</td>
                  <td>Target M.O.</td><td style="text-align:right">{fmt(proj_cfg.get('target_mo',0))}</td></tr>
              <tr><td>Margen Estimado</td><td style="text-align:right">{proj_cfg.get('margen_pct',0)}%</td>
                  <td>Margen de Ahorro</td><td style="text-align:right">{proj_cfg.get('ahorro_pct',0)}%</td></tr>
            </tbody></table>"""

        svc_rows=""
        if svc_via>0: svc_rows+=f"<tr><td>💵 Viáticos</td><td style='text-align:right'>{fmt(svc_via)}</td><td style='text-align:right'>{pct(svc_via)}</td></tr>"
        if svc_gas>0: svc_rows+=f"<tr><td>✈ Gastos de Viaje</td><td style='text-align:right'>{fmt(svc_gas)}</td><td style='text-align:right'>{pct(svc_gas)}</td></tr>"
        if svc_env>0: svc_rows+=f"<tr><td>📦 Envíos Mensajería</td><td style='text-align:right'>{fmt(svc_env)}</td><td style='text-align:right'>{pct(svc_env)}</td></tr>"
        if reas_tot>0: svc_rows+=f"<tr><td>🔀 Reasignaciones</td><td style='text-align:right;color:#c8102e'>{fmt(reas_tot)}</td><td style='text-align:right'>{pct(reas_tot)}</td></tr>"
        if recov_tot>0: svc_rows+=f"<tr><td>♻ Recuperaciones</td><td style='text-align:right;color:green'>-{fmt(recov_tot)}</td><td style='text-align:right;color:green'>-{pct(recov_tot)}</td></tr>"

        html=f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<title>Reporte Ejecutivo — {job_number}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:Arial,sans-serif;font-size:11px;color:#111;padding:24px 28px}}
.header{{display:flex;justify-content:space-between;border-bottom:3px solid #c8102e;padding-bottom:12px;margin-bottom:14px;align-items:flex-start}}
.st{{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;color:#888;margin:14px 0 7px;border-bottom:1px solid #eee;padding-bottom:3px}}
.cg{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:4px}}
.c{{padding:10px 14px;border-radius:6px;border:1px solid #e0e0e0}}
.c.bl{{background:#1f3864;color:#fff;border-color:#1f3864}}.c.gn{{background:rgba(72,199,142,.08);border-color:rgba(72,199,142,.4)}}.c.rd{{background:rgba(200,16,46,.06);border-color:rgba(200,16,46,.3)}}
.cl{{font-size:9px;text-transform:uppercase;letter-spacing:1px;opacity:.7;margin-bottom:4px}}.cv{{font-family:monospace;font-size:16px;font-weight:700}}.cs{{font-size:9px;opacity:.6;margin-top:2px}}
table{{width:100%;border-collapse:collapse;margin-bottom:8px;font-size:10px}}
th{{background:#1f3864;color:#fff;padding:6px 8px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.7px}}
td{{padding:5px 8px;border-bottom:1px solid #ebebeb}}tr:nth-child(even) td{{background:#f9f9f9}}
.gantt{{overflow-x:auto;border:1px solid #eee;border-radius:6px;padding:8px;background:#fafafa}}
.footer{{margin-top:18px;border-top:1px solid #ddd;padding-top:8px;font-size:8px;color:#aaa;display:flex;justify-content:space-between}}
@media print{{body{{padding:10px 12px}}}}
</style></head><body>
<div class="header">
  <div>{logo_tag}<div style="margin-top:6px;font-size:10px;color:#555">PERSICO MÉXICO S.A. DE C.V. · Puebla, Pue.</div></div>
  <div style="text-align:right">
    <div style="font-size:18px;font-weight:900;color:#c8102e">Reporte Ejecutivo</div>
    <div style="font-size:13px;font-weight:700;margin-top:2px">Job: {job_number}</div>
    <div style="font-size:10px;color:#888">{d.get('customer','—')} · {d.get('description','—')}</div>
    <div style="font-size:10px;color:#888">{datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
  </div>
</div>

<div class="st">📊 Resultado Financiero y Operativo</div>
<div class="cg">
  <div class="c bl"><div class="cl">Revenue</div><div class="cv">{fmt(revenue)}</div><div class="cs">{d.get('customer','—')}</div></div>
  <div class="c"><div class="cl">Work Hours Cost</div><div class="cv">{fmt(wh_cost)}</div><div class="cs">{d.get('accum_hours',0):.1f} hrs · {len(d.get('workers',[]))} emp.</div></div>
  <div class="c"><div class="cl">Purchasings + Servicios</div><div class="cv">{fmt(pur_total+svc_total)}</div><div class="cs">{len(d.get('po_items',[]))} OC(s)</div></div>
  <div class="c {'gn' if gm>=0 else 'rd'}"><div class="cl">Gross Margin Financiero</div><div class="cv">{fmt(gm)}</div><div class="cs">{'▲' if gm>=0 else '▼'} {gm_pct:.1f}%</div></div>
  <div class="c bl"><div class="cl">Presupuesto Disponible</div><div class="cv">{fmt(presDisp)}</div></div>
  <div class="c {'gn' if gm_op>=0 else 'rd'}"><div class="cl">GM Operativo</div><div class="cv">{fmt(gm_op)}</div><div class="cs">{'▲' if gm_op>=0 else '▼'} {gm_op_pct:.1f}%</div></div>
</div>

<div class="st">💰 Resumen de Gastos</div>
<table><thead><tr><th>Concepto</th><th style="text-align:right">Monto USD</th><th style="text-align:right">% Presupuesto</th></tr></thead>
<tbody>
  <tr><td>🕐 Mano de Obra</td><td style="text-align:right;font-family:monospace">{fmt(wh_cost)}</td><td style="text-align:right">{pct(wh_cost)}</td></tr>
  <tr><td>🛒 Compras (OCs)</td><td style="text-align:right;font-family:monospace">{fmt(pur_total)}</td><td style="text-align:right">{pct(pur_total)}</td></tr>
  {svc_rows}
  <tr style="font-weight:700;background:#f0f0f0"><td>TOTAL COSTO</td><td style="text-align:right;font-family:monospace">{fmt(d.get('cost',0))}</td><td style="text-align:right">{pct(d.get('cost',0))}</td></tr>
</tbody></table>

<div class="st">📋 Órdenes de Compra ({len(d["po_items"])})</div>
<table><thead><tr><th>No. PO</th><th>Proveedor</th><th style="text-align:right">Subtotal USD</th></tr></thead>
<tbody>{po_rows or "<tr><td colspan='3' style='padding:10px;text-align:center;color:#888'>Sin órdenes de compra registradas</td></tr>"}</tbody>
{'<tfoot><tr style="font-weight:700"><td colspan="2" style="padding:6px 8px">TOTAL POs</td><td style="padding:6px 8px;text-align:right;font-family:monospace">'+fmt(sum(r.get('subtotal_usd',r.get('subtotal',0)) for r in d["po_items"]))+'</td></tr></tfoot>' if d["po_items"] else ''}
</table>

{targ_html}

{'<div class="st">📅 Diagrama de Tiempos (Gantt)</div><div class="gantt">'+gantt_svg+'</div>' if gantt_svg else '<div class="st">📅 Gantt</div><p style="color:#aaa;font-size:10px;padding:8px 0">Sin configuración de timing para este proyecto.</p>'}

<div class="footer">
  <span>Persico Suite · Reporte Ejecutivo · {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}</span>
  <span>Job {job_number} · {d.get('customer','—')}</span>
</div>
<script>window.onload=()=>window.print();</script>
</body></html>"""
        return Response(html, mimetype="text/html",
            headers={"Content-Disposition":f"inline;filename=ejecutivo_{job_number}.html"})
    except Exception as e:
        return jsonify({"error":str(e)}), 500



# ══════════════════════════════════════════════════════════════════
#  MOVIMIENTO APARTADOS → STOCK (Recuperaciones)
# ══════════════════════════════════════════════════════════════════
MOVIMIENTO_FILE = _os.path.join(_DATA, "movimientos_stock.json")

def movimiento_load():
    p = Path(MOVIMIENTO_FILE)
    if p.exists():
        try:
            with open(p,"r",encoding="utf-8") as f: return json.load(f)
        except: return []
    return []

def movimiento_save(records):
    Path(MOVIMIENTO_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(MOVIMIENTO_FILE,"w",encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

@app.route("/api/movimiento-stock-list", methods=["GET"])
def api_get_movimientos():
    try:
        return jsonify({"records": movimiento_load(), "total": len(movimiento_load())})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/movimiento-stock", methods=["POST"])
def api_mover_a_stock():
    """Move items from Apartados → Stock and create recovery records."""
    try:
        data       = request.get_json()
        job        = str(data.get("job","")).strip().upper()
        solicitante= str(data.get("solicitante","")).strip()
        items      = data.get("items",[])
        if not job:   return jsonify({"error":"Job requerido"}), 400
        if not items: return jsonify({"error":"Selecciona al menos un item"}), 400
        with lock:
            now = datetime.datetime.now().isoformat()
            mov_items = []
            for it in items:
                qty = float(it.get("quantity",0))
                if qty <= 0: continue
                uc = float(it.get("unit_cost",0) or 0)
                mov_items.append({
                    "part_number": str(it.get("part_number","")).strip().upper(),
                    "brand":       str(it.get("brand","")).strip().upper(),
                    "description": str(it.get("description","")).strip(),
                    "cat_code":    str(it.get("cat_code","")).strip().upper(),
                    "label_code":  str(it.get("label_code","")).strip().upper(),
                    "quantity":    qty, "unit_cost": uc,
                    "total":       round(qty*uc,2),
                })
            if not mov_items: return jsonify({"error":"Ningún item con cantidad > 0"}), 400

            # Descontar de Apartados
            apartados = apartado_load()
            for it in mov_items:
                pnum=it["part_number"].upper()
                apt=next((a for a in apartados if a.get("part_number","").upper()==pnum),None)
                if apt:
                    je=next((j for j in apt.get("jobs",[]) if j.get("job","").upper()==job),None)
                    if je:
                        je["quantity"]=max(0,float(je.get("quantity",0))-it["quantity"])
                        if je["quantity"]<=0:
                            apt["jobs"]=[j for j in apt.get("jobs",[]) if j.get("job","").upper()!=job]
                    apt["total_quantity"]=max(0,sum(float(j.get("quantity",0)) for j in apt.get("jobs",[])))
                    apt["updated_at"]=now
            apartado_save(apartados)

            # Incrementar Stock
            stock_records = stock_load()
            for it in mov_items:
                existing=next((r for r in stock_records
                    if r.get("part_number","").upper()==it["part_number"].upper() and
                       r.get("manufacturer","").upper()==it["brand"].upper()),None)
                if existing:
                    existing["quantity"]=float(existing.get("quantity",0))+it["quantity"]
                    if it["unit_cost"]>0: existing["last_cost"]=it["unit_cost"]
                    existing["updated_at"]=now
                else:
                    stock_records.append({
                        "id":f"STK-MOV-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                        "manufacturer":it["brand"],"part_number":it["part_number"],
                        "description":it["description"],"last_cost":it["unit_cost"],
                        "quantity":it["quantity"],"unit":"Pieza","section":"","box":"",
                        "label_code":it["label_code"],"recovery_job":job,"created_at":now,
                    })
            stock_save(stock_records)

            # Registrar movimiento
            rec = {
                "id": _doc_next_number("SM"),
                "job":job,"solicitante":solicitante,"items":mov_items,
                "fecha":now[:10],"created_at":now,
            }
            movimientos=movimiento_load(); movimientos.append(rec); movimiento_save(movimientos)

            # Crear registros de Recuperación
            recoveries=recovery_load()
            for it in mov_items:
                recoveries.append({
                    "id":f"RCV-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                    "manufacturer":it["brand"],"part_number":it["part_number"],
                    "description":it["description"],"last_cost":it["unit_cost"],
                    "quantity":it["quantity"],"unit":"Pieza","section":"","box":"",
                    "label_code":it["label_code"],"job":job,
                    "total_value":round(it["quantity"]*it["unit_cost"],2),
                    "stock_id":"","movimiento_id":rec["id"],"created_at":now,
                })
            recovery_save(recoveries)

        return jsonify({"ok":True,"record":rec,"items_moved":len(mov_items)})
    except Exception as e:
        return jsonify({"error":str(e)}),500

@app.route("/api/movimiento-stock/<mov_id>/pdf")
def api_movimiento_pdf(mov_id):
    try:
        records=movimiento_load()
        rec=next((r for r in records if r["id"]==mov_id),None)
        if not rec: return jsonify({"error":"Movimiento no encontrado"}),404
        logo_path=_os.path.join(_BASE,"static","persico_logo.webp"); logo_b64=""
        if _os.path.exists(logo_path):
            import base64 as _b64
            with open(logo_path,"rb") as lf: logo_b64=_b64.b64encode(lf.read()).decode()
        logo_tag=f'<img src="data:image/webp;base64,{logo_b64}" style="height:48px">' if logo_b64 else '<b style="color:#c8102e">PERSICO</b>'
        fmt=lambda v:f"${float(v):,.2f}"
        items=rec.get("items",[]); subtotal=sum(it.get("total",0) for it in items)
        rows="".join(f"<tr><td style='text-align:center'>{i+1}</td><td style='font-family:monospace'>{it.get('cat_code','—')}</td><td style='font-family:monospace'>{it.get('part_number','—')}</td><td>{it.get('brand','—')}</td><td>{it.get('description','—')}</td><td style='text-align:right'>{int(it.get('quantity',0))}</td><td style='text-align:right'>{fmt(it.get('unit_cost',0))}</td><td style='text-align:right;font-weight:600'>{fmt(it.get('total',0))}</td></tr>" for i,it in enumerate(items))
        html=f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>{rec['id']}</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:Arial,font-size:11px;padding:28px}}
table{{width:100%;border-collapse:collapse;margin-bottom:12px}}
thead th{{background:#1f3864;color:#fff;padding:7px 8px;font-size:9px;text-transform:uppercase;text-align:left}}
tbody tr:nth-child(even){{background:#f7f7f7}}td{{padding:6px 8px;border-bottom:1px solid #eee}}</style></head>
<body>
<div style="display:flex;justify-content:space-between;border-bottom:3px solid #f5a623;padding-bottom:12px;margin-bottom:16px">
  <div>{logo_tag}</div>
  <div style="text-align:right"><div style="font-size:10px;text-transform:uppercase;color:#888">Movimiento a Stock</div>
  <div style="font-size:18px;font-weight:900;color:#f5a623">{rec['id']}</div>
  <div style="font-size:10px;color:#888">{rec.get('fecha','')}</div></div>
</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px;padding:10px;background:#f7f7f7;border-radius:6px;font-size:10px">
  <div><b>Job Origen:</b> {rec.get('job','—')}</div><div><b>Solicitante:</b> {rec.get('solicitante','—')}</div>
</div>
<table><thead><tr><th>#</th><th>Código</th><th>No. Parte</th><th>Marca</th><th>Descripción</th><th style="text-align:right">Cant.</th><th style="text-align:right">Costo Unit.</th><th style="text-align:right">Total</th></tr></thead>
<tbody>{rows}</tbody></table>
<div style="display:flex;justify-content:flex-end;gap:20px;padding:10px 14px;background:#1f3864;color:#fff;border-radius:6px;font-weight:700;margin-bottom:28px"><span>TOTAL USD</span><span>{fmt(subtotal)}</span></div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:40px;margin-top:20px">
  <div style="border-top:1px solid #aaa;margin-top:50px;padding-top:6px;font-size:10px;text-align:center"><b>{rec.get('solicitante','—')}</b><br>Solicitante / Autoriza</div>
  <div style="border-top:1px solid #aaa;margin-top:50px;padding-top:6px;font-size:10px;text-align:center">____________________<br>Almacenista / Recibe</div>
</div>
</body></html>"""
        return Response(html,mimetype="text/html",headers={"Content-Disposition":f"inline;filename={rec['id']}.html"})
    except Exception as e: return jsonify({"error":str(e)}),500

# ══════════════════════════════════════════════════════════════════
#  RESPALDO COMPLETO — DESCARGA DIRECTA
# ══════════════════════════════════════════════════════════════════
    """Build an in-memory ZIP with all data files as Excel sheets."""
    import io, zipfile, openpyxl
    buf = io.BytesIO()
    now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")

    def json_to_sheet(wb, data_list, sheet_name):
        if not data_list: return
        ws = wb.create_sheet(sheet_name[:31])
        if not isinstance(data_list[0], dict): return
        keys = list(data_list[0].keys())
        for ci, k in enumerate(keys, 1):
            ws.cell(1, ci, k)
        for ri, row in enumerate(data_list, 2):
            for ci, k in enumerate(keys, 1):
                val = row.get(k)
                if isinstance(val, (dict, list)):
                    val = json.dumps(val, ensure_ascii=False)
                ws.cell(ri, ci, val)

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # --- Simple JSON tables → one xlsx per category
        simple_files = {
            "jobs":             _os.path.join(_DATA, "JOBs", "jobs.json"),
            "quotes":           _os.path.join(_DATA, "QUOTE_REG", "quotes.json"),
            "proveedores":      _os.path.join(_DATA, "proveedores.json"),
            "catalogo_electrico":   _os.path.join(_DATA, "catalogo_electrico.json"),
            "catalogo_mecanico":    _os.path.join(_DATA, "catalogo_mecanico.json"),
            "catalogo_servicios":   _os.path.join(_DATA, "catalogo_servicios.json"),
            "stock":            _os.path.join(_DATA, "stock.json"),
            "apartados":        _os.path.join(_DATA, "apartados.json"),
            "salidas":          _os.path.join(_DATA, "salidas.json"),
            "ingresos":         _os.path.join(_DATA, "ingresos.json"),
            "reassign_orders":  _os.path.join(_DATA, "reassign_orders.json"),
            "recovery":         _os.path.join(_DATA, "recovery.json"),
            "viaticos":         _os.path.join(_DATA, "viaticos.json"),
            "gastos_viaje":     _os.path.join(_DATA, "gastos_viaje.json"),
            "envios":           _os.path.join(_DATA, "envios.json"),
            "project_configs":  _os.path.join(_DATA, "project_configs.json"),
            "generated_pos":    _os.path.join(_DATA, "generated_pos.json"),
            "movimientos_stock":_os.path.join(_DATA, "movimientos_stock.json"),
        }
        for name, path in simple_files.items():
            p = Path(path)
            if not p.exists(): continue
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                if not isinstance(data, list) or not data: continue
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
                json_to_sheet(wb, data, name)
                sheet_buf = io.BytesIO(); wb.save(sheet_buf)
                zf.writestr(f"{name}.xlsx", sheet_buf.getvalue())
            except: pass

        # --- Year-based files
        year_folders = {
            "CPO":  (_os.path.join(_DATA, "CPOs"),    "cpo_{year}.json"),
            "IVP":  (_os.path.join(_DATA, "IVPs"),    "ivp_{year}.json"),
            "WH":   (_os.path.join(_DATA, "WHs"),     "wh_{year}.json"),
            "FX":   (_os.path.join(_DATA, "FX"),      "fx_{year}.json"),
            "IPO":  (_os.path.join(_DATA, "IPOs"),    "po_{year}.json"),
        }
        for prefix, (folder, pattern) in year_folders.items():
            for yr in range(CURRENT_YEAR-3, CURRENT_YEAR+2):
                fp = _os.path.join(folder, pattern.format(year=yr))
                if not _os.path.exists(fp): continue
                try:
                    data = json.loads(Path(fp).read_text(encoding="utf-8"))
                    if not data: continue
                    # Dict → list of rows (FX stores {date:rate})
                    if isinstance(data, dict):
                        data = [{"fecha": k, "tasa": v} for k,v in data.items()]
                    if not isinstance(data, list) or not data: continue
                    wb = openpyxl.Workbook(); wb.remove(wb.active)
                    json_to_sheet(wb, data, f"{prefix}_{yr}"[:31])
                    sheet_buf = io.BytesIO(); wb.save(sheet_buf)
                    zf.writestr(f"{prefix}_{yr}.xlsx", sheet_buf.getvalue())
                except: pass

        # --- PT and SV numbers (folder per number)
        for ptsv_type in ["PT_NUMBERS","SV_NUMBERS"]:
            folder = _os.path.join(_DATA, ptsv_type)
            if not _os.path.exists(folder): continue
            try:
                all_records = []
                for fp in Path(folder).glob("*.json"):
                    try:
                        rec = json.loads(fp.read_text(encoding="utf-8"))
                        if isinstance(rec, list): all_records.extend(rec)
                        elif isinstance(rec, dict): all_records.append(rec)
                    except: pass
                if all_records:
                    wb = openpyxl.Workbook(); wb.remove(wb.active)
                    json_to_sheet(wb, all_records, ptsv_type[:31])
                    sheet_buf = io.BytesIO(); wb.save(sheet_buf)
                    zf.writestr(f"{ptsv_type.lower()}.xlsx", sheet_buf.getvalue())
            except: pass

        # Also include raw JSON of all data files for full restore
        data_root = Path(_DATA)
        for p in data_root.rglob("*.json"):
            if "backup_config" in p.name: continue
            if "users_auth" in p.name: continue  # skip passwords
            try:
                rel = str(p.relative_to(data_root))
                zf.write(str(p), f"_raw_json/{rel}")
            except: pass

    buf.seek(0)
    return buf.getvalue(), f"persico_backup_{now_str}.zip"

def _build_backup_zip():
    """Build an in-memory ZIP with all data files as Excel sheets."""
    import io, zipfile, openpyxl
    buf = io.BytesIO()
    now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")

    def json_to_sheet(wb, data_list, sheet_name):
        if not data_list: return
        ws = wb.create_sheet(sheet_name[:31])
        if not isinstance(data_list[0], dict): return
        keys = list(data_list[0].keys())
        for ci, k in enumerate(keys, 1): ws.cell(1, ci, k)
        for ri, row in enumerate(data_list, 2):
            for ci, k in enumerate(keys, 1):
                val = row.get(k)
                if isinstance(val, (dict, list)): val = json.dumps(val, ensure_ascii=False)
                ws.cell(ri, ci, val)

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        simple_files = {
            "jobs":              _os.path.join(_DATA, "JOBs", "jobs.json"),
            "quotes":            _os.path.join(_DATA, "QUOTE_REG", "quotes.json"),
            "proveedores":       _os.path.join(_DATA, "proveedores.json"),
            "catalogo_electrico":_os.path.join(_DATA, "catalogo_electrico.json"),
            "catalogo_mecanico": _os.path.join(_DATA, "catalogo_mecanico.json"),
            "catalogo_servicios":_os.path.join(_DATA, "catalogo_servicios.json"),
            "stock":             _os.path.join(_DATA, "stock.json"),
            "apartados":         _os.path.join(_DATA, "apartados.json"),
            "salidas":           _os.path.join(_DATA, "salidas.json"),
            "ingresos":          _os.path.join(_DATA, "ingresos.json"),
            "reassign_orders":   _os.path.join(_DATA, "reassign_orders.json"),
            "recovery":          _os.path.join(_DATA, "recovery.json"),
            "viaticos":          _os.path.join(_DATA, "viaticos.json"),
            "gastos_viaje":      _os.path.join(_DATA, "gastos_viaje.json"),
            "envios":            _os.path.join(_DATA, "envios.json"),
            "project_configs":   _os.path.join(_DATA, "project_configs.json"),
            "generated_pos":     _os.path.join(_DATA, "generated_pos.json"),
            "movimientos_stock": _os.path.join(_DATA, "movimientos_stock.json"),
        }
        for name, path in simple_files.items():
            p = Path(path)
            if not p.exists(): continue
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                if not isinstance(data, list) or not data: continue
                wb = openpyxl.Workbook(); wb.remove(wb.active)
                json_to_sheet(wb, data, name)
                sb = io.BytesIO(); wb.save(sb)
                zf.writestr(f"{name}.xlsx", sb.getvalue())
            except: pass

        for prefix, (folder, pattern) in {
            "CPO":(_os.path.join(_DATA,"CPOs"),"cpo_{y}.json"),
            "IVP":(_os.path.join(_DATA,"IVPs"),"ivp_{y}.json"),
            "WH": (_os.path.join(_DATA,"WHs"), "wh_{y}.json"),
            "FX": (_os.path.join(_DATA,"FX"),  "fx_{y}.json"),
            "IPO":(_os.path.join(_DATA,"IPOs"),"po_{y}.json"),
        }.items():
            for yr in range(CURRENT_YEAR-3, CURRENT_YEAR+2):
                fp = _os.path.join(folder, pattern.format(y=yr))
                if not _os.path.exists(fp): continue
                try:
                    data = json.loads(Path(fp).read_text(encoding="utf-8"))
                    if not data: continue
                    if isinstance(data, dict):
                        data = [{"fecha":k,"tasa":v} for k,v in data.items()]
                    if not isinstance(data, list) or not data: continue
                    wb = openpyxl.Workbook(); wb.remove(wb.active)
                    json_to_sheet(wb, data, f"{prefix}_{yr}"[:31])
                    sb = io.BytesIO(); wb.save(sb)
                    zf.writestr(f"{prefix}_{yr}.xlsx", sb.getvalue())
                except: pass

        # Raw JSON for full restore
        for p in Path(_DATA).rglob("*.json"):
            if "users_auth" in p.name: continue
            try:
                rel = str(p.relative_to(Path(_DATA)))
                zf.write(str(p), f"_raw/{rel}")
            except: pass

    buf.seek(0)
    return buf.getvalue(), f"persico_backup_{now_str}.zip"
@app.route("/api/admin/backup", methods=["POST"])
def api_admin_backup():
    """Generate backup ZIP and return as download."""
    if not is_admin(): return jsonify({"error":"Sin permiso"}), 403
    try:
        zip_bytes, zip_name = _build_backup_zip()
        import base64 as _b64
        return jsonify({
            "ok": True,
            "zip_b64": _b64.b64encode(zip_bytes).decode(),
            "filename": zip_name,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route("/api/util/parse-xl-po", methods=["POST"])
def api_parse_xl_po():
    """Parse a PO Excel file (SAE format) and return structured rows."""
    try:
        import io, openpyxl
        f = request.files.get("file")
        if not f: return jsonify({"error":"No se recibió archivo"}), 400
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        rows_data = list(ws.iter_rows(values_only=True))
        if len(rows_data) < 2: return jsonify({"rows":[]})

        # Auto-detect column positions from header row
        hdr = [str(c or '').lower().strip() for c in rows_data[0]]
        def ci(*keywords):
            for kw in keywords:
                for i,h in enumerate(hdr):
                    if kw in h: return i
            return None

        ci_pnum  = ci('part','número','numero','clave')
        ci_desc  = ci('descrip')
        ci_brand = ci('marca','fabricante','manufacturer')
        ci_qty   = ci('cantid','cantidad','qty','quantity')
        ci_cost  = ci('costo','precio','unit','price')
        ci_job   = ci('job')

        def cv(row, idx):
            if idx is None or idx >= len(row): return ''
            v = row[idx]
            return '' if v is None else v

        result = []
        for row in rows_data[1:]:
            pnum = str(cv(row, ci_pnum) or '').strip().upper()
            if not pnum: continue
            qty  = float(cv(row, ci_qty)  or 0)
            cost_raw = str(cv(row, ci_cost) or '0').replace('$','').replace(',','')
            try: cost = float(cost_raw)
            except: cost = 0.0
            result.append({
                "part_number": pnum,
                "description": str(cv(row, ci_desc) or '').strip(),
                "brand":       str(cv(row, ci_brand) or '').strip().upper(),
                "quantity":    qty,
                "unit_cost":   cost,
                "job":         str(cv(row, ci_job) or '').strip().upper(),
            })
        return jsonify({"rows": result, "total": len(result)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route("/api/util/pdf-to-po-json", methods=["POST"])
def api_pdf_to_po_json():
    """Parse a PO PDF and return items as JSON to populate the form directly."""
    try:
        import io, re
        try:
            import pdfplumber
        except ImportError:
            return jsonify({"error": "pdfplumber no está instalado. Agrega 'pdfplumber' a requirements.txt y redespliega."}), 500

        f = request.files.get("file")
        if not f: return jsonify({"error":"No se recibió archivo"}), 400
        full_text = ""
        with pdfplumber.open(io.BytesIO(f.read())) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"

        def find_field(pattern, text, group=1, default=""):
            m = re.search(pattern, text, re.IGNORECASE)
            return m.group(group).strip() if m else default

        oc_num   = find_field(r"OC\s*N[°o]?\s*([\d]+)", full_text)
        job_raw  = ""
        if not job_raw:
            m_job = re.search(r'(\d{3,}-\d{2,}(?:\s+[A-Z]{1,3})?)\s*$', full_text, re.MULTILINE)
            job_raw = m_job.group(1) if m_job else ""
        job_clean = re.sub(r'\s+[A-Z]{1,3}$', '', job_raw.strip()).strip()
        supplier  = find_field(r"PROVEEDOR\s*\(\s*\d+\s*\)\s*\n?([\w\s\.,]+?)(?:\n|RFC|AVENIDA|$)", full_text)
        brand     = supplier.split()[0] if supplier else ""

        items = []
        lines = full_text.split('\n')
        in_items = False
        item_buffer = {}
        desc_extra  = []

        for line in lines:
            line = line.strip()
            if re.match(r'PARTIDA\s+CANTIDAD', line, re.IGNORECASE):
                in_items = True; continue
            if re.match(r'SUBTOTAL|DESCUENTO|TOTAL\b|IVA\b', line, re.IGNORECASE) and in_items:
                if item_buffer:
                    if desc_extra: item_buffer["description"] += " " + " ".join(desc_extra)
                    items.append(item_buffer)
                    item_buffer = {}; desc_extra = []
                break
            if not in_items: continue
            m = re.match(r'^(\d+)\s+([\d,\.]+)\s+(.+?)\s+([\d,\.]+)\s+([\d,\.]+)\s*$', line)
            if m:
                if item_buffer:
                    if desc_extra: item_buffer["description"] += " " + " ".join(desc_extra)
                    items.append(item_buffer); desc_extra = []
                qty    = float(m.group(2).replace(',', ''))
                cost   = float(m.group(4).replace(',', ''))
                middle = m.group(3).strip()
                tokens = middle.split(None, 1)
                first  = tokens[0] if tokens else ""
                if first and (re.search(r'[\d\-]', first) or first.isupper()):
                    pnum = first
                    desc = tokens[1].strip() if len(tokens) > 1 else ""
                else:
                    pnum = ""; desc = middle
                item_buffer = {"part_number":pnum,"description":desc,"brand":brand,
                               "quantity":qty,"unit_cost":cost,"job":job_clean}
                continue
            if item_buffer:
                dibujo = re.search(r'DIBUJO:\s*([\w\s\d\-]+?)(?:\*|$)', line, re.IGNORECASE)
                if dibujo and not item_buffer["part_number"]:
                    item_buffer["part_number"] = dibujo.group(1).strip().replace(' ','-')
                elif line and not re.match(r'\*{2,}', line):
                    desc_extra.append(line)

        if item_buffer:
            if desc_extra: item_buffer["description"] += " " + " ".join(desc_extra)
            items.append(item_buffer)

        if not items:
            return jsonify({"error":"No se encontraron items en el PDF."}), 400
        return jsonify({"ok":True,"oc_num":oc_num,"job":job_clean,
                        "supplier":supplier,"items":items,"total":len(items)})
    except Exception as e:
        return jsonify({"error":str(e)}), 500


def api_pdf_to_po_excel():
    """
    Parse a PO PDF (Persico SAE format) and return a normalized Excel file
    with columns: Número de parte | Descripción | Marca | Cantidad | Costo unitario (USD) | JOB
    """
    try:
        import io, re, pdfplumber, openpyxl

        f = request.files.get("file")
        if not f: return jsonify({"error":"No se recibió archivo"}), 400

        # ── Extract text from PDF
        full_text = ""
        with pdfplumber.open(io.BytesIO(f.read())) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"

        # ── Extract header fields
        def find_field(pattern, text, group=1, default=""):
            m = re.search(pattern, text, re.IGNORECASE)
            return m.group(group).strip() if m else default

        oc_num    = find_field(r"OC\s*N[°o]?\s*([\d]+)", full_text)
        job_raw   = find_field(r"JOB\s*NUMBER\s*\n?([\w\s\-]+?)(?:\n|$)", full_text)
        # Clean job: take only the code part (e.g. "665-02" from "665-02 MF")
        # JOB appears on the line after "FECHA ENTREGA JOB NUMBER" — extract from that line
        job_raw = find_field(
            r'(?:FECHA\s+ENTREGA\s+JOB\s+NUMBER\s*\n.*?([\w\-]+(?:\s+[A-Z]{1,3})?)\s*$'
            r'|JOB\s*NUMBER\s*\n?\s*\S.*?([\w\-]+(?:\s+[A-Z]{1,3})?)\s*$)',
            full_text, group=1)
        # Fallback: find a pattern like 665-01 optionally followed by 2-letter suffix
        if not job_raw:
            m_job = re.search(r'(\d{3,}-\d{2,}(?:\s+[A-Z]{1,3})?)\s*$',
                              full_text, re.MULTILINE)
            job_raw = m_job.group(1) if m_job else ""
        job_clean = re.sub(r'\s+[A-Z]{1,3}$', '', job_raw.strip()).strip()
        supplier  = find_field(r"PROVEEDOR\s*\(\s*\d+\s*\)\s*\n?([\w\s\.,]+?)(?:\n|RFC|AVENIDA|$)", full_text)

        # ── Parse line items
        items = []
        lines = full_text.split('\n')
        in_items = False
        item_buffer = {}
        desc_extra  = []

        for line in lines:
            line = line.strip()
            if re.match(r'PARTIDA\s+CANTIDAD', line, re.IGNORECASE):
                in_items = True; continue
            if re.match(r'SUBTOTAL|DESCUENTO|TOTAL\b|IVA\b', line, re.IGNORECASE) and in_items:
                if item_buffer:
                    if desc_extra: item_buffer["description"] += " " + " ".join(desc_extra)
                    items.append(item_buffer)
                    item_buffer = {}; desc_extra = []
                break
            if not in_items: continue

            m = re.match(r'^(\d+)\s+([\d,\.]+)\s+(.+?)\s+([\d,\.]+)\s+([\d,\.]+)\s*$', line)
            if m:
                if item_buffer:
                    if desc_extra: item_buffer["description"] += " " + " ".join(desc_extra)
                    items.append(item_buffer)
                    desc_extra = []
                qty    = float(m.group(2).replace(',', ''))
                cost   = float(m.group(4).replace(',', ''))
                middle = m.group(3).strip()

                # Split middle: first token = part number, rest = description
                # Valid part number: has digit/hyphen OR is all-uppercase (e.g. SERVICE)
                tokens = middle.split(None, 1)
                first  = tokens[0] if tokens else ""
                if first and (re.search(r'[\d\-]', first) or first.isupper()):
                    pnum = first
                    desc = tokens[1].strip() if len(tokens) > 1 else ""
                else:
                    pnum = ""
                    desc = middle

                item_buffer = {
                    "part_number": pnum,
                    "description": desc,
                    "brand":       supplier.split()[0] if supplier else "",
                    "quantity":    qty,
                    "unit_cost":   cost,
                    "job":         job_clean,
                }
                continue

            if item_buffer:
                dibujo = re.search(r'DIBUJO:\s*([\w\s\d\-]+?)(?:\*|$)', line, re.IGNORECASE)
                if dibujo and not item_buffer["part_number"]:
                    item_buffer["part_number"] = dibujo.group(1).strip().replace(' ', '-')
                elif line and not re.match(r'\*{2,}', line):
                    desc_extra.append(line)

        if item_buffer:
            if desc_extra: item_buffer["description"] += " " + " ".join(desc_extra)
            items.append(item_buffer)

        if not items:
            return jsonify({"error":"No se encontraron items en el PDF. Verifica que sea una OC en formato SAE."}), 400

        # ── Build Excel in standard format
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Items PO"

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill("solid", fgColor="1A1A2E")
        thin     = Side(style='thin', color="CCCCCC")
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)
        data_fnt = Font(name="Calibri", size=10)
        gold_fnt = Font(name="Calibri", size=10, bold=True, color="FFB300")

        headers    = ["Número de parte","Descripción","Marca","Cantidad","Costo unitario (USD)","JOB"]
        col_widths = [20, 52, 16, 12, 22, 14]

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 26

        alt = PatternFill("solid", fgColor="F9F9F9")
        wht = PatternFill("solid", fgColor="FFFFFF")
        for ri, it in enumerate(items, 2):
            fill = alt if ri % 2 == 0 else wht
            vals = [it["part_number"], it["description"], it["brand"],
                    it["quantity"], it["unit_cost"], it["job"]]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = border; cell.fill = fill
                cell.alignment = Alignment(vertical="center", wrap_text=(ci==2))
                if ci == 4:
                    cell.font = data_fnt; cell.alignment = Alignment(horizontal="right")
                elif ci == 5:
                    cell.font = gold_fnt; cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif ci == 1:
                    cell.font = Font(name="Courier New", size=10, bold=True)
                else:
                    cell.font = data_fnt
            ws.row_dimensions[ri].height = 22

        # Save to buffer
        out = io.BytesIO(); wb.save(out); out.seek(0)
        fname = f"PO_{oc_num or 'import'}_normalizada.xlsx"
        return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=fname)
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route("/api/gpo/search", methods=["GET"])
def api_gpo_search():
    """Find a GPO or IPO record by partial number (suffix match)."""
    try:
        q = str(request.args.get("q","")).strip()
        if not q: return jsonify({"error":"Parámetro q requerido"}), 400
        q_clean = q.lstrip("PO-").lstrip("0") or q

        # Search GPO records first
        for r in gpo_load():
            pn = str(r.get("po_number",""))
            if pn.lstrip("PO-").lstrip("0") == q_clean or pn.endswith(q_clean):
                return jsonify({"record": r, "source": "gpo"})

        # Search IPO records (po_YYYY.json) — these are what shows in the PO list
        for yr in range(CURRENT_YEAR - 2, CURRENT_YEAR + 2):
            for r in po_load(yr):
                clave = str(r.get("clave","") or r.get("gpo_number",""))
                if clave.lstrip("PO-").lstrip("0") == q_clean or clave.endswith(q_clean):
                    # Synthesize a GPO-like record from the IPO data for the modal
                    return jsonify({"record": {
                        "po_number":    clave,
                        "supplier_name":r.get("nombre",""),
                        "job":          r.get("entregar_a",""),
                        "moneda":       r.get("moneda","USD"),
                        "fx_rate":      r.get("tipo_cambio",1),
                        "total":        r.get("subtotal",0),
                        "created_at":   r.get("fecha_doc",""),
                        "items":        [],
                        "_from_ipo":    True,
                        "_ipo_year":    yr,
                    }, "source": "ipo"})

        return jsonify({"error": f"No se encontró la orden {q}"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/gpo/<po_number>/split", methods=["POST"])
def api_gpo_split(po_number):
    """
    Divide a PO SAE into individual IPO records (one per item).
    Zeroes the original PO to avoid double counting.
    Each split item gets a unique id so it can be individually managed.
    """
    try:
        data  = request.get_json()
        items = data.get("items", [])
        if not items:
            return jsonify({"error": "No hay items para dividir"}), 400
        po_upper = po_number.upper()
        with lock:
            gpos    = gpo_load()
            gpo_rec = next((r for r in gpos if r.get("po_number","").upper()==po_upper), None)

            # ── Find the correct IPO year (double-break pattern)
            ipo_yr     = CURRENT_YEAR
            po_digits  = ''.join(filter(str.isdigit, po_upper))
            year_found = False
            for search_yr in range(CURRENT_YEAR + 1, CURRENT_YEAR - 4, -1):
                recs = po_load(search_yr)
                for r in recs:
                    clave_r = str(r.get("clave","") or r.get("gpo_number",""))
                    if ''.join(filter(str.isdigit, clave_r)) == po_digits:
                        ipo_yr = search_yr
                        year_found = True
                        break
                if year_found:
                    break

            moneda    = str(data.get("moneda","USD")).upper()
            fx_rate   = float(data.get("tipo_cambio",
                         (gpo_rec or {}).get("fx_rate", 1)) or 1)
            if moneda == "MXN" and fx_rate <= 1:
                fx_rate = 1.0
            supplier  = (gpo_rec or {}).get("supplier_name",
                         data.get("supplier_name",""))
            now       = datetime.datetime.now().isoformat()
            fecha_doc = (gpo_rec or {}).get("created_at","")[:10] or now[:10]

            # ── Load IPO records
            ipo_records = po_load(ipo_yr)
            orig_count  = len(ipo_records)

            # ── Step 1: Zero originals (records that match clave AND are NOT split)
            for r in ipo_records:
                clave_r = str(r.get("clave","") or r.get("gpo_number",""))
                if (''.join(filter(str.isdigit, clave_r)) == po_digits
                        and not r.get("_split_origin")):
                    r["subtotal"]      = 0.0
                    r["subtotal_mxn"]  = 0.0
                    r["estatus"]       = "Dividida"
                    r["_split_zeroed"] = True

            # ── Step 2: Append one record per item
            new_entries = []
            ts = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            for i, it in enumerate(items):
                qty       = float(it.get("quantity", 0) or 0)
                price     = float(it.get("unit_cost",
                             it.get("unit_price", 0)) or 0)
                total     = round(qty * price, 2)
                total_usd = round(total / fx_rate, 2) if moneda == "MXN" else total
                job       = str(it.get("job","")).strip().upper()
                entry = {
                    "_id":                  f"SPLIT-{po_digits}-{i:03d}-{ts}",
                    "clave":                po_upper,
                    "fecha_doc":            fecha_doc,
                    "entregar_a":           job,
                    "nombre":               supplier,
                    "subtotal":             total,
                    "moneda":               moneda,
                    "tipo_cambio":          fx_rate,
                    "subtotal_mxn":         total if moneda=="MXN" else round(total*fx_rate,2),
                    "subtotal_usd":         total_usd,
                    "estatus":              "Emitida",
                    "descuento_financiero": 0,
                    "pct_descuento":        0,
                    "fecha_recepcion":      "",
                    "gpo_number":           po_upper,
                    "gpo_pdf":              i == 0,
                    "part_number":          str(it.get("part_number","")).strip().upper(),
                    "description":          str(it.get("description","")).strip(),
                    "quantity":             qty,
                    "unit_price":           price,
                    "_split_origin":        True,
                    "created_at":           now,
                }
                ipo_records.append(entry)
                new_entries.append(entry)

            # ── Step 3: Save everything once
            po_save(ipo_yr, ipo_records)

            # ── Verify save
            verify = po_load(ipo_yr)
            saved_split = sum(1 for r in verify if r.get("_split_origin") and
                              ''.join(filter(str.isdigit,
                                str(r.get("clave","") or ""))) == po_digits)

            # ── Mark GPO
            if gpo_rec:
                gpo_rec["modificacion_tipo"] = "Dividida"
                gpo_rec["effective_total"]   = 0.0
                gpo_rec["modificacion_at"]   = now
                gpo_rec["modificacion_by"]   = session.get("user","")
                gpo_save(gpos)

        return jsonify({
            "ok":      True,
            "created": len(new_entries),
            "saved":   saved_split,
            "year":    ipo_yr,
            "po_number": po_upper,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500



if __name__ == "__main__":
    print("=" * 60)
    print("  Persico Mex — Suite Unificada")
    print(f"  Job Register    : {JOBS_FOLDER}")
    print(f"  Hourly Rates    : {RATES_FOLDER}")
    print(f"  Quote Register  : {QUOTE_BASE}/quotes.json")
    print(f"  Purchase Orders : {PO_FOLDER}")
    print(f"  Customer POs    : {CPO_FOLDER}")
    print(f"  Work Hours      : {WH_FOLDER}")
    print(f"  Invoiced POs    : {IVP_FOLDER}")
    print(f"  FX / Tipo cambio: {FX_FOLDER}")
    print(f"  URL             : http://localhost:{PORT}")
    print("=" * 60)
    app.run(host=HOST, port=PORT, debug=False)
